# -*- coding: utf-8 -*-
"""
Created on Sun May 22 10:30:01 2016
SC process signups functions
@author: tkc
"""
#%%
import pandas as pd
import numpy as np
import datetime
import re

#%%

def findavailablekeys(df, colname, numkeys):
    '''Pass df and colname, return a defined number of available keys list
    used for players, families, signups, etc.
    '''
    # list comprehension    
    allnums=[i for i in range(1,len(df))]
    usedkeys=df[colname].unique()
    usedkeys=np.ndarray.tolist(usedkeys)
    availkeys=[i for i in allnums if i not in usedkeys]
    if len(availkeys)<numkeys: # get more keys starting at max+1
        needed=numkeys-len(availkeys)
        for i in range(0,needed):
            nextval=int(max(usedkeys)+1) # if no interior vals are available find next one
            availkeys.append(nextval+i)
    return availkeys

def dropcolumns(df1,df2):
    ''' Pass two dfs with df2 being the template.. extra unnecessary columns dropped from df1
    inplace=True modifies both passed and returned df  '''
    cols1=df1.columns.tolist()
    cols2=df2.columns.tolist()
    newdf=df1 # avoids modification of passed df
    uniquelist=[i for i in cols1 if i not in cols2]
    for i,colname in enumerate(uniquelist): # remove cols from df1 that are absent from df2
        # newdf.drop(colname, axis=1, inplace=True) # this modifies both passed and returned dfs
        newdf=newdf.drop(colname, axis=1)
    return newdf

def organizerecruits(df):
    ''' takes a sport-gender and organizes in manner for output to excel signup summary file ''' 
    mycols=['First', 'Last', 'Grade', 'Gender', 'Sport', 'School', 'Phone1', 'Text1','Email1', 'Phone2', 'Text2', 'Email2', 'Plakey', 'Famkey', 'Family'] 
    thisdf=pd.DataFrame(columns=mycols) # temp df for dropping unnecessary columns
    df=dropcolumns(df,thisdf) # drop columns not working
    df.Grade=df.Grade.replace('K',0)
    df=df.sort_values(['Grade'], ascending=True)
    df.Grade=df.Grade.replace(0,'K') # replace K with zero to allow sorting
    df=df[mycols] # put back in desired order
    return df

def organizecontacts(df):
    ''' takes a sport-gender and organizes in manner for output to excel signup summary file ''' 
    mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone', 'Text','Email', 'Phone2', 'Text2', 'Email2', 'Team', 'Plakey', 'Famkey', 'Family'] 
    thisdf=pd.DataFrame(columns=mycols) # temp df for dropping unnecessary columns
    df['Team']=''    
    df=dropcolumns(df,thisdf) # drop columns not working
    df.Grade=df.Grade.replace('K',0)
    df=df.sort_values(['Grade'], ascending=True)
    df.Grade=df.Grade.replace(0,'K') # replace K with zero to allow sorting
    df=df[mycols] # put back in desired order
    return df
    
def organizesignups(df, year):
    ''' takes SCsignup file subset (split by sport) and organizes for output into master signup file ''' 
    mycols=['SUkey','First', 'Last', 'Grade', 'Gender', 'Sport', 'Year', 'Team', 'Plakey','Famkey', 'Family'] 
    thisdf=pd.DataFrame(columns=mycols) # temp df for dropping unnecessary columns
    df=dropcolumns(df,thisdf) # drop columns not working
    df.Grade=df.Grade.replace('K',0)
    df=df.sort_values(['Gender','Grade'], ascending=True) # nested sort gender then grade
    df.Grade=df.Grade.replace(0,'K') # replace K with zero to allow sorting
    # add missing columns to df
    df['Team']=''
    df['Year']= int(year)
    df['SUkey']=0 # column for unique signup key (zero means not yet assigned)
    df=df[mycols] # put back in desired order
    return df

def organizeroster(df):
    ''' Renaming, reorg, delete unnecessary columns for CYC roster output
    already split by sport and year''' 
    df.rename(columns={'First':'Fname','Last':'Lname','Address':'Street','Parish_registration':'Parish of Registration'}, inplace=True)
    df.rename(columns={'Parish_residence':'Parish of Residence','Phone1':'Phone','DOB':'Birthdate','Gender':'Sex'}, inplace=True)
    df.rename(columns={'Email1':'Email'}, inplace=True)
    # replace Girl, Boy with m f
    df.Sex=df.Sex.replace('Girl','F')
    df.Sex=df.Sex.replace('Boy','M')
    mycols=['Fname', 'Lname', 'Street', 'City', 'State', 'Zip', 'Phone', 'Email', 'Birthdate', 'Sex', 'Role', 'Division', 'Grade', 'Team', 'School', 'Parish of Registration', 'Parish of Residence', 'Open/Closed','Coach ID']
    tempdf=pd.DataFrame(columns=mycols) # temp df for dropping unnecessary columns
    df=dropcolumns(df,tempdf) # drop unnecessary columns 
    df=df[mycols] # put back in desired order
    return df
   
def findyearseason(df):
    ''' Pass raw signups and determine year and sports season '''
    # get year from system clock and from google drive timestamp    
    now=datetime.datetime.now()
    val=df.Timestamp[0] # grab first timestamp    
    if val!=datetime.datetime: # if not a timestamp (i.e. manual string entry find one
        while type(val)!=datetime.datetime:
            for i in range(0,len(df)):
                val=df.Timestamp[i]
    year=val.year # use year value from signup timestamps
    if now.year!=val.year:
        print ('Possible year discrepancy: Signups are from ',str(val.year))
    # now find sports season
    mask = np.column_stack([df['Sport'].str.contains("occer", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Fall'
    mask = np.column_stack([df['Sport'].str.contains("rack", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Spring'
    mask = np.column_stack([df['Sport'].str.contains("asket", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Winter'   
    return season, year

def outputduplicates(df):
    '''Prints out names of players with duplicated entries into console... can then delete from google drive signups '''
    tempdf=df.duplicated(['First','Last']) # series with 2nd of duplicated entries as True
    for i in range(0,len(tempdf)):
        if tempdf[i]==True:
            first=df.iloc[i]['First']
            last=df.iloc[i]['Last']
            print('Duplicated signup for player: ', first,' ', last)
    return

def formatphone(df):
    ''' Convert all entered phone numbers in dfs phone columns to 314-xxx-xxxx string and standardize text field '''
    # find phone columns (named phone, phone2, etc.)
    allcols=df.columns
    phlist=[str(s) for s in allcols if 'Phone' in s]
    for i, colname in enumerate(phlist):
        for j in range(0,len(df)):
            num=str(df.iloc[j][colname]) # string with phone number
            if num=='nan':
                continue # skip reformat if nan
            if not re.search(r'(\d+-\d+-\d+)',num):
                num=re.sub("[^0-9]", "", num) # substitute blank for non-number
                if len(num)==7: 
                    num='314'+num # assume 314 area code
                if len(num)==11 and num.startswith('1'): # remove starting 1 if present
                    num=num[1:11]
                if len(num)!=10:
                    print('Bad number: ',num)
                num=num[0:3]+'-'+num[3:6]+'-'+num[6:10]
                df.set_value(j,colname,num) # write back in correct format
    # now change yes in any text field to Y
    txtlist=[str(s) for s in allcols if 'Text' in s]
    for i, colname in enumerate(txtlist):
        for j in range(0,len(df)):
            tempstr=str(df.iloc[j][colname]) # string with phone number
            if tempstr=='yes':
                df.set_value(j,colname,'Y') 
            if tempstr=='Yes':
                df.set_value(j,colname,'Y')     
    return df
    
def standardizeschool(df):
    ''' can pass any frame with school column and standardize name as Cabrini and Soulard''' 
    schstr='frances' + '|' + 'cabrini' + '|' + 'sfca' # multiple school matching string    
    mask = df['School'].str.contains(schstr, na=False, case=False)
    for i in range(0, len(df)):
        if mask[i]==1:
            df=df.set_value(i,'School','Cabrini')
    schstr='soulard'  # multiple school matching string    
    mask = df['School'].str.contains(schstr, na=False, case=False)
    for i in range(0, len(df)):
        if mask[i]==1:
            df=df.set_value(i,'School','Soulard')
    schstr='public'  # multiple school matching string    
    mask = df['School'].str.contains(schstr, na=False, case=False)
    for i in range(0, len(df)):
        if mask[i]==1:
            df=df.set_value(i,'School','SLPS')
    return df

def calculateage(df):
    '''pass df such as Juniorteams with birthdate column/timestamp format 
    return with an added column containing age in years (e.g. 6.1 yrs)'''  
    df=df.reset_index(drop=True)
    mytime=datetime.datetime.now() 
    mytime=datetime.datetime.date(mytime) # convert time to datetime.date
    df['Age']=0.0
    for i in range(0, len(df)):
        dob=df.iloc[i]['Birthdate']  
        if str(dob)=='NaT' or str(dob)=='nan': # skip age calc if DOB is missing
            continue        
        dob=datetime.datetime.date(dob) # convert pandas timestamp dob to datetime.date 
        
        age=mytime-dob # datetime timedelta
        age = round((age.days + age.seconds/86400)/365.2425,1) # get age as float
        df=df.set_value(i,'Age',age)
    return df 
      
def formatnamesnumbers(df):
    '''Switch names to title case, standardize gender, call phone/text reformat and standardize school name''' 
    if 'First' in df:
        df['First']=df['First'].str.title()
    if 'Last' in df:
        df['Last']=df['Last'].str.title()
    if 'Family' in df:
        df['Family']=df['Family'].str.title()
    if 'Pfirst' in df:
        df['Pfirst']=df['Pfirst'].str.title()
    if 'Plast' in df:
        df['Plast']=df['Plast'].str.title()
    if 'Pfirst2' in df:
        df['Pfirst']=df['Pfirst'].str.title()
    if 'Plast2' in df:
        df['Plast2']=df['Plast2'].str.title()
    df=df.replace('Girl','f')
    df=df.replace('Boy','m')
    df=formatphone(df) # call phone reformatting string
    if 'School' in df:
        df=standardizeschool(df) # use "Cabrini" and "Soulard" as school names
    return df

def writecontacts(df, season, signupfile):
    ''' Default data frame with shortened names into which google drive sheets are read; sheets are created with google
    form and contain fresh player data from forms; For paper signups some data will be missing and will be found 
    from existing player database '''
    # Slice by sport (also find season) Basketball (null for winter?), Soccer, Volleyball, Baseball, T-ball, Softball, Track) 
    from openpyxl import load_workbook #
    book=load_workbook(signupfile)
    writer=pd.ExcelWriter(signupfile, engine='openpyxl')
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    if season=='Fall':    
        thismask = df['Sport'].str.contains('soccer', case=False, na=False) & df['Gender'].str.contains('girl', case=False, na=False)
        Girlsoccer=df.loc[thismask]
        Girlsoccer=organizecontacts(Girlsoccer) # organize in correct format for xls file
        Girlsoccer=assignteams(Girlsoccer) # add team assignment
        Girlsoccer.to_excel(writer,sheet_name='Girlsoccer',index=False) # this overwrites existing file
        thismask = df['Sport'].str.contains('soccer', case=False, na=False) & df['Gender'].str.contains('boy', case=False, na=False)
        Boysoccer=df.loc[thismask]
        Boysoccer=organizecontacts(Boysoccer) # organize in correct format for xls file 
        Boysoccer=assignteams(Boysoccer) # add team assignment
        Boysoccer.to_excel(writer,sheet_name='Boysoccer',index=False) # this overwrites existing file
        thismask = df['Sport'].str.contains('v', case=False, na=False) & df['Gender'].str.contains('boy', case=False, na=False)
        BoyVB=df.loc[thismask]
        BoyVB=organizecontacts(BoyVB) # organize in correct format for xls file 
        BoyVB=assignteams(BoyVB) # add team assignment        
        BoyVB.to_excel(writer,sheet_name='BoyVB',index=False) # this overwrites existing file
        thismask = df['Sport'].str.contains('v', case=False, na=False) & df['Gender'].str.contains('girl', case=False, na=False)
        GirlVB=df.loc[thismask]
        GirlVB=organizecontacts(GirlVB) # organize in correct format for xls file 
        GirlVB=assignteams(GirlVB) # add team assignment                
        GirlVB.to_excel(writer,sheet_name='GirlVB',index=False) # this overwrites existing file
         
    if season=='Spring': 
        thismask = df['Sport'].str.contains('baseball', case=False, na=False)
        Baseball=df.loc[thismask]
        Baseball=organizecontacts(Baseball) # organize in correct format for xls file 
        Baseball=assignteams(Baseball) # add team assignment                        
        Baseball.to_excel(writer,sheet_name='Baseball',index=False) # this overwrites existing file
 
        thismask = df['Sport'].str.contains('softball', case=False, na=False)
        Softball=df.loc[thismask]
        Softball=organizecontacts(Softball) # organize in correct format for xls file 
        Softball=assignteams(Softball) # add team assignment                        
        Softball.to_excel(writer,sheet_name='Softball',index=False) # this overwrites existing file
 
        thismask = df['Sport'].str.contains('t-ball', case=False, na=False)
        Tball=df.loc[thismask]
        Tball=organizecontacts(Tball) # organize in correct format for xls file 
        Tball=assignteams(Tball) # add team assignment                        
        Tball.to_excel(writer,sheet_name='Tball',index=False) # this overwrites existing file

        thismask = df['Sport'].str.contains('track', case=False, na=False)
        Track=df.loc[thismask]
        Track=organizecontacts(Track) # organize in correct format for xls file 
        Track=assignteams(Track) # add team assignment                        
        Track.to_excel(writer,sheet_name='Track',index=False) # this overwrites existing file
    
    if season=='Winter': # currently only basketball
        Basketball=organizecontacts(df) # organize in correct format for xls file 
        Basketball=assignteams(Basketball) # add team assignment             
        Basketball.to_excel(writer,sheet_name='Basketball',index=False) # this overwrites existing file
 
    writer.save() # saves xls file with all modified data   
    return 
    
def createsignups(df, Mastersignups, season, year):     
    ''' pass SCsignup and add signups to master list, also returns list of current player keys by sport
    typically use writesignupstoExcel instead
    '''   
    df.Grade=df.Grade.replace('K',0)# replace K with zero to allow sorting
    if season=='Fall':
        mask = np.column_stack([df['Sport'].str.contains("occer", na=False)])
        Allsoccer=df.loc[mask.any(axis=1)]
        Allsoccer=Allsoccer.reset_index(drop=True)
        for i in range(0,len(Allsoccer)):
            Allsoccer.set_value(i, 'Sport', 'soccer')
        Allsoccer=organizesignups(Allsoccer, year)
                
        mask = np.column_stack([df['Sport'].str.contains("olleyball", na=False)])
        Allvb=df.loc[mask.any(axis=1)]
        Allvb=Allvb.reset_index(drop=True)
        for i in range(0,len(Allvb)):
            Allvb.set_value(i, 'Sport', 'VB')
        Allvb=organizesignups(Allvb, year)
        
        # concatenate new soccer and VB signups with master signups file
        colorder=Mastersignups.columns.tolist() # desired column order
        Mastersignups=pd.concat([Allsoccer,Mastersignups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Mastersignups=pd.concat([Allvb,Mastersignups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Mastersignups=Mastersignups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Mastersignups=Mastersignups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
        # Signups=Signups.reset_index(drop=True)
        
        # add unique SUkey (if not already assigned)
        neededkeys = Mastersignups[(Mastersignups['SUkey']==0)] # filter by year 
        availSUkeys=SC.findavailablekeys(Mastersignups, 'SUkey', len(neededkeys)) # get necessary # of unique SU keys      
        keycounter=0
        for index, row in neededkeys.iterrows():
            Mastersignups.set_value(index,'SUkey',availSUkeys[keycounter]) # reassign SU key in source master list
            keycounter+=1 # move to next available key
            
        Mastersignups.to_csv('master_signups.csv', index=False)
        return Mastersignups
        
    if season=='Winter':
        AllBball=df # no need to filter by sport in winter
        for i in range(0,len(AllBball)):
            AllBball.set_value(i, 'Sport', 'Bball')
        AllBball=organizesignups(AllBball, year)
        
        # concatenate new Bball signups with master signups file
        colorder=Mastersignups.columns.tolist() # desired column order
        Mastersignups=pd.concat([AllBball,Mastersignups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Mastersignups=Mastersignups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Mastersignups=Mastersignups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
        # Signups=Signups.reset_index(drop=True)
        # add unique SUkey (if not already assigned)
        neededkeys = Mastersignups[(Mastersignups['SUkey']==0)] # filter by year 
        availSUkeys=SC.findavailablekeys(Mastersignups, 'SUkey', len(neededkeys)) # get necessary # of unique SU keys      
        keycounter=0
        for index, row in neededkeys.iterrows():
            Mastersignups.set_value(index,'SUkey',availSUkeys[keycounter]) # reassign SU key in source master list
            keycounter+=1 # move to next available key

        Mastersignups.to_csv('master_signups.csv', index=False)
        return Mastersignups
        
    if season=='Spring':
        mask = np.column_stack([df['Sport'].str.contains("track", case=False, na=False)])
        Track=df.loc[mask.any(axis=1)]
        Track=Track.reset_index(drop=True)
        for i in range(0,len(Track)):
            Track.set_value(i, 'Sport', 'Track')
        Track=organizesignups(Track, year)
            
        mask = np.column_stack([df['Sport'].str.contains("soft", case=False, na=False)])
        Softball=df.loc[mask.any(axis=1)]
        Softball=Softball.reset_index(drop=True)
        for i in range(0,len(Softball)):
            Softball.set_value(i, 'Sport', 'Softball')
        Softball=organizesignups(Softball, year)
        
        mask = np.column_stack([df['Sport'].str.contains("base", case=False, na=False)])
        Baseball=df.loc[mask.any(axis=1)]
        Baseball=Baseball.reset_index(drop=True)
        for i in range(0,len(Baseball)):
            Baseball.set_value(i, 'Sport', 'Baseball')
        Baseball=organizesignups(Baseball, year)
        
        mask = np.column_stack([df['Sport'].str.contains("t-ball", case=False, na=False)])
        Tball=df.loc[mask.any(axis=1)]
        Tball=Tball.reset_index(drop=True)
        for i in range(0,len(Tball)):
            Tball.set_value(i, 'Sport', 'Tball')
        Tball=organizesignups(Tball, year)
        
        # concatenate new track and ?ball signups with master signups file
        colorder=Mastersignups.columns.tolist() # desired column order
        Mastersignups=pd.concat([Track,Mastersignups], ignore_index=True) # concat the two frames
        Mastersignups=pd.concat([Softball,Mastersignups], ignore_index=True) # concat the two frames
        Mastersignups=pd.concat([Baseball,Mastersignups], ignore_index=True) # concat the two frames\
        Mastersignups=pd.concat([Tball,Mastersignups], ignore_index=True) # concat the two frames
        Mastersignups=Mastersignups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Mastersignups=Mastersignups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)

        # add unique SUkey (if not already assigned)
        neededkeys = Mastersignups[(Mastersignups['SUkey']==0)] # filter by year 
        availSUkeys=SC.findavailablekeys(Mastersignups, 'SUkey', len(neededkeys)) # get necessary # of unique SU keys      
        keycounter=0
        for index, row in neededkeys.iterrows():
            Mastersignups.set_value(index,'SUkey',availSUkeys[keycounter]) # reassign SU key in source master list
            keycounter+=1 # move to next available key

        # Signups=Signups.reset_index(drop=True)
        Mastersignups.to_csv('master_signups.csv', index=False)
        return Mastersignups

def createrosters(df, season, year, players, teams, famcontact):
    ''' From Mastersignups of this season creates Cabrini CYC roster and transfers (for separate sports)
    and all junior sports (calculates ages for Judge Dowd);  pulls info merged from famcontact, players, teams, and coaches
    teams should already be assigned using teams xls and assigntoteams function'''    
    if season=='Fall':  
        df = df[(df['Year']==year)] # filter by year 
        # now grab all extra info needed for CYC rosters 
        # Street, City, State, Zip, Phone, email, Parishreg, parishres from fam-contact        
        df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))        
        # get division from Teams xls
        teams.rename(columns={'CYCname':'Team'}, inplace=True)
        df=pd.merge(df, teams, how='left', on=['Team'], suffixes=('','_r2'))        
        # DOB, School  from players.csv
        df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))        
        # set Role=Player
        df['Role']='Player' # add column for role
        df['Open/Closed']='Closed'
        df['Coach ID']=''
        
        # Find Cabrini CYC names (containing hyphen)        
        thismask = df['Team'].str.contains('-', case=False, na=False)
        CabriniCYC=df.loc[thismask] # all  players on Cabrini CYC teams (for single soccer roster file creation)
        
        # Now split by sport and save 
        thismask = CabriniCYC['Sport'].str.contains('soccer', case=False, na=False)
        Soccer=CabriniCYC.loc[thismask] # all soccer players        
        Soccer=organizeroster(Soccer) # reformat this mess as single CYC roster
        fname='Cabrini_soccer_rosters_'+str(year)+'.csv'    
        Soccer.to_csv(fname, index=False)

        thismask = CabriniCYC['Sport'].str.contains('VB', case=False, na=False)
        VB=CabriniCYC.loc[thismask] # all VB players        
        VB=organizeroster(VB) # reformat this mess as single CYC roster
        fname='Cabrini_VB_rosters_'+str(year)+'.csv'    
        VB.to_csv(fname, index=False)
        # Find CYC transfers and junior teams
        thismask = df['Team'].str.contains('-', case=False, na=False)
        Others=df.loc[~thismask] # all teams that are not Cabrini CYC level (junior and transfers)
        
        # Cabrini transferred players (teams are school names without a number)
        # non-CYC cabrini junior teams start with number
        pattern=r'[0-9]' # starts with letter
        thismask = Others['Team'].str.contains(pattern, na=True) # flag  nans and set to true (usually jr teams w/o assignment)
        Transfers=Others.loc[~thismask]
        Juniorteams=Others.loc[thismask] # junior non-CYC cabrini team has number but not hyphen
        
        # output single roster for all transferred soccer players        
        SoccerTransfers= Transfers[Transfers['Sport'].str.contains('soccer', case=False, na=False)]
        SoccerTransfers=organizeroster(SoccerTransfers)
        SoccerTransfers=SoccerTransfers.sort_values(['Team', 'Sex', 'Grade'], ascending=True)
        fname='CYCSoccer_transfers'+str(year)+'.csv'    
        SoccerTransfers.to_csv(fname, index=False)
        # now output CYC roster for all transferred VB players 
        VBTransfers= Transfers[Transfers['Sport'].str.contains('VB', case=False, na=False)]
        VBTransfers=organizeroster(VBTransfers)
        VBTransfers=VBTransfers.sort_values(['Team', 'Sex', 'Grade'], ascending=True)
        fname='CYCVB_transfers'+str(year)+'.csv'    
        VBTransfers.to_csv(fname, index=False)
        
        # now output all junior teams in same format (sometimes needed by Judge Dowd)
        # also calculate current age
        Juniorteams=organizeroster(Juniorteams)
        Juniorteams=calculateage(Juniorteams)
        fname='Cabrini_junior_teams_'+str(year)+'.csv'    
        Juniorteams.to_csv(fname, index=False)
        return
        
    if season=='Winter':
        df = df[(df['Year']==year)] # filter by year 
        # now grab all extra info needed for CYC rosters 
        # Street, City, State, Zip, Phone, email, Parishreg, parishres from fam-contact        
        df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))        
        # get division from Teams xls
        teams.rename(columns={'CYCname':'Team'}, inplace=True)
        df=pd.merge(df, teams, how='left', on=['Team'], suffixes=('','_r2'))        
        # DOB, School  from players.csv
        df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))        
        # set Role=Player
        df['Role']='Player' # add column for role
        df['Open/Closed']='Closed'
        df['Coach ID']=''
        # Find Cabrini CYC names (containing hyphen)        
        thismask = df['Team'].str.contains('-', case=False, na=False)
        CabriniCYC=df.loc[thismask] # all  players on Cabrini CYC teams
        
        # Now split by sport (basketball only and save)
        thismask = CabriniCYC['Sport'].str.contains('basketball', case=False, na=False)
        Basketball=CabriniCYC.loc[thismask] # all soccer players        
        Basketball=organizeroster(Basketball) # reformat this mess as single CYC roster
        fname='Cabrini_basketball_rosters_'+str(year)+'.csv'    
        Basketball.to_csv(fname, index=False)
        
        # Find CYC transfers and junior teams
        thismask = df['Team'].str.contains('-', case=False, na=False)
        Others=df.loc[~thismask] # all teams that are not Cabrini CYC level (junior and transfers)
        
        # Cabrini transferred players (teams are school names without a number)
        # non-CYC cabrini junior teams start with number
        pattern=r'[0-9]' # starts with letter
        thismask = Others['Team'].str.contains(pattern, na=True) # flag  nans and set to true (usually jr teams w/o assignment)
        Transfers=Others.loc[~thismask]
        Juniorteams=Others.loc[thismask] # junior non-CYC cabrini team has number but not hyphen
        
        # output single roster for all transferred soccer players        
        Bballtransfers= Transfers[Transfers['Sport'].str.contains('soccer', case=False, na=False)]
        Bballtransfers=organizeroster(Bballtransfers)
        Bballtransfers=Bballtransfers.sort_values(['Team', 'Sex', 'Grade'], ascending=True)
        fname='CYC_basketball_transfers'+str(year)+'.csv'    
        Bballtransfers.to_csv(fname, index=False)
        
        # now output all junior teams in same format (sometimes needed by Judge Dowd)
        # also calculate current age
        Juniorteams=organizeroster(Juniorteams)
        Juniorteams=calculateage(Juniorteams)
        fname='Cabrini_junior_teams_'+str(year)+'.csv'    
        Juniorteams.to_csv(fname, index=False)
        return
        
    if season=='Spring':
        df = df[(df['Year']==year)] # filter by year 
        # now grab all extra info needed for CYC rosters 
        # Street, City, State, Zip, Phone, email, Parishreg, parishres from fam-contact        
        df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))        
        # get division from Teams xls
        teams.rename(columns={'CYCname':'Team'}, inplace=True)
        df=pd.merge(df, teams, how='left', on=['Team'], suffixes=('','_r2'))        
        # DOB, School  from players.csv
        df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))        
        # set Role=Player
        df['Role']='Player' # add column for role
        df['Open/Closed']='Closed'
        df['Coach ID']=''
        # Find Cabrini CYC names (containing hyphen)        
        thismask = df['Team'].str.contains('-', case=False, na=False)
        allCYCball=df.loc[thismask] # all  players on Cabrini CYC teams
        
        # all CYC softball, baseball, petball teams can be in single file (track and Tball already out)
        allCYCball=organizeroster(allCYCball) # reformat this mess as single CYC roster
        fname='Cabrini_allCYCball_rosters_'+str(year)+'.csv'    
        allCYCball.to_csv(fname, index=False)
        
        # track rosters
        Track= df[df['Sport'].str.contains('track', case=False, na=False)]
        Track=organizeroster(Track) # format same as CYC roster  
        # TODO calculate track age group        
        # calculate age for track team
        Track=calculateage(Track)
        fname='Track_roster_'+str(year)+'.csv'    
        Track.to_csv(fname, index=False)
        
        # Don't need any formal roster for Tball... just contact sheet per team
        return    

def assignteams(df, Teams, sport):
    '''Pass contacts summary and assign team name (similar to assigntoteams called by mastersignups
    Teams tab must have those for current year; merge based on grade, gender and sport from teams tab 
    (which only contains names from this season/year to avoid screwing up old custom team assignments''' 
    Teamsmult=makemultiteam(Teams) # makes duplicates team entries to match both grades
    Teamsmult['Grade']=Teamsmult['Grade'].astype('str') # convert grade back to string
    Teamsmult.Grade=Teamsmult.Grade.replace('0','K', regex=True) # convert 0 back to K
    df['Grade']=df['Grade'].astype('str')
    # left merge keeps all master_signups entries
    df=pd.merge(df, Teamsmult, how='left', on=['Grade','Gender','Sport'], suffixes=('','_r'))
    df['Team']=df['Team'].astype('str') # convert grade back to string
    # now copy over CYC team name from Teams_coaches to this df ... skip copying if null
    for i in range(0, len(df)):
        if df.iloc[i]['CYCname']!='nan':
            df.set_value(i, 'Team', df.iloc[i]['CYCname'])
    # now drop extra columns
    mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone', 'Text', 'Email', 'Phone2', 'Text2', 'Email2', 'Team', 'Plakey','Famkey', 'Family'] 
    tempdf=pd.DataFrame(columns=mycols) # formatted to match usual contact sheet
    df=dropcolumns(df,tempdf)
    return df

def makemultiteam(df):
    '''Small utility called by assigntoteams to make temp teams df that has separate entry for each grade if team is mixed grade
    then merge to assign teams is straightforward'''
    df['Graderange']=df['Graderange'].astype('str')
    df.Graderange=df.Graderange.replace('K','0', regex=True) # replace K with 0
    mask=df['Graderange'].str.len() > 1
    multiteams=df.loc[mask]
    multiteams['Grade']=multiteams['Grade'].astype('int')
    multiteams=multiteams.reset_index(drop=True)
    # Use this to make a multiple entry at lower grade(s).. normally just one extra
    for i in range(0,len(multiteams)):
        # TODO make sure it's not 3 grades (i.e. K-2)
        multiteams.set_value(i,'Grade',multiteams.iloc[i]['Grade']-1)
    # now combine with original df
    df=pd.concat([df,multiteams], ignore_index=True)
    return df

def assigntoteams(df, season, year, Teams):
    '''Finds CYC team name based on year, grade, gender and sport from teams tab 
    (which only contains names from this season/year to avoid screwing up old custom team assignments''' 
    Teamsmult=makemultiteam(Teams) # makes duplicates team entries to match both grades
    Teamsmult['Grade']=Teamsmult['Grade'].astype('str') # convert grade back to string
    Teamsmult.Grade=Teamsmult.Grade.replace('0','K', regex=True) # convert 0 back to K
    df['Grade']=df['Grade'].astype('str')
    # left merge keeps all master_signups entries
    df=pd.merge(df, Teamsmult, how='left', on=['Year','Grade','Gender','Sport'], suffixes=('','_r'))
    df['Team']=df['Team'].astype('str') # convert grade back to string
    # now copy over CYC team name from Teams_coaches to mastersignups ... skip copying if null
    for i in range(0, len(df)):
        # entry will be nan for prior years/seasons
        if df.iloc[i]['CYCname']!='nan':
            df.set_value(i, 'Team', df.iloc[i]['CYCname'])
    # now drop extra columns
    mycols=['First', 'Last', 'Grade', 'Gender', 'Sport', 'Year', 'Team', 'Plakey','Famkey', 'Family'] 
    tempdf=pd.DataFrame(columns=mycols) # temp df for dropping unnecessary columns
    df=dropcolumns(df,tempdf) # drop columns not working
    df.to_csv('master_signups.csv', index=False) # save/overwrite existing csv
    return df
    

def findrecruits(df, players, famcontact, season, year, signupfile):
    '''Read list of signed-up player keys from xls file; compare with last year's set of 
    players from master Signups log '''
    mycols=df.columns.tolist()
    if season=='Fall':  
        Recruits=pd.DataFrame(columns=mycols) # single file with info on all recruits
        soccer=df[df['Sport']=='soccer'] # filter for soccerd
        plakeylist=soccer.Plakey.unique() # ndarray with list of unique soccer players      
        keylist=plakeylist.tolist()        
        for i, key in enumerate(keylist):
            match=soccer[soccer['Plakey']==key]
            # recruits ... played in year -1 but not in year
            if year-1 in match.Year.unique() and year not in match.Year.unique():
                match=match[0:1]
                Recruits=pd.concat([Recruits,match], ignore_index=True)
        vb=df[df['Sport']=='VB'] # filter for soccer
        plakeylist=vb.Plakey.unique()        
        keylist=plakeylist.tolist()        
        for i, key in enumerate(keylist):
            match=vb[vb['Plakey']==key]
            # recruits ... played in year -1 but not in year
            if year-1 in match.Year.unique() and year not in match.Year.unique():
                match=match[0:1]
                Recruits=pd.concat([Recruits,match], ignore_index=True)            
    # plakey, famkey, first, last, grade, gender, 
    Recruits.Grade=Recruits.Grade.replace('K',0) # replace K with zero to allow sorting
    Recruits.Grade=Recruits.Grade.astype(int)
    # adjust grade such that players current grade is in list
    for i in range(len(Recruits)):
        grade=Recruits.iloc[i]['Grade']
        Recruits=Recruits.set_value(i,'Grade',grade+1) 
    # join with famcontact on famkey to get contact info (emails, phones, etc.)
    # Inner join on famkey adds the necessary info
    Recruits=pd.merge(Recruits, famcontact,how='inner', on='Famkey', suffixes=('','_r'))
    Recruits['School']='' # temporarily add school column (not yet looked up from players)    
    Recruits=organizerecruits(Recruits) # reformat into same as Excel signups,  summary
    # now need to look up school from players.csv    
    Recruits=pd.merge(Recruits, players, how='inner', on='Plakey', suffixes=('','_r'))    
    Recruits.drop('School', axis=1, inplace=True)
    Recruits.rename(columns={'School_r':'School'}, inplace=True)
    Recruits=organizerecruits(Recruits) # reformat into same as Excel signups,  summary      
    Recruits=Recruits.sort_values(['Sport', 'Gender', 'Grade'], ascending=True)
    # now write recruits to tab in master signups file
    from openpyxl import load_workbook
    book=load_workbook(signupfile)
    writer=pd.ExcelWriter(signupfile, engine='openpyxl')
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    Recruits.to_excel(writer,sheet_name='Recruits',index=False) # this overwrites existing file
    writer.save() # saves xls file with all modified data   
    
def findmissinginfo(df, players, famcontact):
    ''' using player and family keys, update nan values in SC signups (mainly for paper/word-of-mouth entries
    needed for writesignupstoExcel '''
    for i in range(0, len(df)):
        if type(df.iloc[i]['Timestamp'])!=datetime.datetime:
            thisplakey=df.iloc[i]['Plakey']
            thisfamkey=df.iloc[i]['Famkey']
            # get first, dob, school from master players list
            match = players[(players['Plakey']==thisplakey)]
            if len(match)==1:
                df=df.set_value(i,'First', match.iloc[0]['First']) # probably same but doublecheck/ set to CYC card value
                df=df.set_value(i,'DOB', match.iloc[0]['DOB'])
                df=df.set_value(i,'School', match.iloc[0]['School'])
            # get address, zip, parish, phone/text, email, phone2, text2, email2 from famcontact 
            match = famcontact[(famcontact['Famkey']==thisfamkey)]
            if len(match)==1:
                df=df.set_value(i,'Address', match.iloc[0]['Address'])
                df=df.set_value(i,'Zip', match.iloc[0]['Zip'])
                df=df.set_value(i,'Parish', match.iloc[0]['Parish_registration'])
                df=df.set_value(i,'Phone', match.iloc[0]['Phone1'])
                df=df.set_value(i,'Text', match.iloc[0]['Text1'])
                df=df.set_value(i,'Email', match.iloc[0]['Email1'])
                df=df.set_value(i,'Phone2', match.iloc[0]['Phone2'])
                df=df.set_value(i,'Text2', match.iloc[0]['Text2'])
                df=df.set_value(i,'Email2', match.iloc[0]['Email2'])
    return df
    
def findplayernumbers(SCsignup, players):
    '''Find player key from players df using multiple levels of matching (Plakey already initialized)
    if not a perfect match on all characters, create some data output structure to resolve possible problems'''
        
    if 'Plakey' not in SCsignup.columns: # add player key column to sign-ups file if not present
        SCsignup.insert(0,'Plakey',0) # insert as first col and init to 0
    if 'Famkey' not in SCsignup.columns: # add family key to sign-ups file
        SCsignup.insert(1,'Famkey',0) # add as 2nd col and init to 0
    if 'Family' not in SCsignup.columns: # add family key to sign-ups file
        SCsignup.insert(2,'Family', '') # add as 2nd col and init to blank 
    for su in range(0,len(SCsignup)):
        if SCsignup.iloc[su]['Plakey']==0: # some may already have assigned player number from prior run
            first=SCsignup.iloc[su]['First']
            last=SCsignup.iloc[su]['Last']
            DOB=SCsignup.iloc[su]['DOB'] # usually datetime format
            # TODO case insensitive matches to avoid DeFord problem
            match = players[(players['Last']==last) & (players['First']==first) & (players['DOB']==DOB)]
            if len(match)==1: # single perfect match on all 3 
                SCsignup=SCsignup.set_value(su,'Plakey',match.iloc[0]['Plakey'])
                SCsignup=SCsignup.set_value(su,'Famkey',match.iloc[0]['Famkey'])
                SCsignup=SCsignup.set_value(su,'Family',match.iloc[0]['Family'])
                continue # move to next signup
            elif len(match)==0: # try last and DOB match (first name error)
                match = players[(players['Last']==last) & (players['DOB']==DOB)]
                if len(match)==1: # single perfect match on last and DOB
                    SCsignup=SCsignup.set_value(su,'Plakey',match.iloc[0]['Plakey'])
                    SCsignup=SCsignup.set_value(su,'Famkey',match.iloc[0]['Famkey'])
                    SCsignup=SCsignup.set_value(su,'Family',match.iloc[0]['Family'])
                    continue # move to next signup
                elif len(match)>1: # multiple match on last and DOB (twins problem with slight first name difference)
                    SCsignup=SCsignup.set_value(su,'Plakey',match.iloc[0]['Plakey'])
                    SCsignup=SCsignup.set_value(su,'Famkey',match.iloc[0]['Famkey'])
                    SCsignup=SCsignup.set_value(su,'Family',match.iloc[0]['Family'])
                    continue 
                elif len(match)==0: # no last/DOB match so check first/last (possible DOB error)
                    match = players[(players['Last']==last) & (players['First']==first)] #check for possible DOB entry error
                    if len(match)==1: # first/last exact match but not DOB
                        print('DOB entry error suspected for found player: ', first, ' ', last)
                        SCsignup=SCsignup.set_value(su,'Plakey',match.iloc[0]['Plakey'])
                        SCsignup=SCsignup.set_value(su,'Famkey',match.iloc[0]['Famkey'])
                        SCsignup=SCsignup.set_value(su,'Family',match.iloc[0]['Family'])                   
                        continue
                        # TODO maybe write these to a check DOB file??
                    if len(match)>1: # first/last matches multiple players (name is too common?)
                        print('No DOB match. First last matches multiple players:',first, ' ',last)
                    else:
                        print('Likely new player :',first,' ',last,'.' ) #no first/last match, no last/DOB match
                        try: # try appending to existing newplayers dataframe... if fail, then create dataframe 
                            newplayers=newplayers.append(SCsignup.iloc[su])
                        except:
                            newplayers=SCsignup.iloc[[su]] # new df with current row copied over
                    # generate entries for new players list (maybe temporarily split)
                    continue
            elif len(match)>1: # somehow duplicate entries in players dataframe
                print('Remove duplicated entry for ', first," ",last ,' from players list.')
                continue
    try:
        isinstance(newplayers,pd.DataFrame) # check for non-existant dataframe
    except:
        newplayers=pd.DataFrame # in case no new players are found, create and return empty frame
    return SCsignup, newplayers # same df but with all available player numbers added in playkey column

def findfamily(newplayers,famcontact):
    ''' For confirmed new players, find existing family name and key (if it exists)''' 
    newplayers=newplayers.reset_index(drop=True) # reset index to avoid problems
    if 'Famkey' not in newplayers.columns: # add player key column to sign-ups file if not present
        newplayers.insert(1,'Famkey',0)
    if 'Family' not in newplayers.columns: # add player key column to sign-ups file if not present
        newplayers.insert(0,'Family','')
    for su in range(0,len(newplayers)):
        first=newplayers.iloc[su]['First']         
        last=newplayers.iloc[su]['Last'] 
        phone1=newplayers.iloc[su]['Phone']
        phstr=last + '|' + phone1 # multiple phone matching string
        phone2=str(newplayers.iloc[su]['Phone2']) # need to test for nan
        if phone2!='nan': # if entered add 2nd phone to matching string
            phstr=phstr + '|' + phone2 
        mask = famcontact['Phone1'].str.contains(phstr, na=False, case=False) | famcontact['Phone2'].str.contains(phstr, na=False) | famcontact['Phone3'].str.contains(phstr, na=False)
        match=famcontact.loc[mask] # filter df with above phone matching mask
        if len(match)==1:
            print('Phone match of player ', first, last,' to family ', match.iloc[0]['Family'])
            newplayers.set_value(su,'Family',match.iloc[0]['Family']) # 
            newplayers.set_value(su,'Famkey',match.iloc[0]['Famkey'])
            continue # famkey assigned and move to next if phone match
        mask = famcontact['Family'].str.contains(last, na=False, case=False) # check for last name match
        match=famcontact.loc[mask] # filter df with above phone matching mask
        if len(match)>0: # check for possible family matches (even though no phone match)
            for i in range(0,len(match)):
                tempstr=''
                tempstr=tempstr+str(match.iloc[i]['Famkey'])+' ' 
                tempstr=tempstr+str(match.iloc[i]['Family'])+' ' # string with family key # and name for all possibles
            print('Possible match of player ', first, last,' to family ', tempstr,'not yet assigned') # lists all
            # don't assign number but can be done manually if reasonable
        else:
            print('No match for player ', first, last)
    return newplayers # send back same df with family name and key

def addplayers(df, players):
    ''' gets info from confirmed newplayers,reformat and adds to main players list '''
    # playerentries = pd.DataFrame(index=np.arange(len(newplayers)), columns=players.columns)
    mytime=datetime.datetime.now()
    datestr='_' + str(mytime.day) + mytime.strftime("%B") + str(mytime.year)[2:] # current date as in 3Jun16
    filestr='players'+datestr+'.bak'
    players.to_csv(filestr, index=False) # backup existing players master list
    df=df.iloc[:,0:10]   # cut off extra columns at end 
    df.drop('Timestamp', axis=1, inplace=True)
    df['Gradeadj']=0 # add grade adjust col and init to zero
    colorder=players.columns.tolist() # desired column order
    players=pd.concat([df,players]) # concat the two frames (if same names, column order doesn't matter)
    players=players[colorder] # put back in original order
    players=players.reset_index(drop=True)
    players=players.to_csv('players.csv',index =False)
    return players # full master list with new entries
    
def addfamilies(df,famcontact, fambills):
    ''' df contains new families to add to master family contact and family billing tables '''
    df=df.reset_index(drop=True) # first reset index for proper for loops
    dfcon=df # make copy of original for famcontact below
    # backup existing family contact and billing tables    
    mytime=datetime.datetime.now()
    datestr='_' + str(mytime.day) + mytime.strftime("%B") + str(mytime.year)[2:] # current date as in 3Jun16
    filestr='family_contact'+datestr+'.bak'
    famcontact.to_csv(filestr, index=False) # backup existing players master list
    filestr='family_bill'+datestr+'.bak'
    fambills.to_csv(filestr, index=False) # backup existing players master list
    
    # update family billing
    datestr=str(mytime.month)+'/'+str(mytime.day)+'/'+str(mytime.year)
    df=dropcolumns(df,fambills) # drops all cols from df that are not in fambills
    # df=df.iloc[:,0:3]
    # df.drop('Plakey', axis=1, inplace=True)
    df['Startdate']=datestr # add and initialize other necessary columns
    df['Lastupdate']=datestr
    df['Startbal']=0
    df['Currbalance']=0
    df['Billing_note']=''
    colorder=fambills.columns.tolist()
    fambills=pd.concat([fambills,df]) # concat the two frames (if same names, column order doesn't matter)
    fambills=fambills[colorder] # put back in original order
    fambills=fambills.reset_index(drop=True)
    fambills=fambills.to_csv('family_bill.csv',index =False)
    
    # update family contact
    dfcon.rename(columns={'Plakey': 'Players', 'Parish': 'Parish_registration', 'Phone': 'Phone1', 'Text': 'Text1', 'Email': 'Email1',}, inplace=True)
    dfcon=dropcolumns(dfcon, famcontact) # drop unnecessary columns from dfcon (not in famcontact)
    #dfcon.drop('Timestamp', axis=1, inplace=True)
    #dfcon.drop('First', axis=1, inplace=True)
    #dfcon.drop('Last', axis=1, inplace=True)
    #dfcon.drop('DOB', axis=1, inplace=True)
    #dfcon.drop('Gender', axis=1, inplace=True)
    #dfcon.drop('School', axis=1, inplace=True)
    #dfcon.drop('Grade', axis=1, inplace=True)
    #dfcon.drop('AltPlacement', axis=1, inplace=True)
    #dfcon.drop('Ocstatus', axis=1, inplace=True)
    #dfcon.drop('Othercontact', axis=1, inplace=True)
    #dfcon.drop('Coach', axis=1, inplace=True)
    #dfcon.drop('Coach2', axis=1, inplace=True)
    #dfcon.drop('Sport', axis=1, inplace=True)
    dfcon['City']='St. Louis'
    dfcon['State']='MO'
    dfcon['Parish_residence']=''
    dfcon['Pfirst3']=''
    dfcon['Plast3']=''
    dfcon['Phone3']=''
    dfcon['Text3']=''
    dfcon['Phone4']=''
    dfcon['Text4']=''
    dfcon['Email3']=''
    
    colorder=fambills.columns.tolist()
    famcontact=pd.concat([famcontact,dfcon]) # concat the two frames (if same names, column order doesn't matter)
    famcontact=famcontact[colorder] # put back in original order
    famcontact=famcontact.reset_index(drop=True)
    famcontact=famcontact.to_csv('family_contact.csv',index =False)
    return famcontact, fambills 
    
def addnewplafams(newplayers, players, famcontact, fambills):
    '''Newplayers contains google drive entries for players that don't match master players list 
    Manual/visual check should be done before adding to players list ''' 
    ''' TODO # double check again to ensure these haven't been added... pass to findplayernumbers above
    SCsignup, doublecheck = findplayernumbers(newplayers,players) 
    if len(doublecheck)!=len(newplayers):
        print("Doublecheck to ensure that these players haven't already been entered")
        break '''
    # first assign new player and family keys (index already reset by findfamily)
    # get new lists of available player and family keys    
    availplakeys=findavailablekeys(players, 'Plakey', len(newplayers)) # gets needed # of new plakey values
    availfamkeys=findavailablekeys(famcontact, 'Famkey', len(newplayers))

    # assign new player keys and create df with info to add to master players 
    
    for i,val in enumerate(availplakeys): # i is row of newplayers, val is next available key
        newplayers.set_value(i,'Plakey', val)

    # create family name for new families
    for i in range(0,len(newplayers)): # all new players
        if newplayers.iloc[i]['Famkey']==0: # only assign names to new families (skip those with existing)
            last=str(newplayers.iloc[i]['Last'])
            plast=str(newplayers.iloc[i]['Plast']) # parent last name
            if last==plast: # same name for parent and player
                newplayers.set_value(i,'Family', newplayers.iloc[i]['Last']) # assign player last name as family name
            elif str(newplayers.iloc[i]['Plast'])=='nan': # parent last name missing
                newplayers.set_value(i,'Family', newplayers.iloc[i]['Last']) # assign player last name as family name
            elif plast in last: # if kid name is hyphenated multiname, use the hyphenated name
                newname=str(newplayers.iloc[i]['Last'])+'_'+str(newplayers.iloc[i]['Plast'])
                newplayers.set_value(i,'Family', newplayers.iloc[i]['Last'])
            else: # make a new family name for different first/last
                newname=last+'_'+plast
                newplayers.set_value(i,'Family', newname)

    # find rows in newplayers that need a family
    newfams=[] # keep a list of the new family numbers (for write to famcontact)
    for i in range(0,len(newplayers)):
        if newplayers.iloc[i]['Famkey']==0:
            newfams.append(i)
    for i,val in enumerate(availfamkeys): # this misses some unassigned keys but that's fine
        if i in newfams:
            newplayers.set_value(i,'Famkey', val) # assign new unique famkey to players without families
            
    # from newplayers entries, create entries for master players table
    players=addplayers(newplayers,players) # update master players list, save and return

    # slice newplayers to knock out those w/ existing families (using above list)
    df=newplayers[newplayers.index.isin(newfams)]
    
    # update master families lists, save and return
    famcontact, fambills =addfamilies(df,famcontact, fambills) 
    
    return players, famcontact, fambills # go ahead and pass back modified versions to main

def comparefamkeys(players,famcontact, fambills):
    '''Utility script to compare family contacts, family bills and players list '''
    fams_pla=players.Famkey.unique()
    fams_pla=np.ndarray.tolist(fams_pla)
    fams_con=famcontact.Famkey.unique()
    fams_con=np.ndarray.tolist(fams_con)
    fams_bill=fambills.Famkey.unique()
    fams_bill=np.ndarray.tolist(fams_bill)
    # compare contacts and billing ... should be identical
    billonly=[i for i in fams_bill if i not in fams_con]
    cononly=[i for i in fams_con if i not in fams_bill]
    noplayers=[i for i in fams_con if i not in fams_pla]
    for i,val in enumerate(billonly):
        print("Famkey ", val, " in family billing but not in family contacts.")
    for i,val in enumerate(cononly):
        print("Famkey ", val, " in family contacts but not in family billing.")
    for i,val in enumerate(noplayers):
        print("Famkey ", val, " in family contacts but not found among players.")
    # look for different names between family contacts and family bills
    for i in range(0,len(famcontact)):
        famkey=famcontact.iloc[i]['Famkey'] # grab this key
        family=famcontact.iloc[i]['Family']
        family=family.title()   # switch to title case
        family=family.strip() # remove whitespace
        match=fambills[fambills['Famkey']==famkey]
        if len(match)==1: # should be a match already assuming above section finds no discrepancies
            family2=match.iloc[0]['Family']
            family2=family2.title()   # switch to title case
            family2=family2.strip() # remove whitespace
            if family!=family2: # different family names with same key
                print("Family key ", str(famkey), ": ", family, " or ", family2)
    # Now check for family name discrepancies between players and famcontact
    for i in range(0,len(famcontact)):
        famkey=famcontact.iloc[i]['Famkey'] # grab this key
        family=famcontact.iloc[i]['Family']
        family=family.title()   # switch to title case
        family=family.strip() # remove whitespace
        match=players[players['Famkey']==famkey]
        if len(match)==1: # should be a match already assuming above section finds no discrepancies
            family2=match.iloc[0]['Family']
            family2=family2.title()   # switch to title case
            family2=family2.strip() # remove whitespace
            if family!=family2: # different family names with same key
                print("Family key ", str(famkey), ": ", family, " or ", family2)
    return

#%% Legacy or one-time use functions
    
def createsignupsold(df, Signups, season, year):     
    ''' pass SCsignup and add signups to master list, also returns list of current player keys by sport
    typically use writesignupstoExcel instead
    '''   
    df.Grade=df.Grade.replace('K',0)# replace K with zero to allow sorting
    if season=='Fall':
        mask = np.column_stack([df['Sport'].str.contains("occer", na=False)])
        Allsoccer=df.loc[mask.any(axis=1)]
        Allsoccer=Allsoccer.reset_index(drop=True)
        for i in range(0,len(Allsoccer)):
            Allsoccer.set_value(i, 'Sport', 'soccer')
        Allsoccer=organizesignups(Allsoccer, year)
        Soccerlist=Allsoccer.Plakey.unique()
        Soccerlist=Soccerlist.tolist()
        
        mask = np.column_stack([df['Sport'].str.contains("olleyball", na=False)])
        Allvb=df.loc[mask.any(axis=1)]
        Allvb=Allvb.reset_index(drop=True)
        for i in range(0,len(Allvb)):
            Allvb.set_value(i, 'Sport', 'VB')
        Allvb=organizesignups(Allvb, year)
        VBlist=Allvb.Plakey.unique()
        VBlist=VBlist.tolist() # construct list of player keys for this season

        # concatenate new soccer and VB signups with master signups file
        colorder=Signups.columns.tolist() # desired column order
        Signups=pd.concat([Allsoccer,Signups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Signups=pd.concat([Allvb,Signups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Signups=Signups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Signups=Signups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
        # Signups=Signups.reset_index(drop=True)
        Signups.to_csv('master_signups.csv', index=False)
        return Soccerlist, VBlist, Signups
        
    if season=='Winter':
        AllBball=df # no need to filter by sport in winter
        for i in range(0,len(AllBball)):
            AllBball.set_value(i, 'Sport', 'Bball')
        AllBball=organizesignups(AllBball, year)
        Bballlist=AllBball.Plakey.unique()
        Bballlist=Bballlist.tolist() # create list of keys of all Bball players
        
        # concatenate new Bball signups with master signups file
        colorder=Signups.columns.tolist() # desired column order
        Signups=pd.concat([AllBball,Signups], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
        Signups=Signups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Signups=Signups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
        # Signups=Signups.reset_index(drop=True)
        Signups.to_csv('master_signups.csv', index=False)
        return Bballlist, Signups
        
    if season=='Spring':
        mask = np.column_stack([df['Sport'].str.contains("rack", na=False)])
        Track=df.loc[mask.any(axis=1)]
        Track=Track.reset_index(drop=True)
        for i in range(0,len(Track)):
            Track.set_value(i, 'Sport', 'Track')
        Track=organizesignups(Track, year)
        Tracklist=Track.Plakey.unique()
        Tracklist=Tracklist.tolist()            
            
        mask = np.column_stack([df['Sport'].str.contains("soft", case=False, na=False)])
        Softball=df.loc[mask.any(axis=1)]
        Softball=Softball.reset_index(drop=True)
        for i in range(0,len(Softball)):
            Softball.set_value(i, 'Sport', 'Softball')
        Softball=organizesignups(Softball, year)
        SBlist=Softball.Plakey.unique()
        SBlist=SBlist.tolist()  
        
        mask = np.column_stack([df['Sport'].str.contains("base", case=False, na=False)])
        Baseball=df.loc[mask.any(axis=1)]
        Baseball=Baseball.reset_index(drop=True)
        for i in range(0,len(Baseball)):
            Baseball.set_value(i, 'Sport', 'Baseball')
        Baseball=organizesignups(Baseball, year)
        BBlist=Baseball.Plakey.unique()
        BBlist=BBlist.tolist()  
        
        mask = np.column_stack([df['Sport'].str.contains("t-ball", case=False, na=False)])
        Tball=df.loc[mask.any(axis=1)]
        Tball=Tball.reset_index(drop=True)
        for i in range(0,len(Tball)):
            Tball.set_value(i, 'Sport', 'Tball')
        Tball=organizesignups(Tball, year)
        TBlist=Tball.Plakey.unique()
        TBlist=TBlist.tolist()  
        
        # concatenate new track and ?ball signups with master signups file
        colorder=Signups.columns.tolist() # desired column order
        Signups=pd.concat([Track,Signups], ignore_index=True) # concat the two frames
        Signups=pd.concat([Softball,Signups], ignore_index=True) # concat the two frames
        Signups=pd.concat([Baseball,Signups], ignore_index=True) # concat the two frames\
        Signups=pd.concat([Tball,Signups], ignore_index=True) # concat the two frames
        Signups=Signups[colorder] # put back in original order
        
        # drop duplicates and save master signups file        
        Signups=Signups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
        # Signups=Signups.reset_index(drop=True)
        Signups.to_csv('master_signups.csv', index=False)
        return Tracklist, SBlist, BBlist, TBlist, Signups


    
    