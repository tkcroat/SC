# -*- coding: utf-8 -*-
"""
Created on Wed May 24 16:15:24 2017
Sponsors Club  messaging functions

@author: tkc
"""
import pandas as pd
import smtplib
import numpy as np
import datetime
import tkinter as tk
import glob
import re
import sys
import math
import textwrap
from tkinter import filedialog
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pkg.SC_signup_functions import findcards
from openpyxl import load_workbook
import pkg.SC_config as cnf

def emailparent_tk(teams, season, year):
    ''' Inteface for non-billing email messages to parents (non-generic)
    Message types include:
        recruit - specific inquiry about player from last year not yet signed up; needs signupfile w/ recruits tab
        assign - notify of team assignment, optional recruit for short team, CYC card notify; teams/cards/mastersignups
        missinguni - ask about missing uniforms; missingunifile
        unireturn - generic instructions for uniform return; mastersignups w/ unis issued
        askforcards - check for CYC card on file and ask
        other -- Generic single all team+coaches message (can have $SCHOOL, $GRADERANGE,$COACHINFO, $SPORT, $PLAYERLIST) 
    8/9/17  works for team assignments
    TODO test recruit, missing unis, unireturn
    
    args:
        teams - df w/ active teams
        season -'Winter', 'Fall' or 'Spring'
        year - starting sport year i.e. 2019 for 2019-20 school year
    '''
#%%
    # first print out existing info in various lines
    root = tk.Tk()
    root.title('Send e-mail to parents')
    messageframe=tk.LabelFrame(root, text='Message options')
    unifilename=tk.StringVar()
    try:
        unifiles=glob.glob('missingunilist*') # find most recent uniform file name
        if len(unifiles)>1:
            unifile=findrecentfile(unifiles) # return single most recent file
        else:
            unifile=unifiles[0]
        # find most recent missing uni file name
        unifilename.set(unifile)
    except: # handle path error
        unifilename.set('missingunilist.csv')
                
    recruitbool=tk.BooleanVar()  # optional recruiting for short teams
    emailtitle=tk.StringVar()  # e-mail title 
    mtype=tk.StringVar()  # coach message type
    messfile=tk.StringVar() # text of e-mail message
    transmessfile=tk.StringVar() # text of e-mail message for transfers
    extravar=tk.StringVar() # use depends on message type... normally filename
    extraname=tk.StringVar() # name for additional text entry box (various uses mostly filenames)
    extraname.set('Extra_file_name.txt') # default starting choice
    choice=tk.StringVar()  # test or send -mail 
    
    def chooseFile(txtmess, ftypes):
        ''' tkinter file chooser (passes message string for window and expected
        file types as tuple e.g. ('TXT','*.txt')
        '''
        root=tk.Tk() # creates pop-up window
        root.update() # necessary to close tk dialog after askopenfilename is finished
        # tk dialog asks for a single station file
        full_path = tk.filedialog.askopenfilename(title = txtmess, filetypes=[ ftypes] )
        root.destroy() # closes pop up window
        return full_path

    def choose_message():
        # choose existing message (.txt file)
        root=tk.Tk() # creates pop-up window
        root.update() # necessary to close tk dialog after askopenfilename is finished
        # tk dialog asks for a single station file
        full_path = tk.filedialog.askopenfilename(title = 'Choose message file', filetypes=[ ('TXT','*.txt')] )
        root.destroy() # closes pop up window
        return full_path
    
    # Functions to enable/disable relevant checkboxes depending on radiobutton choice
    def Assignopts():
        ''' Display relevant choices for team assignment notification/cyc card/ short team recruiting '''
        recruitcheck.config(state=tk.NORMAL)  
        extraentry.config(state=tk.DISABLED)
        extraname.set('n/a')
        messfile.set('parent_team_assignment.txt')
        transmessfile.set('parent_team_transfer.txt')
        emailtitle.set('Fall $SPORT for $FIRST')
    def Recruitopts():
        ''' Display relevant choices for specific player recruiting'''
        recruitcheck.config(state=tk.NORMAL)
        extraentry.config(state=tk.DISABLED)
        messfile.set('player_recruiting.txt')
        transmessfile.set('n/a')
        extraname.set('n/a')
        emailtitle.set('Cabrini-Soulard sports for $FIRST this fall?')    
    def Missingopts():
        ''' Display relevant choices for ask parent for missing uniforms '''
        recruitcheck.config(state=tk.DISABLED)        
        extraentry.config(state=tk.NORMAL)
        messfile.set('finish_me.txt')
        transmessfile.set('n/a')
        extraname.set('Missing uni file name')
        extravar.set('missing_uni.csv')
        # TODO look up most recent uni file?
        emailtitle.set("Please return $FIRST's $SPORT uniform!")
    def Schedopts():
        ''' Display relevant choices for sending schedules (game and practice) to parents '''
        recruitcheck.config(state=tk.DISABLED)
        # Used here for name of master file schedule
        extraentry.config(state=tk.NORMAL)
        messfile.set('parent_game_schedule.txt')
        transmessfile.set('n/a')
        extraname.set('Game schedule file')
        extravar.set('Cabrini_2017_schedule.csv')
        emailtitle.set("Game schedule for Cabrini $GRADERANGE $GENDER $SPORT")
    def Cardopts():
        ''' Display relevant choices for asking parent for missing CYC cards '''
        recruitcheck.config(state=tk.DISABLED)
        # Used here for name of master file schedule
        extraentry.config(state=tk.DISABLED)
        messfile.set('CYCcard_needed.txt')
        transmessfile.set('n/a')
        extraname.set('')
        extravar.set('')
        emailtitle.set("CYC card needed for $FIRST")
    def Otheropts():
        ''' Display relevant choices for other generic message to parents '''
        recruitcheck.config(state=tk.DISABLED)
        # Used here for name of master file schedule
        extraentry.config(state=tk.NORMAL)
        messfile.set('temp_message.txt')
        transmessfile.set('n/a')
        extraname.set('')
        extravar.set('')
        emailtitle.set("Message from Cabrini Sponsors Club")
    def Allopts():
        ''' Display relevant choices for generic message to all sports parents '''
        recruitcheck.config(state=tk.DISABLED)
        extraentry.config(state=tk.NORMAL)
        messfile.set('temp_message.txt')
        transmessfile.set('n/a')
        extraname.set('')
        extravar.set('')
        emailtitle.set("Message from Cabrini Sponsors Club")
    # E-mail title and message file name
    rownum=0
    tk.Label(messageframe, text='Title for e-mail').grid(row=rownum, column=0)
    titleentry=tk.Entry(messageframe, textvariable=emailtitle)
    titleentry.config(width=50)
    titleentry.grid(row=rownum, column=1)
    rownum+=1
    tk.Label(messageframe, text='messagefile').grid(row=rownum, column=0)
    messentry=tk.Entry(messageframe, textvariable=messfile)
    messentry.config(width=50)
    messentry.grid(row=rownum, column=1)
    rownum+=1
    tk.Label(messageframe, text='Transfer messagefile').grid(row=rownum, column=0)
    transmessentry=tk.Entry(messageframe, textvariable=transmessfile)
    transmessentry.config(width=50)
    transmessentry.grid(row=rownum, column=1)
    rownum+=1    
    # Choose counts, deriv, both or peaks plot
    tk.Radiobutton(messageframe, text='Team assignment', value='Assign', variable = mtype, command=Assignopts).grid(row=rownum, column=0)
    tk.Radiobutton(messageframe, text='Recruit missing', value='Recruit', variable = mtype, command=Recruitopts).grid(row=rownum, column=1)
    tk.Radiobutton(messageframe, text='Missing uni', value='Missing', variable = mtype, command=Missingopts).grid(row=rownum, column=2)
    tk.Radiobutton(messageframe, text='Send schedule', value='Schedule', variable = mtype, command=Schedopts).grid(row=rownum, column=3)
    rownum+=1
    tk.Radiobutton(messageframe, text='Ask for cards', value='Cards', variable = mtype, command=Cardopts).grid(row=rownum, column=1)
    tk.Radiobutton(messageframe, text='Other team message', value='Other', variable = mtype, command=Otheropts).grid(row=rownum, column=1)
    tk.Radiobutton(messageframe, text='All sport parents', value='All', variable = mtype, command=Allopts).grid(row=rownum, column=2)
    rownum+=1    
    tk.Label(messageframe, text=extraname.get()).grid(row=rownum, column=0)
    extraentry=tk.Entry(messageframe, textvariable=extravar)
    extraentry.grid(row=rownum, column=1)
    # Extra file chooser button 
    # button arg includes file type extension .. get from messfile
    try:
        ft = extraname.get().split('.')[-1]
        ftypes =("%s" %ft.upper(), "*.%s" %ft)
    except:
        ftypes =("CSV" , "*.*") # default to all files 
    # TODO fix extra file chooser
    d=tk.Button(messageframe, text='Choose file', command=chooseFile('Choose extra file', ftypes) )
    d.grid(row=rownum, column=2)

    recruitcheck=tk.Checkbutton(messageframe, variable=recruitbool, text='Recruit more players for short teams?')
    recruitcheck.grid(row=rownum, column=3) # can't do immediate grid or nonetype is returned
    rownum+=1    
    messageframe.grid(row=0, column=0)
    # Specific team selector section using checkboxes
    teamframe=tk.LabelFrame(root, text='Team selector')
    teamdict=shortnamedict(teams)
    teamlist=[] # list of tk bools for each team
    # Make set of bool/int variables for each team
    for i, val in enumerate(teamdict):
        teamlist.append(tk.IntVar())
        if '#' not in val:
            teamlist[i].set(1) # Cabrini teams checked by default
        else:
            teamlist[i].set(0) # transfer team

    # make checkbuttons for each team
    for i, val in enumerate(teamdict):
        thisrow=i%5+1+rownum # three column setup
        thiscol=i//5
        thisname=teamdict.get(val,'')
        tk.Checkbutton(teamframe, text=thisname, variable=teamlist[i]).grid(row=thisrow, column=thiscol)
    rownum+=math.ceil(len(teamlist)/5)+2
    # Decision buttons bottom row
    def chooseall(event):
        ''' Select all teams '''
        for i, val in enumerate(teamdict):
            teamlist[i].set(1)
    def clearall(event):
        ''' deselect all teams '''
        for i, val in enumerate(teamdict):
            teamlist[i].set(0)
    def abort(event):
        choice.set('abort')        
        root.destroy()  
    def test(event):
        choice.set('test')        
        root.destroy()  
    def KCtest(event):
        choice.set('KCtest')        
        root.destroy()  
    def send(event):
        choice.set('send')        
        root.destroy()  
    rownum+=1
    d=tk.Button(teamframe, text='All teams')
    d.bind('<Button-1>', chooseall)
    d.grid(row=rownum, column=0)
    
    d=tk.Button(teamframe, text='Clear teams')
    d.bind('<Button-1>', clearall)
    d.grid(row=rownum, column=1)
    teamframe.grid(row=1, column=0)
    choiceframe=tk.LabelFrame(root)
    d=tk.Button(choiceframe, text='Abort')
    d.bind('<Button-1>', abort)
    d.grid(row=rownum, column=2)

    d=tk.Button(choiceframe, text='Test')
    d.bind('<Button-1>', test)
    d.grid(row=rownum, column=3)

    d=tk.Button(choiceframe, text='KCtest')
    d.bind('<Button-1>', KCtest)
    d.grid(row=rownum, column=4)
    
    d=tk.Button(choiceframe, text='Send')
    d.bind('<Button-1>', send)
    d.grid(row=rownum, column=5)
    choiceframe.grid(row=2, column=0)

    root.mainloop()
#%%
    mychoice=choice.get()
    
    if mychoice!='abort':
        kwargs={}
        if mychoice=='KCtest':
            # this is a true send test but only to me
            kwargs.update({'KCtest':True})
            mychoice='send'
        kwargs.update({'choice':mychoice})  # test or send
        emailtitle=emailtitle.get()
        messagefile='messages\\'+messfile.get()
        
        # Handle selection of team subsets
        selteams=[]
        for i, val in enumerate(teamdict):
            if teamlist[i].get()==1:
                selteams.append(val)
        # Filter teams based on checkbox input
        teams=teams[teams['Team'].isin(selteams)]
        # drop duplicates in case of co-ed team (m and f entries)
        teams=teams.drop_duplicates(['Team','Sport'])
        # Now deal with the different types of messages
#%%
        if mtype.get()=='Schedule':
            # Send practice and game schedules
            try:
                sched=pd.read_csv(extravar.get())
            except:
                print('Problem opening schedule and other required files for sending game schedules')
                fname=filedialog.askopenfilename(title='Select schedule file.')
                sched=pd.read_csv(fname)
            # fields=pd.read_excel(cnf._INPUT_DIR+'\\Teams_coaches.xlsx', sheetname='Fields')
            fields=pd.read_csv(cnf._INPUT_DIR+'\\fields.csv')
            Mastersignups = pd.read_csv(cnf._INPUT_DIR+'\\master_signups.csv', encoding='cp437')
            #coaches=pd.read_excel('Teams_coaches.xlsx', sheetname='Coaches')
            coaches=pd.read_csv(cnf._INPUT_DIR+'\\coaches.csv')
            # INTERNAL TESTING 
            # Mastersignups=Mastersignups[Mastersignups['Last']=='Croat']
            famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
            with open(messagefile, 'r') as file:
                blankmess=file.read()
            # open and send master CYC schedule 
            sendschedule(teams, sched, fields, Mastersignups, coaches, year, famcontact, emailtitle, blankmess, **kwargs)
        if mtype.get()=='Recruit':
            try:
                famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
            except:
                print('Problem loading family contacts')
            try: # Recruits stored in CSV 
                Recruits=pd.read_csv(cnf._OUTPUT_DIR+'\\%s%s_recruits.csv' %(season, year))
                print('Loaded possible recruits from csv file')
            except:
                fname=filedialog.askopenfilename(title='Select recruits file.')
                if fname.endswith('.csv'): # final move is query for file
                    Recruits=pd.read_csv(fname)
                else:
                    print('Recruits file needed in csv format.')
                return
            emailrecruits(Recruits, famcontact, emailtitle, messagefile, **kwargs)
            
        if mtype.get()=='Assign':
            # Notify parents needs teams, mastersignups, famcontacts
            if recruitbool.get():
                kwargs.update({'recruit':True})
            try:
                Mastersignups = pd.read_csv(cnf._INPUT_DIR+'\\master_signups.csv', encoding='cp437')
                #coaches=pd.read_excel(cnf._INPUT_DIR+'\\Teams_coaches.xlsx', sheetname='Coaches')
                coaches=pd.read_csv(cnf._INPUT_DIR+'\\coaches.csv', encoding='cp437')
                # INTERNAL TESTING 
                # Mastersignups=Mastersignups[Mastersignups['Last']=='Croat']
                famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
                with open(messagefile, 'r') as file:
                    blankmess=file.read()
                tranmessagefile='messages\\'+transmessfile.get()
                with open(tranmessagefile, 'r') as file:
                    blanktransmess=file.read()
            except:
                print('Problem loading mastersignups, famcontacts')
                return
            notifyfamilies(teams, Mastersignups, coaches, year, famcontact, emailtitle, blankmess, blanktransmess, **kwargs)
        if mtype.get()=='Unis':
            try:
                missing=pd.read_csv(messfile.get(), encoding='cp437')
                oldteams=pd.read_excel(cnf._INPUT_DIR+'\\Teams_coaches.xlsx', sheetname='Oldteams') # loads all old teams in list
                kwargs.update({'oldteams':oldteams,'missing':missing})                
            except:
                print('Problem loading missingunis, oldteams')
                return
            # TODO Finish ask for missing uniforms script
            askforunis(teams, Mastersignups, year, famcontact, emailtitle, blankmess, **kwargs)
        if mtype.get()=='Cards':
            try:
                famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
                Mastersignups = pd.read_csv(cnf._INPUT_DIR+'\\master_signups.csv', encoding='cp437')
                with open(messagefile, 'r') as file:
                    blankmess=file.read()
            except:
                print('Problem loading famcontacts, mastersignups, or blank message')
                return
            # TODO Finish ask for missing uniforms script
            askforcards(teams, Mastersignups, year, famcontact, emailtitle, blankmess, **kwargs)
        if mtype.get()=='Other':
            try:
                famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
                Mastersignups = pd.read_csv(cnf._INPUT_DIR+'\\master_signups.csv', encoding='cp437')
                coaches=pd.read_excel(cnf._INPUT_DIR+'\\Teams_coaches.xlsx', sheetname='Coaches')
                with open(messagefile, 'r') as file:
                    blankmess=file.read()
            except:
                print('Problem loading mastersignups, coaches, ')
                return
            # TODO Finish ask for missing uniforms script
            sendteammessage(teams, year, Mastersignups, famcontact, coaches, emailtitle, blankmess, **kwargs)
        if mtype.get()=='All':
            try:
                famcontact= pd.read_csv(cnf._INPUT_DIR+'\\family_contact.csv', encoding='cp437')
                Mastersignups = pd.read_csv(cnf._INPUT_DIR+'\\master_signups.csv', encoding='cp437')
                #coaches=pd.read_excel(cnf._INPUT_DIR+'\\Teams_coaches.xlsx', sheetname='Coaches')
                coaches=pd.read_excel(cnf._INPUT_DIR+'\\coaches.csv')
                with open(messagefile, 'r') as file:
                    blankmess=file.read()
            except:
                print('Problem loading mastersignups, coaches, ')
                return
            # TODO Finish ask for missing uniforms script
            sendallmessage(season, year, Mastersignups, famcontact, coaches, emailtitle, blankmess, **kwargs)
    return

''' TESTING of notifyfamilies
[sport, team, graderange, coachinfo, playerlist] =cabteamlist[6]      i=6
index=thisteam.index[0]
row=thisteam.loc[index] 
'''	

def askforcards(teams, Mastersignups, year, famcontact, emailtitle, blankmess, **kwargs):
    ''' Notifying players that need cards and ask for them via custom e-mail (one per player) 
    kwargs:
        choice - 'send' or 'test'
    '''
    choice=kwargs.get('choice','test')
    if choice=='send':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('parent_email_log.txt','w', encoding='utf-8')
    
    # this years signups only (later match for sport and team)
    Mastersignups=Mastersignups[Mastersignups['Year']==year]
    # drop non-CYC K and 1 level teams
    teams=teams[teams['Grade']>=2]
    # Make list of sport/team/school/graderange
    teamlist=[]
    for index, row in teams.iterrows():
        # get school 
        if '#' not in teams.loc[index]['Team']:
            school='Cabrini'
        else:
            school=teams.loc[index]['Team'].split('#')[0]
        # Get sport, team, graderange, coach info (first/last/e-mail), playerlist
        teamlist.append([teams.loc[index]['Sport'], teams.loc[index]['Team'], school,
            teams.loc[index]['Graderange']])
    # dict. with each team and its players
    cards=findcards() # find all player cards
    if not cards: # terminate if no cards are found (path error?)
        print("Error opening CYC card image database")
        return
    # Drop all player nums found in cards 
    cardslist=list(cards.keys())
    cardslist=[i for i in cardslist if '-' not in i]
    cardslist=[int(i) for i in cardslist]
    # Only keep signups without cards
    Mastersignups=Mastersignups[~Mastersignups['Plakey'].isin(cardslist)]
    CYCSUs=pd.DataFrame()
    for i, [sport, team, school, graderange] in enumerate(teamlist):
        CYCSUs=CYCSUs.append(Mastersignups[(Mastersignups['Sport']==sport) & (Mastersignups['Team']==team)])
    # only one notice needed per player
    CYCSUs=CYCSUs.drop_duplicates('Plakey')
    CYCSUs=pd.merge(CYCSUs, famcontact, on='Famkey' , how='left', suffixes =('','_2'))

    for index, row in CYCSUs.iterrows():
        # Replace first name in e-mail title (default e-mail title is fall $SPORT for $FIRST)
        thistitle=emailtitle.replace('$FIRST', row.First)
        thistitle=thistitle.replace('$LAST', row.Last)
        # custom message for individual player on this team
        thismess=blankmess.replace('$FIRST', row.First)
        thismess=thismess.replace('$LAST', row.Last)
        recipients=getemailadds(row)
        # Create custom email message (can have multiple sports in df)
        if choice=='send':
            # add From/To/Subject to actual e-mail
            thisemail='From: Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>\nTo: '
            thisemail+=', '.join(recipients)+'\nSubject: '+thistitle+'\n'
            thisemail+=thismess
            thisemail=thisemail.encode('utf-8')
            for i,addr in enumerate(recipients): # Send message to each valid recipient in list
                try:
                    smtpObj.sendmail('sfcasponsorsclub@gmail.com', addr, thisemail)
                    print ('Message sent to ', addr)
                except:
                    print('Message to ', addr, ' failed.')
            if not recipients:
                print('No email address for ', row.First, row.Last)
        else: # Testing mode ... just write to log w/o e-mail header and such
            logfile.write(thistitle+'\n')
            logfile.write(thismess+'\n')
    # close log file (if testing mode)
    if choice!='send':
        logfile.close()
    else:
        pass
        # TODO fix this attempted close
        # smtpObj.quit() # close SMTP connection
    return

def sendallmessage(season, year, Mastersignups, famcontact, coaches, emailtitle, blankmess, **kwargs):
    ''' Top level messaging function for notifying families of team assignment/ CYC card
    + optional short-team-player-recruiting
    via custom e-mail; one per player
    currently not including SMS 
    kwargs:
        choice - 'send' or 'test'
    '''
    choice=kwargs.get('choice','test')
    if choice=='send':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('allparent_email_log.txt','w', encoding='utf-8')
    # Get all email addresses from recent parents (default last 3 seasons)
    recipients=makeemaillist(Mastersignups, famcontact, season, year, SMS=False)
    # add all coach emails
    coachemails=np.ndarray.tolist(coaches.Email.unique())
    coachemails=[i for i in coachemails if '@' in i]
    recipients.extend(coachemails)
    recipients=set(recipients)
    recipients=list(recipients) # unique only

    # Create custom email message (can have multiple sports in df)
    if choice=='send':
        if 'KCtest' in kwargs: # internal only send test
            recipients=['kcroat@gmail.com','tkc@wustl.edu']
        msg=MIMEText(blankmess,'plain')
        # msg = MIMEMultipart('alternative') # message container
        msg['Subject'] = emailtitle
        msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'                
        msg['To'] = 'Cabrini Sports Parents <sfcasponsorsclub@gmail.com>'  
        msg['Bcc']=','.join(recipients)  # single e-mail or list
        # Simultaneous send to all in recipient list
        smtpObj.sendmail('sfcasponsorsclub@gmail.com', recipients, msg.as_string())
        print ('Message sent to ', ','.join(recipients))

    else: # Testing mode
        tempstr='Test message to: '+', '.join(recipients)
        logfile.write(tempstr+'\n')
        logfile.write(blankmess)
    # close log file (if testing mode)
    if choice!='send':
        logfile.close()
    else:
        pass
        # smtpObj.quit() # close SMTP connection
    return

def sendteammessage(teams, year, Mastersignups, famcontact, coaches, emailtitle, blankmess, **kwargs):
    ''' Top level messaging function for notifying families of team assignment/ CYC card
    + optional short-team-player-recruiting
    via custom e-mail; one per player
    currently not including SMS 
    kwargs:
        choice - 'send' or 'test'
        recruit - T or F -- add recruiting statement for short teams
        mformat -  not really yet used ... just sending as text not html 
    '''
    choice=kwargs.get('choice','test')
    if choice=='send':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('team_email_log.txt','w', encoding='utf-8')
    
    # this years signups only (later match for sport and team)
    Mastersignups=Mastersignups[Mastersignups['Year']==year]
    # drop extra co-ed K or other entries
    teams=teams.drop_duplicates(['Team'])

    myteams=pd.merge(teams, coaches, on='Coach ID', how='left', suffixes=('','_2'))
    # Make list of sport/team/school/graderange/coachinfo/playerlist
    teamlist=[]
    for index, row in myteams.iterrows():
        # get school 
        if '#' not in myteams.loc[index]['Team']:
            school='Cabrini'
            try:
                coachinfo=myteams.loc[index]['Fname']+' '+ myteams.loc[index]['Lname']+' ('+myteams.loc[index]['Email']+')'
            except:
                coachinfo=''
        else:
            school=myteams.loc[index]['Team'].split('#')[0]
            coachinfo=''
        # Get sport, team, graderange, coach info (first/last/e-mail), playerlist
        teamlist.append([row.Sport, row.Team, school,
            gradetostring(row.Graderange), coachinfo, row.Playerlist])
    # Separate notification for each signup  is OK
    for i, [sport, team, school, graderange, coach, playerlist] in enumerate(teamlist):
        thisteam=Mastersignups[(Mastersignups['Sport']==sport) & (Mastersignups['Team']==team)]
        thisteam=pd.merge(thisteam, famcontact, on='Famkey' , how='left', suffixes =('','_2'))
        # Cabrini team base message
        thisteammess=blankmess
        thistitle=emailtitle
        # Make team-specific replacements in message text and e-mail title
        for j, col in enumerate(['$SPORT', '$TEAMNAME', '$SCHOOL', '$GRADERANGE', '$COACH', '$PLAYERLIST']):
            thisteammess=thisteammess.replace(col, textwrap.fill(teamlist[i][j], width=100))
            thistitle=thistitle.replace(col, teamlist[i][j])
        # get coach emails
        recipients=getcoachemails(team, teams, coaches, **{'asst':True})
        # Now get all unique team email addresses (single message to coach and team)
        recipients=getallteamemails(thisteam, recipients)
        
        # Create custom email message (can have multiple sports in df)
        if choice=='send':
            msg=MIMEText(thisteammess,'plain')
            # msg = MIMEMultipart('alternative') # message container
            msg['Subject'] = emailtitle
            msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'                
            # part2=MIMEText(thismess_html,'alternate')                
            msg['To']=','.join(recipients)  # single e-mail or list
            # Simultaneous send to all in recipient list
            smtpObj.sendmail('sfcasponsorsclub@gmail.com', recipients, msg.as_string())
            print ('Message sent to ', ','.join(recipients))
            if not recipients:
                print('No email addresses for team', team)
        else: # Testing mode ... just write to log w/o e-mail header and such
            logfile.write(thistitle+'\n')
            logfile.write(thisteammess+'\n')
    # close log file (if testing mode)
    if choice!='send':
        logfile.close()
    else:
        pass
        # TODO fix this attempted close
        # smtpObj.quit() # close SMTP connection
    return

def makeemaillist(Mastersignups, famcontact, thisseason, thisyear, SMS=False):
    '''Return active and inactive families (mainly for e-mail contact list 
    active if has player in 3 prior sport-seasons (includes current )
    
    '''
    # TODO generalize to n prior sports seasons
    thisyearSU=Mastersignups[Mastersignups['Year']==thisyear] # take all form 
    lastyearSU=Mastersignups[Mastersignups['Year']==(thisyear-1)]
    lastyearSU=lastyearSU[lastyearSU['Grade']!=8]
    seasonlist=['Fall', 'Winter', 'Spring']    
    pos=seasonlist.index(thisseason)
    activeseasons=seasonlist[pos:]
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    activesports=[]
    for i, season in enumerate(activeseasons):
        sportlist=sportsdict.get(season)
        activesports.extend(sportlist)
    lastyearSU=lastyearSU[lastyearSU['Sport'].isin(activesports)] # last year's signups incl.
    allSU=pd.concat([thisyearSU,lastyearSU],ignore_index=True)
    activefams=allSU.Famkey.unique()
    emaillist=[]
    match=famcontact[famcontact['Famkey'].isin(activefams)]
    emails=match.Email1.unique()
    emails=np.ndarray.tolist(emails)
    emaillist.extend(emails)
    emails=match.Email2.unique()
    emails=np.ndarray.tolist(emails)
    emaillist.extend(emails)
    emails=match.Email3.unique()
    emails=np.ndarray.tolist(emails)
    emaillist.extend(emails)
    emaillist=set(emaillist) # eliminate duplicates
    emaillist=list(emaillist)
    emaillist=[x for x in emaillist if str(x) != 'nan'] # remove nan
    emaillist=[x for x in emaillist if str(x) != 'none'] # remove nan
    if not SMS: # Drop SMS
        emaillist=[x for x in emaillist if not str(x).startswith('314')]
        emaillist=[x for x in emaillist if not str(x).startswith('1314')]
    return emaillist

def getcabsch(sched, teams, coaches, fields, **kwargs):
    ''' Return Cabrini containing subset of teams from master schedule
    manual save... can then feed csv to sendschedule
    kwargs:
        sport -- Soccer, VB or whatever
        div--- division 5G
        school  - Cabrini 
    #TESTING sched=fullsched.copy()
    '''
    if 'school' in kwargs:
        if kwargs.get('school','')=='Cabrini':
            # drop transfer teams w/ #
            teams=teams[~teams['Team'].str.contains('#')]
    if 'sport' in kwargs:
        sport=kwargs.get('sport','')
        teams=teams[teams['Sport']==sport]
    if 'div' in kwargs:
        div=kwargs.get('div','')
        grade=int(div[0])
        if div[1].upper()=='G':
            gender='f'
        elif div[1].upper()=='B':
            gender='m'
        teams=teams[(teams['Grade']==grade) & (teams['Gender']==gender)]
    # perform any team filtering 
    sched=sched.rename(columns={'Start':'Time','Venue':'Location','Sched Name':'Division', 
                                'Visitor':'Away'})
    teamdict=findschteams(sched, teams, coaches)
    cabsched=pd.DataFrame()
    for key, [div, schname] in teamdict.items():
        match=sched[(sched['Division'].str.startswith(div)) & ((sched['Home'].str.contains(schname)) | (sched['Away'].str.contains(schname)))]
        if 'Cabrini' not in schname:
            newname=schname.split('/')[0]+'-Cabrini'
            match['Home']=match['Home'].str.replace(schname,newname)
            match['Away']=match['Away'].str.replace(schname,newname)
        # add team column via assign
        match=match.assign(Team=key)
        # Why isn't team col being copied?
        cabsched=cabsched.append(match, ignore_index=True)
        print(len(match),' games for team', str(schname))
    cabsched['Home']=cabsched['Home'].str.replace('St Frances','')
    cabsched['Away']=cabsched['Away'].str.replace('St Frances','')
    cabsched=cabsched.sort_values(['Division','Date','Time'])
    # now sort 
    myCols=['Date','Time','Day','Location','Division','Home','Away','Team']
    # add col if missing from CYC schedule
    for miss in [i for i in myCols if i not in cabsched.columns]:
        print(miss,'column missing from full CYC schedule')
        cabsched[miss]=''
    cabsched=cabsched[myCols] # set in above preferred order
    flist=np.ndarray.tolist(cabsched.Location.unique())
    missing=[s for s in flist if s not in fields['Location'].tolist()]
    if len(missing)>0:
        print('Address missing from fields table:',','.join(missing))
    # convert to desired string format here (write-read cycle makes it a string anyway)
    # cabsched.Time=cabsched.Time.apply(lambda x:datetime.time.strftime(x, format='%I:%M %p'))
    #cabsched['Date']=cabsched['Date'].dt.strftime(date_format='%d-%b-%y')
    return cabsched

def detectschchange(sched1, sched2):
    '''Compare two schedule versions and return unique rows (changed games)

    '''
    # Convert both to datetime/timestamps if in string format (probably %m/%d/%Y)
    if type(sched1.iloc[0]['Date'])==str:
        try:
            sched1['Date']=sched1['Date'].apply(lambda x:datetime.datetime.strptime(x, "%m/%d/%Y"))
        except:
            print('Problem with string to datetime conversion for', sched1.iloc[0]['Date'])
    if type(sched2.iloc[0]['Date'])==str:
        try:
            sched2['Date']=sched2['Date'].apply(lambda x:datetime.datetime.strptime(x, "%m/%d/%Y"))
        except:
            print('Problem with string to datetime conversion for', sched2.iloc[0]['Date'])
    if type(sched2.iloc[0]['Time'])==str:
        try:
            # convert to timestamp
            sched2['Time']=sched2['Time'].apply(lambda x:datetime.datetime.strptime(x, "%H:%M:%S").time())
            # convert to datetime.time
            sched2['Time']=sched2['Time'].apply(lambda x:datetime.time(x))
        except:
            print('Problem with string to datetime conversion for', sched2.iloc[0]['Date'])
    # all columns by default, false drops both duplicates leaving unique rows
    bothsch=pd.concat([sched1,sched2])
    alteredrows=bothsch.drop_duplicates(keep=False)
    alteredrows=alteredrows.sort_values(['Date','Time','Division'])
    return alteredrows

def makefieldtable(df, fields):
    ''' Make separate table of field addresses for all fields in 
    given team's schedule  (called by sendschedule)'''
    venues=np.ndarray.tolist(df.Location.unique())
    venues=[s.strip() for s in venues]
    ft=pd.DataFrame()
    ft['Location']=venues
    ft=pd.merge(ft, fields, how='left', on=['Location'])
    ft=ft[['Location','Address']]
    return ft

def notifyfamilies(teams, Mastersignups, coaches, year, famcontact, emailtitle, blankmess, blanktransmess, **kwargs):
    ''' Top level messaging function for notifying families of team assignment/ CYC card
    + optional short-team-player-recruiting
    via custom e-mail; one per player
    currently not including SMS 
    kwargs:
        choice - 'send' or 'test'
        recruit - T or F -- add recruiting statement for short teams
    '''
    choice=kwargs.get('choice','test')
    if choice=='send':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('parent_email_log.txt','w', encoding='utf-8')
    
    # this years signups only (later match for sport and team)
    Mastersignups=Mastersignups[Mastersignups['Year']==year]

    myteams=pd.merge(teams, coaches, on='Coach ID', how='left', suffixes=('','_2'))
    # Make list of sport/team/school/graderange/coachinfo/playerlist
    teamlist=[]
    for index, row in myteams.iterrows():
        # get school 
        if '#' not in myteams.loc[index]['Team']:
            school='Cabrini'
            coachinfo=myteams.loc[index]['Fname']+' '+ myteams.loc[index]['Lname']+' ('+myteams.loc[index]['Email']+')'
        else:
            school=myteams.loc[index]['Team'].split('#')[0]
            coachinfo=''
        # Get sport, team, graderange, coach info (first/last/e-mail), playerlist
        teamlist.append([row.Sport.lower(), row.Team, school, gradetostring(row.Graderange),
            coachinfo, row.Playerlist])
    # dict. with each team and its players
    cards=findcards() # find all player cards
    if not cards: # terminate if no cards are found (path error?)
        print("Error opening CYC card image database")
        return
    # Separate notification for each signup  is OK
    for i, [sport, team, school, graderange, coachinfo, playerlist] in enumerate(teamlist):
        thisteam=Mastersignups[(Mastersignups['Sport']==sport) & (Mastersignups['Team']==team)]
        thisteam=pd.merge(thisteam, famcontact, on='Famkey' , how='left', suffixes =('','_2'))
        if '#' not in team:
            # Cabrini team base message
            thisteammess=blankmess
        else: # base message for transferred players
            thisteammess=blanktransmess
        thisteamtitle=emailtitle
        # Make team-specific replacements
        for j, col in enumerate(['$SPORT', '$TEAMNAME', '$SCHOOL', '$GRADERANGE', '$COACH', '$PLAYERLIST']):
            thisteammess=thisteammess.replace(col, textwrap.fill(teamlist[i][j], width=100))
            thisteamtitle=thisteamtitle.replace(col, teamlist[i][j])
        # Check if Cabrini team is short of players (max grade, sport, numplayers)
        try:
            recmess=makerecmess(team, thisteam['Grade'].max(), sport, len(thisteam))
        except:
            recmess='' # handles empty teams during testing
        # Either blank inserted or generic needs more players request (same for whole team)
        thisteammess=thisteammess.replace('$RECRUIT','\n'+recmess)

        for index, row in thisteam.iterrows():
            # Replace first name in e-mail title (default e-mail title is fall $SPORT for $FIRST)
            thistitle=thisteamtitle.replace('$FIRST', row.First)
            thistitle=thistitle.replace('$SPORT', row.Sport)
            
            # Check for each players CYC card if necessary (also for older transfer teams)
            thiscardmess=makecardmess(row, cards)
            # custom message for individual player on this team
            thismess=thisteammess.replace('$FIRST', row.First)
            thismess=thismess.replace('$LAST', row.Last)
            # message is blank if on file or not required and 
            thismess=thismess.replace('$CYCCARD', '\n'+thiscardmess)
            recipients=getemailadds(row)
            # Create custom email message (can have multiple sports in df)
            if choice=='send':
                # add From/To/Subject to actual e-mail
                msg=MIMEText(blankmess,'plain')
                # msg = MIMEMultipart('alternative') # message container
                msg['Subject'] = thistitle
                msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'                
                msg['To'] = 'Cabrini Sports Parents <sfcasponsorsclub@gmail.com>'  
                msg['Bcc']=','.join(recipients)  # single e-mail or list
                # Simultaneous send to all in recipient list
                smtpObj.sendmail('sfcasponsorsclub@gmail.com', recipients, msg.as_string())
                print ('Message sent to ', ','.join(recipients))

                thisemail='From: Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>\nTo: '
                thisemail+=', '.join(recipients)+'\nSubject: '+thistitle+'\n'
                thisemail+=thismess
                thisemail=thisemail.encode('utf-8')
                for i,addr in enumerate(recipients): # Send message to each valid recipient in list
                    try:
                        smtpObj.sendmail('sfcasponsorsclub@gmail.com', addr, thisemail)
                        print ('Message sent to ', addr)
                    except:
                        print('Message to ', addr, ' failed.')
                if not recipients:
                    print('No email address for ', row.First, row.Last)
            else: # Testing mode ... just write to log w/o e-mail header and such
                logfile.write(thistitle+'\n')
                logfile.write(thismess+'\n')
    # close log file (if testing mode)
    if choice!='send':
        logfile.close()
    else:
        pass
        # TODO fix this attempted close
        # smtpObj.quit() # close SMTP connection
    return

def makecardmess(row, cards):
    ''' Determine if card is needed and add generic message to that effect (called by emailparent_tk, notifyparent)
    row is Series
    '''
    cmess=("$FIRST $LAST needs a CYC ID card to play on this team and we do not have one in our files."
    "If your child already has this ID card, please take a picture of it and e-mail to sfcasponsorsclub@gmail.com."
    "If you don't have one, you can get one online at: https://idcards.cycstl.net/ or at uniform night.  "
    "For this you need: 1) picture of the child 2) child's birth certificate (or birth document) and 3) $5 fee")
    if str(row.Plakey) in cards:
        return '' # already on file
    # Now handle teams that don't need CYC cards (generally K or 1st)
    if '-' not in row.Team: # non-CYC level teams and transfer teams
        if '#' not in row.Team: # non-CYC cabrini team
            return '' # junior team doesn't require card
        else: # determine grade level for transfer team
            tempstr=row.Team
            tempstr=tempstr.split('#')[1][0:1]
            tempstr=tempstr.replace('K','0')
            try:
                grade=int(tempstr)
                if grade<2: # judge dowd or junior transfer team
                    return ''
            except:
                print("couldn't determine grade for transfer team")
                return ''
    # all remaining players need a card 
    cmess=cmess.replace('$FIRST',row.First)
    cmess=cmess.replace('$LAST',row.Last)
    cmess=textwrap.fill(cmess, width=100)
    return cmess

'''TESTING
makerecmess('teamname', 2, 'T-ball', 14)  
textwrap.fill(recmess, width=80)
'''
def makerecmess(team, grade, sport, numplayers):
    ''' Figure out if team is short of players (based on grade level, sport, Cabteam or not)
    '''
    recmess=('This team could use more players. If you know anyone who is interested,'  
             'please inform us at sfcasponsorsclub@gmail.com.')
    recmess=textwrap.fill(recmess, width=100)
    if '#' in team: # no recruiting for transfer teams
        return ''
    if grade=='K':
        grade=0
    else:
        grade=int(grade)
    if sport=='VB': # 8 for all grades
        if numplayers<8:
            return recmess
    if sport=='Soccer':
        if grade>=5: # 11v11 so need 16
            if numplayers<16:
                return recmess
        elif grade<=4 and grade>=2:    # 8v8 from 2nd to 4th so 12 is OK
            if numplayers<13:
                return recmess
        elif grade==1: # 7v7 so 11 is OK
            if numplayers<12:
                return recmess
        else: # k is 6v6 so 10 is OK
            if numplayers<11:
                return recmess
    if sport=='Basketball': # 5v5 for all grades so 10 is good
        if numplayers<11:
            return recmess
    if sport=='T-ball': # 9v9 ish so 13 is good
        if numplayers<14:
            return recmess        
    if sport=='Baseball': # 9v9 ish so 13 is good
        if numplayers<14:
            return recmess            
    if sport=='Softball': # 9v9 ish so 13 is good
        if numplayers<14:
            return recmess            
    return ''
    
def emailcoach_tk(teams, coaches, gdrivedict):
    ''' tk interface for e-mails to team coaches
    some required datasets (players, famcontact, mastersignups) are directly loaded depending on choice
    message types (mtypes) are: 
        unis - send summary of missing uniforms to team coaches
        contacts - send contacts and current google drive link
        bills -  send summary of outstanding bills 
    '''
    root = tk.Tk()
    root.title('Send e-mail to coaches')

    unifilename=tk.StringVar()
    try:
        unifiles=glob.glob('missingunilist*') # find most recent uniform file name
        if len(unifiles)>1:
            unifile=findrecentfile(unifiles) # return single most recent file
        else:
            unifile=unifiles[0]
        # find most recent missing uni file name
        unifilename.set(unifile)
    except: # handle path error
        unifilename.set('missingunilist.csv')
        
    billname=tk.StringVar() # file
    try:
        billfiles=glob.glob('billlist*')
        if len(billfiles)>1:
            billfile=findrecentfile(billfiles) # return single most recent file
        else:
            billfile=billfiles[0]        
        # find most recent billlist file name
        billname.set(billfile)
    except:
        billname.set('billist.csv')
        
    asstbool=tk.BooleanVar()  # optional labelling of elements
    emailtitle=tk.StringVar()  # e-mail title 
    mtype=tk.StringVar()  # coach message type
    messfile=tk.StringVar() # text of e-mail message
    choice=tk.StringVar()  # test or send -mail 
    # Functions to enable/disable relevant checkboxes depending on radiobutton choice
    def Uniopts():
        ''' Disable irrelevant checkboxes '''
        billentry.config(state=tk.DISABLED)
        unientry.config(state=tk.NORMAL)
        messfile.set('coach_email_outstanding_unis.txt')
        # clear current team selector... this autoloads oldteams
        for i, val in enumerate(teamdict):
            teamlist[i].set(0)
        emailtitle.set('Return of uniforms for your Cabrini team')
    def Contactopts():
        ''' Disable irrelevant checkboxes '''
        billentry.config(state=tk.DISABLED)
        unientry.config(state=tk.DISABLED)
        messfile.set('coach_email_contacts.txt')
        emailtitle.set('Contact list for your Cabrini team')    
    def Billopts():
        ''' Disable irrelevant checkboxes '''
        billentry.config(state=tk.NORMAL)
        unientry.config(state=tk.DISABLED)
        messfile.set('coach_email_outstanding_bills.txt')
        emailtitle.set('Fees still owed by your Cabrini team')
    def Otheropts():
        ''' Display relevant choices for other generic message to parents '''
        billentry.config(state=tk.DISABLED)
        unientry.config(state=tk.DISABLED)
        messfile.set('temp_message.txt')
        emailtitle.set('Message from Sponsors Club') 
    # e-mail title and message file name
    rownum=0
    tk.Label(root, text='Title for e-mail').grid(row=rownum, column=0)
    titleentry=tk.Entry(root, textvariable=emailtitle)
    titleentry.config(width=30)
    titleentry.grid(row=rownum, column=1)
    rownum+=1
    tk.Label(root, text='messagefile').grid(row=rownum, column=0)
    messentry=tk.Entry(root, textvariable=messfile)
    messentry.config(width=30)
    messentry.grid(row=rownum, column=1)
    rownum+=1
    # Choose counts, deriv, both or peaks plot (radio1)
    tk.Radiobutton(root, text='Missing uniforms', value='Unis', variable = mtype, command=Uniopts).grid(row=rownum, column=0)
    tk.Radiobutton(root, text='Send contact info', value='Contacts', variable = mtype, command=Contactopts).grid(row=rownum, column=1)
    tk.Radiobutton(root, text='Send bill info', value='Bills', variable = mtype, command=Billopts).grid(row=rownum, column=2)
    tk.Radiobutton(root, text='Other message', value='Other', variable = mtype, command=Otheropts).grid(row=rownum, column=3)
    rownum+=1
    asstcheck=tk.Checkbutton(root, variable=asstbool, text='Email asst coaches?')
    asstcheck.grid(row=rownum, column=0) # can't do immediate grid or nonetype is returned
    rownum+=1
    tk.Label(root, text='Bill_list file name').grid(row=rownum, column=0)
    billentry=tk.Entry(root, textvariable=billname)
    billentry.grid(row=rownum, column=1)
    rownum+=1
    tk.Label(root, text='Missing uni file name').grid(row=rownum, column=0)
    unientry=tk.Entry(root, textvariable=unifilename)
    unientry.grid(row=rownum, column=1)
    rownum+=1
    # insert team selector 
        # Specific team selector section using checkboxes
    teamdict=shortnamedict(teams)
    teamlist=[] # list of tk bools for each team
    # Make set of bool/int variables for each team
    for i, val in enumerate(teamdict):
        teamlist.append(tk.IntVar())
        if '#' not in val:
            teamlist[i].set(1) # Cabrini teams checked by default
        else:
            teamlist[i].set(0) # transfer team

    # make checkbuttons for each team
    for i, val in enumerate(teamdict):
        thisrow=i%5+1+rownum # three column setup
        thiscol=i//5
        thisname=teamdict.get(val,'')
        tk.Checkbutton(root, text=thisname, variable=teamlist[i]).grid(row=thisrow, column=thiscol)
    rownum+=math.ceil(len(teamlist)/5)+2

    # Decision buttons bottom row
    def chooseall(event):
        ''' Select all teams '''
        for i, val in enumerate(teamdict):
            teamlist[i].set(1)
    def clearall(event):
        ''' deselect all teams '''
        for i, val in enumerate(teamdict):
            teamlist[i].set(0)
    def abort(event):
        choice.set('abort')        
        root.destroy()
    def test(event):
        choice.set('test')        
        root.destroy()  
    def KCtest(event):
        choice.set('KCtest')
        root.destroy()
    def send(event):
        choice.set('send')        
        root.destroy()  
    rownum+=1
    d=tk.Button(root, text='All teams')
    d.bind('<Button-1>', chooseall)
    d.grid(row=rownum, column=0)
    
    d=tk.Button(root, text='Clear teams')
    d.bind('<Button-1>', clearall)
    d.grid(row=rownum, column=1)
    
    d=tk.Button(root, text='Abort')
    d.bind('<Button-1>', abort)
    d.grid(row=rownum, column=2)

    d=tk.Button(root, text='Test')
    d.bind('<Button-1>', test)
    d.grid(row=rownum, column=3)

    d=tk.Button(root, text='KCtest')
    d.bind('<Button-1>', KCtest)
    d.grid(row=rownum, column=4)
    
    d=tk.Button(root, text='Send')
    d.bind('<Button-1>', send)
    d.grid(row=rownum, column=5)

    root.mainloop()

    if choice.get()!='abort':
        kwargs={}
        if choice.get()=='KCtest':
            kwargs.update({'KCtest':True}) 
            kwargs.update({'choice':'send'}) 
        else:
            kwargs.update({'choice':choice.get()}) # send, KCtest (internal) or test (to log file) 
        if asstbool.get()==True:
            kwargs.update({'asst':True}) # Optional send to asst. coaches if set to True
        emailtitle=emailtitle.get()
        messagefile='messages\\'+messfile.get()
        # Handle selection of team subsets
        selteams=[]
        for i, val in enumerate(teamdict):
            if teamlist[i].get()==1:
                selteams.append(val)
        # Filter teams based on checkbox input
        teams=teams[teams['Team'].isin(selteams)]
        teams=teams.drop_duplicates(['Team','Sport'])
        if mtype.get()=='Contacts':
            mtype='contacts'
            try:
                Mastersignups = pd.read_csv('master_signups.csv', encoding='cp437')
                players= pd.read_csv('players.csv', encoding='cp437')
                famcontact= pd.read_csv('family_contact.csv', encoding='cp437')
                kwargs.update({'SUs':Mastersignups,'players':players,'famcontact':famcontact})
            except:
                print('Problem loading mastersignups, players, famcontact')
                return
        elif mtype.get()=='Bills':
            mtype='bills'
            try:
                Mastersignups = pd.read_csv('master_signups.csv', encoding='cp437')
                billlist=pd.read_csv(billfile.get(), encoding='cp437')
                kwargs.update({'bills':billlist, 'SUs':Mastersignups})                
                kwargs.update({'SUs':Mastersignups,'players':players,'famcontact':famcontact})
            except:
                print('Problem loading billlist, mastersignups')
                return
        elif mtype.get()=='Unis':
            mtype='unis'
            try:
                missing=pd.read_csv(unifilename.get(), encoding='cp437')
                oldteams=pd.read_excel('Teams_coaches.xlsx', sheetname='Oldteams') # loads all old teams in list
                kwargs.update({'oldteams':oldteams,'missing':missing})                
            except:
                print('Problem loading missingunis, oldteams')
                return
        elif mtype.get()=='Other':
            # nothing special to load?
            pass 
        emailcoaches(teams, coaches, mtype, emailtitle, messagefile, gdrivedict, **kwargs)
    return

def maketextsched(sched,teams, coaches, fields, messagefile, logfile, **kwargs):
    ''' Concise textable game schedule for cell only people from extracted Cabrini schedule'''
    # Convert dates/ times from timestamp to desired string formats for proper output
    if type(sched.iloc[0]['Time'])==datetime.time:
        sched.Time=sched.Time.apply(lambda x:datetime.time.strftime(x, format='%I:%M %p'))
    else:
        print('Time format is', type(sched.iloc[0]['Time']))
    if type(sched.iloc[0]['Date'])==datetime.time:
        sched['Date']=sched['Date'].dt.strftime(date_format='%d-%b-%y')
    else:
        print('Date format is', type(sched.iloc[0]['Date']))
    if 'div' in kwargs:
        div=kwargs.get('div','')
        grade=int(div[0])
        if div[1].upper()=='G':
            gender='f'
        elif div[1].upper()=='B':
            gender='m'
        teams=teams[(teams['Grade']==grade) & (teams['Gender']==gender)]
    log=open(logfile,'w', encoding='utf-8')
    myteams=pd.merge(teams, coaches, on='Coach ID', how='left', suffixes=('','_2'))
    # Make list of sport/team/school/graderange/coachinfo/playerlist
    teamlist=[]
    # Open generic message header 
    with open('messages\\'+messagefile, 'r') as file:
        blankmess=file.read()
    for index, row in myteams.iterrows():
        # get school 
        if '#' not in myteams.loc[index]['Team']:
            school='Cabrini'
            try:
                coachinfo=myteams.loc[index]['Fname']+' '+ myteams.loc[index]['Lname']+' ('+myteams.loc[index]['Email']+')'
            except:
                coachinfo=''
        else:
            school=myteams.loc[index]['Team'].split('#')[0]
            coachinfo=''
        # Get gender
        if row.Gender.lower()=='f':
            gender='girls'
        elif row.Gender.lower()=='m':
            gender='boys'
        else:
            print('Problem finding team gender')
        grrang=str(myteams.loc[index]['Graderange'])
        if len(grrang)==2:
            grrang=grrang[0]+'-'+grrang[1]
        if grrang.endswith('2'):
            grrang+='nd'
        elif grrang.endswith('3'):
            grrang+='rd'
        else: 
            grrang+='th'
        grrang=grrang.replace('0','K')
        # Get sport, team, graderange, coach info (first/last/e-mail), playerlist
        teamlist.append([myteams.loc[index]['Sport'], myteams.loc[index]['Team'], school,
            grrang, gender, coachinfo, myteams.loc[index]['Playerlist']])
        # get dictionary of teams found/matched in CYC schedule
    for i, [sport, team, school, graderange, gender, coachinfo, playerlist] in enumerate(teamlist):
        # Either have cabrini only schedule or full CYC schedule
        if 'Team' in sched:
            thissched=sched[sched['Team']==team].copy()
            thissched=thissched[['Date','Time','Day', 'Location']]
        else:
            print("Couldn't find schedule for", school, str(graderange), sport, team)
            continue
        if len(thissched)==0:
            print('Games not found for ', team)
            continue
        # TODO construct textable message in log 
        games=''
        for index, row in thissched.iterrows():
            # output date, day, time, location 
            games+=row.Date+'   '+row.Day+'  '+row.Time+'   '+row.Location+'\n'
        thismess=blankmess.replace('$SCHEDULE', games)
        thismess=thismess.replace('$GRADERANGE', graderange)
        thismess=thismess.replace('$GENDER', gender)
        thismess=thismess.replace('$SPORT', sport)
        # now create/ insert location and address table
        thisft=makefieldtable(thissched, fields)
        myfields=''
        for index, row in thisft.iterrows():
            # output date, day, time, location 
            myfields+=row.Location+'   '+row.Address+'\n'
        thismess=thismess.replace('$FIELDTABLE', myfields)
        log.write(thismess+'\n')
    log.close()
    return

''' TESTING
teamnamedict=findschteams(sched, teams, coaches)
'''

''' TESTING
sched=pd.read_csv('Cabrini_Bball2018_schedule.csv')
sport, team, school, graderange, gender, coachinfo, playerlist=teamlist[0]    i=0
recipients=['tkc@wustl.edu','kcroat@gmail.com']
'''

def sendschedule(teams, sched, fields, Mastersignups, coaches, year, famcontact, emailtitle, blankmess, **kwargs):
    ''' Top level messaging function for notifying families of team assignment/ CYC card
    + optional short-team-player-recruiting
    via custom e-mail; one per player
    currently not including SMS 
    kwargs:
        choice - 'send' or 'test'
        recruit - T or F -- add recruiting statement for short teams
        mformat -  not really yet used ... just sending as text not html 
    '''
    # convert date- time from extracted schedule to desired str format
    # type will generally be string (if reloaded) or timestamp (if direct from prior script)
    ''' if already string just keep format the same, if timestamp or datetime then convert below
    if type(sched.iloc[0]['Time'])==str:
        sched.Time=pd.to_datetime(sched.Time, format='%H:%M:%S') # convert string to timestamp
        '''
    if type(sched.iloc[0]['Time'])!=str:
        # Then convert timestamp to datetime to desired string format
        sched.Time=sched.Time.apply(lambda x:pd.to_datetime(x).strftime(format='%I:%M %p'))
        if type(sched.iloc[0]['Date'])==str:
            try:
                sched.Date=pd.to_datetime(sched.Date, format='%m/%d/%Y')
            except:
                try:
                    sched.Date=pd.to_datetime(sched.Date, format='%Y-%m-%d')
                except:
                    print('Difficulty converting date with format', type(sched.iloc[0]['Date']))
        # convert to desired date string format
        sched['Date']=sched['Date'].dt.strftime(date_format='%d-%b-%y')
    choice=kwargs.get('choice','test')
    if choice=='send' or choice=='KCtest':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('parent_email_log.txt','w', encoding='utf-8')
    
    # this years signups only (later match for sport and team)
    Mastersignups=Mastersignups[Mastersignups['Year']==year]
    # Should be only one entry per coach
    myteams=pd.merge(teams, coaches, on='Coach ID', how='left', suffixes=('','_2'))
    # Make list of sport/team/school/graderange/coachinfo/playerlist
    teamlist=[]
    for index, row in myteams.iterrows():
        # get school 
        if '#' not in myteams.loc[index]['Team']:
            school='Cabrini'
            try:
                coachinfo=row.Fname+' '+ row.Lname+' ('+row.Email+')'
            except:
                coachinfo=''
        else:
            school=row.Team.split('#')[0]
            coachinfo=''
        # Get gender
        if row.Gender.lower()=='f':
            gender='girl'
        elif row.Gender.lower()=='m':
            gender='boys'
        else:
            print('Problem finding team gender')
        # Get sport, team, graderange, coach info (first/last/e-mail), playerlist
        teamlist.append([row.Sport, row.Team, school, gradetostring(row.Graderange),
            gender, coachinfo, row.Playerlist])
    # get dictionary of teams found/matched in CYC schedule
    teamnamedict=findschteams(sched, teams, coaches)
    # TESTING  sport, team, school, graderange, gender, coachinfo, playerlist=teamlist[i]    i=2
    for i, [sport, team, school, graderange, gender, coachinfo, playerlist] in enumerate(teamlist):
        # Either have cabrini only schedule or full CYC schedule
        if 'Team' in sched:
            thissched=sched[sched['Team']==team].copy()
            # shorten team name
            thissched['Home']=thissched['Home'].str.split('/').str[0]
            thissched['Away']=thissched['Away'].str.split('/').str[0]
            thissched['Home']=thissched['Home'].str.strip()
            thissched['Away']=thissched['Away'].str.strip()
            # Times/dates already reformatted
            thissched=thissched[['Date','Time','Day','Home','Away','Location']]
        else: # handle if an unsorted CYC schedule (not Cab only)
            if team in teamnamedict:
                [div,schname]=teamnamedict.get(team,'')
                thissched=getgameschedule(div,schname, sched)
                thissched=thissched[['Date','Time','Day','Division','Home','Away','Location']]
            else:
                print("Couldn't find schedule for", school, str(graderange), sport, team)
                continue
        if len(thissched)==0:
            print('Games not found for ', team)
            continue
        thisteam=Mastersignups[(Mastersignups['Sport']==sport) & (Mastersignups['Team']==team)]
        thisteam=pd.merge(thisteam, famcontact, on='Famkey' , how='left', suffixes =('','_2'))
        # Make all team-specific replacements in message body and email title
        thisteammess=blankmess
        thistitle=emailtitle
        # have to use caution due to $TEAMTABLE (common) and $TEAMNAME (rarely used)
        for j, col in enumerate(['$SPORT', '$TEAMNAME', '$SCHOOL', '$GRADERANGE', '$GENDER', '$COACH', '$PLAYERLIST']):
            if j!='$SPORT':
                val=teamlist[i][j]
            else: # lower-case sport name for replace
                val=teamlist[i][j].lower()
            try:
                thisteammess=thisteammess.replace(col, textwrap.fill(val, width=100))
                thistitle=thistitle.replace(col, val)
            except:
                print("Problem with teamname", val)
                continue
        # Convert thissched to string table and insert into message
        thisteammess=thisteammess.replace('$SCHEDULE', thissched.to_string(index=False, justify='left'))
        #Make and insert field table 
        thisft=makefieldtable(thissched, fields)
        thisteammess=thisteammess.replace('$FIELDTABLE', thisft.to_string(index=False, justify='left'))
        # Get coach emails
        recipients=getcoachemails(team, teams, coaches, **{'asst':True})
        # Now get all unique team email addresses (single message to coach and team)...drops nan
        recipients=getallteamemails(thisteam, recipients)
        if choice=='KCtest': # internal send test
            recipients=['tkc@wustl.edu','kcroat@gmail.com']
            choice='send'
        # Create custom email message (can have multiple sports in df)
        if choice=='send':
            try: # single simultaneous e-mail to all recipients
                msg=MIMEText(thisteammess,'plain')
                msg['Subject'] = thistitle
                msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'  
                msg['To']=','.join(recipients)
                smtpObj.sendmail('sfcasponsorsclub@gmail.com', recipients, msg.as_string())
                print ('Message sent to ', ','.join(recipients))
            except:
                print('Message to ', team, 'failed.')
            if not recipients:
                print('No email addresses for team ', team)
        else: # Testing mode ... just write to log w/o e-mail header and such
            logfile.write(thistitle+'\n')
            logfile.write(thisteammess+'\n')
    # close log file (if testing mode)
    if choice!='send':
        logfile.close()
    else:
        pass
        # TODO fix this attempted close
        # smtpObj.quit() # close SMTP connection
    return

# TESTING

def makegcals(sched, teams, coaches, fields, season, year, duration=1, **kwargs):
    ''' Turn standard CYC calendar into google calendar 
    description: 1-2 girls soccer vs opponent   
    kwargs: 
        div - get only calendar for given division
        school - Cabrini ... drop transfer teams w/ #
        splitcals - separate calendar for each team (default True), 
        '''
    #TODO ... Test after alteration of address field
    if 'school' in kwargs:
        if kwargs.get('school','')=='Cabrini':
            # drop transfer teams w/ #
            teams=teams[~teams['Team'].str.contains('#')]
    if 'div' in kwargs:
        div=kwargs.get('div','')
        grade=int(div[0])
        if div[1].upper()=='G':
            gender='f'
        elif div[1].upper()=='B':
            gender='m'
        teams=teams[(teams['Grade']==grade) & (teams['Gender']==gender)]
    # ensure correct formats for separate date and time columns
    if type(sched.iloc[0]['Date'])==str:
        try: # format could be 10/18/2018 0:00
            sched.Date=sched.Date.str.split(' ').str[0]
            sched.Date=pd.to_datetime(sched.Date, format='%m/%d/%Y') 
        except:
            pass
        try:
            sched.Date=pd.to_datetime(sched.Date, format='%m/%d/%Y') 
        except:
            pass
        try:
            sched.Date=pd.to_datetime(sched.Date, format='%Y-%m-%d')
        except:
            print('Problem converting date format of ', sched.iloc[0]['Date'])
    # gcal needs %m/%d/%y (not zero padded)
    sched['Date']=sched['Date'].dt.strftime(date_format='%m/%d/%Y')
    if type(sched.iloc[0]['Time'])==str:
        try:
            sched.Time=pd.to_datetime(sched.Time, format='%H:%M %p')
        except:
            try:            
                sched.Time=pd.to_datetime(sched.Time, format='%H:%M:%S') # convert string to timestamp
            except:
                print('Failed conversion of time column... check format')
    # Common reformatting of all gcals
    sched=sched.rename(columns={'Date':'Start Date','Time':'Start Time'})

    # Calculate end time while still a timestamp
    sched['End Time']=pd.to_datetime(sched['Start Time']) + datetime.timedelta(hours=1)
    sched['End Time']=pd.to_datetime(sched['End Time'])
    sched['End Time']=sched['End Time'].apply(lambda x:pd.to_datetime(x).strftime('%I:%M %p'))

    # Then convert timestamp to datetime to desired string format
    sched['Start Time']=sched['Start Time'].apply(lambda x:pd.to_datetime(x).strftime(format='%I:%M %p'))
    # Standard google calendar column names
    gcalcols=['Subject','Start Date', 'Start Time', 'End Date','End Time', 'All Day Event', 'Description', 'Location','Private']
    sched['All Day Event']='FALSE'
    sched['Private']='FALSE'
    sched['End Date']=sched['Start Date']
    # append short address to location field
    sched=pd.merge(sched, fields, on='Location', how='left', suffixes=('','_2'))
    # replace blank address (in case not found but shouldn't happen)
    sched['Address']=sched['Address'].replace(np.nan,'')
    sched['Location']=sched['Location']+' '+sched['Address']
    # Cabrini extracted schedule has team name column
    teamlist=np.ndarray.tolist(sched.Team.unique())
    shortnames=shortnamedict2(teams)
    # Get head coach email for team from coaches list
    teams=pd.merge(teams, coaches, how='left', on=['Coach ID'], suffixes=('','_2'))
    # Optional single calendar format
    if not kwargs.get('splitcal', True):
        combocal=pd.DataFrame(columns=gcalcols)
    for i, team in enumerate(teamlist):
        thissch=sched[sched['Team']==team]
        # Need to get associated sport from teams
        match=teams[teams['Team']==team]
        if len(match)==1:
            sport=match.iloc[0]['Sport']
            email=match.iloc[0]['Email']
        else:
            sport=''
            email=''
            print('Sport not found for team', team)
            # skip these teams (usually non-Cabrini team w/ Cab players)
            continue
        # Make unique description column
        descr=shortnames.get(team,'')+' '+ sport.lower()
        # Use 1-2nd girl soccer as calendar event title/subject
        thissch['Subject']=descr
        # Prepend grade/gender sport string to team opponents
        
        thissch['Description']=thissch['Home'].str.split('/').str[0] +' vs '+thissch['Away'].str.split('/').str[0]
        # prepend string 1-2 girls soccer to each event
        thissch['Description']=thissch['Description'].apply(lambda x:descr+': '+x)
        cancel='Contact '+str(email)+' for cancellation/reschedule info'
        # Add line w/ coach email for cancellation
        thissch['Description']=thissch['Description'].apply(lambda x:x+'\r\n'
               + cancel)
        thissch=thissch[gcalcols]
        if kwargs.get('splitcal', True): # separate save of gcal for each team 
            fname='gcal_'+descr+'.csv'
            thissch.to_csv(fname, index=False)
        else: # add to single jumbo cal
            combocal=pd.concat([combocal, thissch], ignore_index=True)
    if not kwargs.get('splitcal', True):
        fname='Cabrini_gcal_'+season.lower()+str(year)+'.csv'
        combocal.to_csv(fname, index=False)
    return 

def getgameschedule(div, schname, sched):
    ''' Find and extract game schedule for team with matching name '''
    sched=sched.rename(columns={'Game Time':'Time','Division Name':'Division', 'Field Name':'Location','Visitor':'Away','AwayTeam':'Away','Home Team':'Home'})
    thissched=sched[sched['Division'].str.startswith(div)]
    thissched=thissched[(thissched['Home'].str.contains(schname)) | (thissched['Away'].str.contains(schname))]
    # already sorted and date-time strings already converted to preferred format in getcabsch 
    return thissched

def findschteams(sched, teams, coaches):
    ''' Find team names as listed in schedule or elsewhere based on division (e.g. 1B) 
    plus coach and/or school 
    return dictionary with all internal names and associated CYC schedule div & name '''
    sched=sched.rename(columns={'Game Time':'Time','Field Name':'Location','AwayTeam':'Away','Home Team':'Home'})
    sched=sched[pd.notnull(sched['Home'])] # drop unscheduled games

    # Get identifying info for all teams
    myteams=pd.merge(teams, coaches, on='Coach ID', how='left', suffixes=('','_2'))
    # Make list of sport/team/school/graderange/coachinfo/playerlist
    # Need double entry for double-rostered teams
    double=myteams.copy()
    double=double[double['Team'].str.contains('!')]
    for index, row in double.iterrows():
        doublename=row.Team
        tok=doublename.split('-')
        name1=tok[0]+'-'+tok[1]+'-'+tok[2].split('!')[0]+'-'+tok[3]
        name2=tok[0]+'-'+tok[1]+'-'+tok[2].split('!')[1]+'-'+tok[3]
        double=double.set_value(index, 'Team', name2)
        myteams['Team']=myteams['Team'].str.replace(doublename, name1)
        myteams=myteams.append(double)

    # First find all Cabrini teams 
    teamnamedict={}
    for index, row in myteams.iterrows():
        # get division
        if '-' in row.Team:
            school='Cabrini'
            coach=str(row.Lname)
            try:
                tok=row.Team.split('-')
                div=tok[2]
            except:
                print("Couldn't find division for", row.Team)
                continue
        # non-cabrini team w/ transfers
        elif '#' in row.Team:
            school=row.Team.split('#')[0]
            coach=str(row.Coach)
            if '??' in coach:
                coach='nan'
            if row.Gender=='m':
                div=str(row.Grade)+'B'
            else:
                div=str(row.Grade)+'G'
        else: # typically junior teams
            print("no schedule for ", row.Team)
            continue
        # Get sport, school, division, coach last nameteam, graderange, coach info (first/last/e-mail), playerlist
        thisdiv=sched[sched['Division'].str.startswith(div)]
        # On rare occasions teams can only have away games
        divteams=np.ndarray.tolist(thisdiv['Home'].unique())
        divteams.extend(np.ndarray.tolist(thisdiv['Away'].unique()))
        divteams=set(divteams)
        divteams=list(divteams)
        # find this schools teams
        thisteam=[team for team in divteams if school.lower() in team.lower()]
        # handle multiple teams per grade
        if len(thisteam)>1:
            thisteam=[team for team in thisteam if coach.lower() in team.lower()]
            if len(thisteam)>1: # Same last name?  use exact coach match
                coach=str(myteams.loc[index]['Coach'])
                thisteam=[team for team in thisteam if coach in team.lower()]
        if len(thisteam)==1: # found unique name match
            # Need division and name due to duplicates problem
            try:
                teamnamedict.update({row.Team: [div, thisteam[0].strip()]})
            except:
                print("Couldn't hash", row.Team)
        else:
            print("Couldn't find unique schedule team name for", row.Team, div)
    return teamnamedict

def shortnamedict(teams):
    ''' From teams list, make shortened name dictionary for tk display (i.e. 1G-Croat or 3G-Ambrose)'''
    teamdict={}
    for index, row in teams.iterrows():
        # Get coach name or school
        if '#' in teams.loc[index]['Team']:
            name=teams.loc[index]['Team'].split('#')[0]
        else:
            name=str(teams.loc[index]['Coach'])
        if teams.loc[index]['Gender']=='m':
            gend='B'
        else:
            gend='G'
        grrange=str(teams.loc[index]['Graderange'])
        grrange=grrange.replace('0','K')
        thisname=grrange+gend+'-'+name
        teamdict.update({teams.loc[index]['Team']:thisname})
    return teamdict

def shortnamedict2(teams):
    ''' From teams list, make shortened name dictionary for gcal  (i.e. 1-2 girls)'''
    teamdict={}
    for index, row in teams.iterrows():
        if teams.loc[index]['Gender']=='m':
            gend=' boys'
        else:
            gend=' girls'
        grrange=str(teams.loc[index]['Graderange'])
        grrange=grrange.replace('0','K')
        if len(grrange)>1:
            grrange=grrange[0]+'-'+grrange[1]
        if grrange.endswith('2'):
            grrange+='nd'
        elif grrange.endswith('3'):
            grrange+='rd'
        try:
            if int(grrange[-1]) in range(4,9):
                grrange+='th'
        except:
            pass # p
        thisname=grrange+gend
        teamdict.update({teams.loc[index]['Team']:thisname})
    return teamdict

def findrecentfile(filelist):
    ''' Return most recently dated file from list of autonamed files .. date format is always 27Jan17 '''
    dates=[s.split('_')[1].split('.')[0] for s in filelist]
    try:
        dates=[datetime.datetime.strptime(val, "%d%b%y") for val in dates]
        datepos=dates.index(max(dates)) # position of newest date (using max)
        newfile=filelist[datepos]
    except:
        print('File date comparison failed... using first one')
        newfile=filelist[0]
    return newfile
'''TESTING
missing=pd.read_csv('missingunilist_29Dec17.csv')
'''
def emailcoaches(teams, coaches, mtype, emailtitle, messagefile, gdrivedict, **kwargs):
    ''' Send e-mails to all coach: types are contacts, bills, unis (missing uniforms info)
    various dfs are passed via kwargs when necessary 
    4/1/17 works for contacts and unpaid bill summary
    HTML message w/ plain text alternate
    kwargs:  choice -- send, test or KCtest (real internal send )
    '''
    choice=kwargs.get('choice','test') # send or test (KCtest kwarg true set separately)
    print(choice)
    if choice=='send': # true send or internal live send to tkc@wustl
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('coach_email_log.txt','w', encoding='utf-8')
    # Iterate over teams 
    teaminfo=[]  # list w/ most team info
    if mtype=='unis' and 'missing' in kwargs: 
        # Iterate only over old teams with missing uniforms
        missing=kwargs.get('missing',pd.DataFrame())
        for index, row in missing.iterrows():
            if row.Gender.lower()=='f':
                gend='girls'
            else:
                gend='boys'
            # coach and graderange are nan... not needed for missing unis
            teaminfo.append([row.Year, row.Sport, gradetostring(row.Grade), gend, row.Team, 'coach', 
                             'graderange', row.Number])
        # Replace teams with oldteams
        teams=kwargs.get('oldteams',pd.DataFrame())
    else: # iterate through current teams for contacts, bills, cards
        for index, row in teams.iterrows():
            if row.Gender.lower()=='f':
                gend='girls'
            else:
                gend='boys'
            
            teaminfo.append([row.Year, row.Sport, gradetostring(row.Grade), gend, row.Team, row.Coach, 
                             gradetostring(row.Graderange), row.Number])
    with open(messagefile,'r') as mess:
        message=mess.read() # generic e-mail message body with limited find/replace
    # insert up to date google drive link for this season (various options)
    if 'GDRIVE' in message: # global replacements (for all teams)
        for key, val in gdrivedict.items():
            message=message.replace(key, val)
    for i, [year, sport, grade, gender, team, coach, graderange, numplayers] in enumerate(teaminfo):
        # Make all team-specific replacements in message body and email title
        thisteammess=message
        thistitle=emailtitle
        for j, col in enumerate(['$YEAR', '$SPORT', '$GRADE', '$GENDER', '$TEAMNAME', '$COACH', '$GRADERANGE', '$NUMBER']):
            thisteammess=thisteammess.replace(col, str(teaminfo[i][j]))
            thistitle=thistitle.replace(col, str(teaminfo[i][j]))
        # Replace missing CYC cards list if requested by message
        if '$MISSINGCARDS' in message:
            if 'SUs' not in kwargs:
                print('Signups needed to find missing cards')
                return
            carddict=findcards()
            SUs=kwargs.get('SUs','')
            missingstr=findmissingcards(team, SUs, carddict)
            thisteammess=thisteammess.replace('$MISSINGCARDS', missingstr)
        # Get head coach e-mail address (list optionally incl. assistants in kwargs)
        if 'KCtest' not in kwargs:
            coachemail=getcoachemails(team, teams, coaches, **kwargs)
        else:
            if i==0: # send first message only as live test
                coachemail=['tkc@wustl.edu','kcroat@gmail.com']
            else:
                coachemail=[]
        if coachemail==[]: # list of head and asst coaches
            print('No valid coach e-mail for '+ team +'\n')
            continue
        # handle the special message cases 
        if mtype=='bills': # replacement of teamtable for bills
            if 'SUs' and 'bills' not in kwargs:
                print('Signups and billing list needed for e-mail send bill to coaches option.')
                return
            SUs=kwargs.get('SUs','')
            bills=kwargs.get('bills','')
            teambilltable=makebilltable(team, bills, SUs)
            if teambilltable=='': # team's all paid up, skip e-mail send
                print('All players paid for team'+team+'\n')
                continue # no e-mail message sent
            thismess_html=thisteammess.replace('$TEAMTABLE', teambilltable.to_html(index=False))
            thismess_plain=thisteammess.replace('$TEAMTABLE', teambilltable.to_string(index=False))
        elif mtype=='contacts': # replacement of teamtable for contact
            if 'SUs' and 'players' and 'famcontact' not in kwargs:
                print('Signups, player and family contact info needed for contact lists to coaches option.')
                return
            SUs=kwargs.get('SUs','')
            # Collapse track sub-teams to single track team
            SUs['Team']=SUs['Team'].str.replace(r'track\d+', 'Track', case=False)
            SUs=SUs[SUs['Year']==year] # in case of duplicated team name, filter by year
            players=kwargs.get('players','')
            famcontact=kwargs.get('famcontact','')
            contacttable=makecontacttable(team, SUs, players, famcontact)
            # Find/replace e-mail addresses
            # Convert df to text
            thismess_html=thisteammess.replace('$TEAMTABLE', contacttable.to_html(index=False))
            thismess_plain=thisteammess.replace('$TEAMTABLE', contacttable.to_string(index=False))
        elif mtype=='unis': # replacement of teamtable for uniform returns
            # Probably need current and former teams
            if 'SUs' and 'oldteams' and 'missing' not in kwargs:
                print('Signups, old teams and missing uniform list needed for e-mail unis to coaches option.')
                return
            # in this case teams iterator with have old not current teams
            unitable=makeunitable(team, missing)
            thismess_html=thisteammess.replace('$TEAMTABLE', unitable.to_html(index=False))
            thismess_plain=thisteammess.replace('$TEAMTABLE', unitable.to_string(index=False))
        else: # generic message w/o $TEAMTABLE
            thismess_html=thisteammess
            thismess_plain=thisteammess
        if choice=='send':
            try:
                # Create message container - the correct MIME type is multipart/alternative.
                msg = MIMEMultipart('alternative') # message container
                msg['Subject'] = emailtitle
                msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'                
                part1=MIMEText(thismess_plain,'plain')
                part2=MIMEText(thismess_html,'alternate')                
                msg['To']=','.join(coachemail) # single e-mail or list
                msg.attach(part1) # plain text
                msg.attach(part2) # html (last part is preferred)
                # Simultaneous send to all in recipient list
                smtpObj.sendmail('sfcasponsorsclub@gmail.com', coachemail, msg.as_string())
                print ('Message sent to ', ','.join(coachemail))
            except:
                print('Message to ', ','.join(coachemail), ' failed.')
        else: # testing only... open log file
            logfile.write(emailtitle+'\n')
            logfile.write(thismess_plain+'\n')
    return

def gradetostring(val):
    ''' Turns grade or grade ranges into strings with appropriate ending 23 becomes '2-3rd' '''
    if len(str(val))==2:
        val=str(val)[0]+'-'+str(val)[1]
    else:
        val=str(val)
    if val.endswith('1'):
        val+='st'
    elif val.endswith('2'):
        val+='nd'
    elif val.endswith('3'):
        val+='rd'
    else: 
        val+='th'
    return val 

def getallteamemails(df, emails):
    ''' Get all unique e-mails associated with team 
    passed emails contains coach emails already extracted
    email1, email2, and email3 columns all present in family contacts'''
    
    emails=np.ndarray.tolist(df.Email1.unique())
    emails.extend(np.ndarray.tolist(df.Email2.unique()))
    emails.extend(np.ndarray.tolist(df.Email3.unique()))
    emails=set(emails)
    emails=list(emails)
    emails=[i for i in emails if str(i)!='nan']
    return emails

def getcoachemails(team, teams, coaches, **kwargs):
    ''' Returns head coach e-mail for given team and optionally asst coaches '''
    teams=teams.drop_duplicates('Team') # drop coed team duplicate
    thisteam=teams[teams['Team']==team]
    emails=[]
    IDs=[]
    if len(thisteam)!=1:
        print(team, 'not found in current teams list')
        return emails # blank list
    thisteam=thisteam.dropna(subset=['Coach ID'])
    if len(thisteam)!=1:
        print('Coach ID not found for', team)
        return emails # blank list
    if thisteam.iloc[0]['Coach ID']!='': # possibly blank
        IDs.append(thisteam.iloc[0]['Coach ID'])
    thisteam=thisteam.dropna(subset=['AssistantIDs'])
    if kwargs.get('asst', False): # optional send to asst coaches
        if len(thisteam)==1: # grab asst IDs if they exist
            asstIDs=thisteam.iloc[0]['AssistantIDs'] 
            asstIDs=[str(s).strip() for s in asstIDs.split(",")]
            IDs.extend(asstIDs)
    # now find email addresses for this set of CYC IDs
    thesecoaches=coaches[coaches['Coach ID'].isin(IDs)]
    thesecoaches=thesecoaches.dropna(subset=['Email'])
    emails=np.ndarray.tolist(thesecoaches.Email.unique()) # surely won't have blank string
    return emails    
    
def makeunitable(team, missing):
    ''' Make missing uniform table for auto-emailing to head coach; looping over old teams and
    unis identified as not yet returned from prior seasons
    '''
    # Could be problem with non-unique team 
    thisteam=missing[missing['Team']==team]
    mycols=['First', 'Last', 'Issue date', 'Sport', 'Year', 'Uniform#', 'Team']
    thisteam=thisteam[mycols]
    thisteam=thisteam.replace(np.nan,'')
    return thisteam
    
def makecontacttable(team, SUs, players, famcontacts):
    ''' Make team contacts list for auto-emailing to head coach 
    first, last, grade, school, phone/text/email 1&2 '''
    # Find subset of all signups from this team
    thisteam=SUs[SUs['Team']==team]
    # Get school from players
    thisteam=pd.merge(thisteam, players, on='Plakey', how='left', suffixes=('','_r'))
    # Get other contact info from family contacts
    thisteam=pd.merge(thisteam, famcontacts, on='Famkey', how='left', suffixes=('','_r'))
    mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone1', 'Text1','Email1', 'Phone2', 'Text2', 'Email2']
    thisteam=thisteam[mycols]
    thisteam=thisteam.replace(np.nan,'')
    return thisteam

def makebilltable(team, billlist, Mastersignups):
    ''' Create billing message summary tabl for individual team, return as message string, called by 
    e-mail loop or e-mail log test '''
    thisteam=billlist[billlist['Teams'].str.contains(team)]
    if len(thisteam)==0:
        return '' # pass blank string and test for it (skip e-mail )
    # Construct comment with "also owes for Sibling basketball"
    thisteam['Comments']=''
    for index, row in thisteam.iterrows():
        SUs=thisteam.loc[index]['SUkeys'] # this family's current season signups
        if ',' in SUs: # Need to match the sibling (not on this team)
            SUkeys=[int(i) for i in SUs.split(',')]
            theseSUs=Mastersignups[Mastersignups['SUkey'].isin(SUkeys)]
            otherSUs=theseSUs[theseSUs['Team']!=team] # other family signups... summarize in comments
            otherSUs=otherSUs.sort_values(['Sport'], ascending=True)
            tempstr='also owes for: '
            for ind, ro in otherSUs.iterrows():
                tempstr+=otherSUs.loc[ind]['First']+' '
                tempstr+=otherSUs.loc[ind]['Sport'].lower()
            thisteam=thisteam.set_value(index,'Comments',tempstr)    
    mycols=['Family','Players','Charges','CurrPayments','Balance','Email1','Phone1','Comments']
    thisteam['Balance']*=-1
    thisteam=thisteam[mycols] # summary for insertion into e-mail to team's coaches
    teamsum=thisteam.to_html() # convert to html table for insertion
    return teamsum # table with this team's currently outstanding bills

def emailrecruits(Recruits, famcontact, emailtitle, messagefile, **kwargs):
    ''' Top level messaging function for recruiting players via custom e-mail; one for each player (not by family)
    currently not including SMS '''
    choice=kwargs.get('choice','test') # send or test
    if choice=='send':
        smtpObj = smtplib.SMTP('smtp.gmail.com', 587) # port 587
        smtpObj.ehlo() # say hello
        smtpObj.starttls() # enable encryption for send 
        print('Enter password for sponsors club gmail ')
        passwd=input()
        smtpObj.login('sfcasponsorsclub@gmail.com', passwd)
    else: # testing only... open log file
        logfile=open('player_recruit_email_log.txt','w', encoding='utf-8')
    # Get address, full e-mail/phone list via family merge
    Recruits=pd.merge(Recruits, famcontact, how='left', on='Famkey', suffixes=('','_2'))  
    Recruits=Recruits[pd.notnull(Recruits['Email1'])]
    Recs=Recruits.groupby(['First','Last'])
    for [first, last], row in Recs:
        recipients=getemailadds(row.iloc[0]) # list of recipients
        # customized email title w/ first name commonly used
        thistitle=emailtitle.replace('$FIRST', first)
        # create custom email message (can have multiple sports in df)
        try:
            thismess=makerecinfomessage(row, messagefile)
        except:
            print("Rec message error for", first, last)
        if choice=='send':
            msg=MIMEText(thismess,'plain')
            msg['Subject'] = thistitle
            msg['From'] = 'Cabrini Sponsors Club <sfcasponsorsclub@gmail.com>'                
            # part2=MIMEText(thismess_html,'alternate')                
            msg['To']=','.join(recipients)  # single e-mail or list
            # Add From/To/Subject to actual e-mail
            try:
                smtpObj.sendmail('sfcasponsorsclub@gmail.com', recipients, msg.as_string())
                print ('Message sent to ', ','.join(recipients))
            except:
                print('Message to ', ','.join(recipients), ' failed.')
            if not recipients:
                print('No email address for ', first, last)
        else: # testing mode
            logfile.write(thistitle+'\n')
            logfile.write(thismess+'\n')
    if choice!='send':
        logfile.close()
    else: # TODO need to close connection?
        pass
    return

def makerecinfomessage(row, messagefile):
    ''' Make e-mail message for family from billrow (passed as Series)
    pass family's row from bill, signup details, payment logbook, and list of email recipients 
    row is groupby df (not Series)
    '''
    #TODO 8/27 this will need some testing after redesign
    sports=np.ndarray.tolist(row.Sport.unique())
    sports=[s.lower() if s!='VB' else 'volleyball' for s in sports]
    phonelist=[]
    for i, col in enumerate(['Phone1','Phone2']):
        if str(row.iloc[0][col])!='nan':
            phonelist.append(row.iloc[0][col])
    emaillist=[]
    for i, col in enumerate(['Email1','Email2']):
        if str(row.iloc[0][col])!='nan':
            emaillist.append(row.iloc[0][col])
    # load default message and then make custom substitutions depending on contents of each row in 
    with open(messagefile,'r') as file:
        # email title, recipients, 
        message=file.read()
        message=message.replace('$SPORTS', ' and '.join(sports))
        message=message.replace('$FIRST', row.iloc[0]['First'])
        message=message.replace('$LAST', row.iloc[0]['Last'])
        # contact info is tagged in message so recruited parent can check it
        message=message.replace('$PHONELIST', ', '.join(phonelist))
        message=message.replace('$EMAILLIST', ', '.join(emaillist))
        try:
            message=message.replace('$ADDRESS', row.iloc[0]['Address']+', '+str(row.iloc[0]['Zip']) )
        except:
            print('No address for',row.iloc[0]['First'],row.iloc[0]['Last'])
            message=message.replace('$ADDRESS','') 
        return message

def getemailadds(thisrow):
    '''Find email address(es) from series (single family row from bill or other df
    and return as list ''' 
    email1=str(thisrow.Email1)
    email2=str(thisrow.Email2)
    recipients=[]
    if '@' in email1:
        recipients.append(email1)
    if '@' in email2:
        recipients.append(email2)
    return recipients

''' TESTING
i=0   team=teamlist[i]
'''
def gamecardmaker(teams, coaches, Mastersignups, sched, pastelist, gctemplate):
    ''' Somewhat generic insertion into existing excel file template
    open Excel_python_insert_template'''

    teamlist=np.ndarray.tolist(sched.Team.unique())
    # only need teams in schedule list (Cabrini CYC only)
    teams=teams[teams['Team'].isin(teamlist)]
    # get player lists (last, first), player number list, and coach list (last, first)
    teaminfo=[]
    myplayers=pd.merge(teams, Mastersignups, how='inner', on=['Team','Sport','Year'], suffixes=('','_2'))
    removelist=[]
    for i, team in enumerate(teamlist):
        thisteam=myplayers[myplayers['Team']==team]
        thisteam=thisteam.sort_values(['Last'])
        plalist=[]
        planumlist=[]
        for index, row in thisteam.iterrows():
            plalist.append(row.Last+', '+row.First)
            planumlist.append(row['Uniform#'])
    # now need to get coach(es) last, first
        headID=[]
        match=teams[teams['Team']==team]
        match=match.drop_duplicates(['Team']) # handle co-ed duplicates
        match=match[pd.notnull(match['Coach ID'])]
        if len(match)==1:
            headID.append(match.iloc[0]['Coach ID'])
        else:
            print('No coach found for team', team)
            removelist.append(team)
            # Remove from above teamlist
            continue
        match=match[pd.notnull(match['AssistantIDs'])]
        asstIDs=match.iloc[0]['AssistantIDs'].split(',')
        hcoach=coaches[coaches['Coach ID'].isin(headID)]
        asstcoaches=coaches[coaches['Coach ID'].isin(asstIDs)]
        coachlist=[]
        for index, row in hcoach.iterrows():
            coachlist.append(row.Lname+', '+row.Fname)
        for index, row in asstcoaches.iterrows():
            coachlist.append(row.Lname+', '+row.Fname)
        teaminfo.append([plalist, planumlist, coachlist])
    for i, team in enumerate(removelist):
        teamlist.remove(team)
    for i, team in enumerate(teamlist):
        thissch=sched[sched['Team']==team]
        maketeamgamecards(team, thissch, teaminfo[i], gctemplate, pastelist)
    return

''' TESTING
info=teaminfo[0]  team=teamlist[0]
row=thissch.loc[0]
'''            
def maketeamgamecards(teamname, thissch, info, gctemplate, pastelist):
    ''' Make xls file full of game card specific to this team
    info contains 1) playerlist (last, first) 2) player number list 3) head (and asst) coach names
    gctemplate - sheet w/ basic structure
    pastelist-  list of items and locations for find replace ''' 
    plalist=info[0]
    planums=info[1]
    planums=[str(i) if str(i)!='nan' else '' for i in planums] # change any np.nans
    coachlist=info[2]
    # need new excel file for this team's cards
    gcname=teamname+'_gamecards.xlsx'
    book=load_workbook(gctemplate)
    '''
    writer=pd.ExcelWriter(gctemplate, engine='openpyxl', datetime_format='mm/dd/yy', date_format='mm/dd/yy')
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    '''
    templsheet=book.active
    thissch=thissch.reset_index(drop=True)
    # TODO generic version w/o game schedule?
    for index, row in thissch.iterrows():
        # Nake a new worksheet in existing opened workbook
        newsheet=book.copy_worksheet(templsheet)
        thistitle='Game'+str(index+1)
        newsheet.title=thistitle
        newsheet.page_margins.left=0.45
        newsheet.page_margins.right=0.45
        newsheet.page_margins.top=0.2
        newsheet.page_margins.bottom=0.2
        newsheet.page_margins.header=0.05
        newsheet.page_margins.footer=0.05
        gamedt=[row.Day + " "+ row.Date+" "+':'.join(row.Time.split(':')[0:2])]
        gdate=row.Date
        gtime=':'.join(row.Time.split(':')[0:2])
        thislocation=[row.Location]
        thisdiv=[row.Division]
        if 'Cabrini' in row.Home:
            homeflag=True
            try:
                oteam=[row.Away.split('/')[0]] # all openpyxl pastes are handled as list
                ocoach=[row.Away.split('/')[1]]
            except: # for JD schedule
                oteam=[row.Away]
                ocoach=[]
        else:
            homeflag=False
            try:
                oteam=[row.Home.split('/')[0]]
                ocoach=[row.Home.split('/')[1]]
            except:
                oteam=[row.Home]
                ocoach=[]
        newsheet=makethisgc(newsheet, pastelist, plalist, planums, coachlist, gamedt, gdate, gtime, 
               thislocation, thisdiv, homeflag, ocoach, oteam, teamname)
    # Setting page margins correctly

    # remove game_generic template and save as "teamname"_gamecards
    book.remove_sheet(book.active)
    book.save(gcname)
    print('Saved game cards for ', teamname)
    return

''' TESTING
row=pastelist.iloc[0]
newsheet['A1'].value
newsheet.cell(row=1,column=1).value='kapow'
newsheet['E42'].value
startrow=10  startcol=5  celldir='Down'
'''

def makethisgc(newsheet, pastelist, plalist, planums, coachlist, gamedt, gdate, gtime, 
               thislocation, thisdiv, homeflag, ocoach, oteam, teamname):
    ''' Handle find replace for this worksheet '''
    # print('Coachlist is', ",".join(coachlist), ' for team', teamname)
    for index, row in pastelist.iterrows():
        if row.Data=='plalist':
            newsheet=pastechunk(newsheet, plalist, row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='planums':
            newsheet=pastechunk(newsheet, planums, row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='coaches':
            newsheet=pastechunk(newsheet, coachlist, row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='location':
            newsheet=pastechunk(newsheet, thislocation, row.Startrow, row.Startcol, row.Direction)        
        elif row.Data=='datetime':
            newsheet=pastechunk(newsheet, gamedt, row.Startrow, row.Startcol, row.Direction)        
        elif row.Data=='date':
            newsheet=pastechunk(newsheet, gdate, row.Startrow, row.Startcol, row.Direction)        
        elif row.Data=='time':
            newsheet=pastechunk(newsheet, gtime, row.Startrow, row.Startcol, row.Direction) 
        elif row.Data=='league':
            newsheet=pastechunk(newsheet, thisdiv, row.Startrow, row.Startcol, row.Direction)        
        elif row.Data=='teamname':
            newsheet=pastechunk(newsheet,  teamname, row.Startrow, row.Startcol, row.Direction)        
        elif row.Data=='headcoach': # used for BBall
            newsheet=pastechunk(newsheet, coachlist[0], row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='asstcoach1' and len(coachlist)>1: # used for BBall
            newsheet=pastechunk(newsheet, coachlist[1], row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='asstcoach2' and len(coachlist)>2: # used for BBall
            newsheet=pastechunk(newsheet, coachlist[2], row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='hcoach':
            if homeflag:
                newsheet=pastechunk(newsheet, coachlist[:1], row.Startrow, row.Startcol, row.Direction)
            else:
                newsheet=pastechunk(newsheet, ocoach, row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='vcoach':
            if homeflag:
                newsheet=pastechunk(newsheet, ocoach, row.Startrow, row.Startcol, row.Direction)
            else:
                newsheet=pastechunk(newsheet, coachlist[:1], row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='hteam':
            if homeflag:
                newsheet=pastechunk(newsheet, ['St. Frances Cabrini'], row.Startrow, row.Startcol, row.Direction)
            else:
                newsheet=pastechunk(newsheet, oteam, row.Startrow, row.Startcol, row.Direction)
        elif row.Data=='vteam':
            if homeflag:
                newsheet=pastechunk(newsheet, oteam, row.Startrow, row.Startcol, row.Direction)
            else:
                newsheet=pastechunk(newsheet, ['St. Frances Cabrini'], row.Startrow, row.Startcol, row.Direction)
    return newsheet

def pastechunk(newsheet, mylist, startrow, startcol, celldir):
    ''' write list to cells in sheet '''
    if type(mylist)==str:
        mylist=[mylist] # convert to list if inadvertantly passed as string
    if celldir=='Down':
        for i, val in enumerate(mylist):
            # using zero-based indexing for rows and columns
            thisrow=startrow+i
            newsheet.cell(row=thisrow, column=startcol).value=mylist[i]
    else:
        for i, val in enumerate(mylist):
            thiscol=startcol+i
            newsheet.cell(row=startrow, column=thiscol).value=mylist[i]
    return newsheet

def detectSMS(recipients):
    '''Determine if primary (first) e-mail address is very likely SMS (9 to 10 leading numbers)'''
    if len(recipients)==0:
        return False
    tempstr=recipients[0].split('@')[0]
    SMSmatch=re.match(r'\d{9}', tempstr) 
    if SMSmatch:
        return True
    else:
        return False

def findmissingcards(team, SUs, carddict):
    ''' Finds players on team without CYC card on file 
    return as comma sep first/last'''
    thisteam=SUs[SUs['Team']==team]
    plakeys=np.ndarray.tolist(thisteam.Plakey.unique())
    plakeys=[str(i) for i in plakeys]
    # Compare as strings (since dict keys are strings)
    missingkeys=[i for i in plakeys if i not in carddict.keys()]
    match=thisteam[thisteam['Plakey'].isin(missingkeys)]
    missinglist=[]
    for index, row in match.iterrows():
        missinglist.append(row.First+' '+row.Last)
    missingstr=','.join(missinglist)
    return missingstr