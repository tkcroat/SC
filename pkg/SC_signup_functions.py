# -*- coding: utf-8 -*-
"""
Created on Sun May 22 10:30:01 2016
SC process signups functions
@author: tkc
"""
#%%
import pandas as pd
import numpy as np
from datetime import datetime, date
import re, glob, math
from openpyxl import load_workbook # writing to Excel
from PIL import Image, ImageDraw, ImageFont
import tkinter as tk

import pkg.SC_config as cnf # _OUTPUT_DIR and _INPUT_DIR
#%%
def combinephrases(mylist):
    ''' Combine list of phrases using commas & and '''
    if len(mylist)==1:
        return str(mylist[0])
    elif len(mylist)==2:
        tempstr=str(mylist[0])+ ' and ' +str(mylist[1])
        return tempstr
    else:
        rest=mylist[:-1]
        rest=[str(i) for i in rest]
        last=mylist[-1]
        tempstr=', '.join(rest) +' and ' + str(last)
        return tempstr#%%

def writetoxls(df, sheetname, xlsfile):
    ''' Generic write of given df to specified tab of given xls file '''
    book=load_workbook(xlsfile)
    writer=pd.ExcelWriter(xlsfile, engine='openpyxl', datetime_format='mm/dd/yy', date_format='mm/dd/yy')
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer,sheet_name=sheetname,index=False) # this overwrites existing file
    writer.save() # saves xls file with all modified data
    return

def loadtransfers(df, signups):
    ''' Load transferred players and add to signups (then run player ID);
    transfers added as normal players but need fake billing entries 
    '''
    df=df.rename(columns={'Fname':'First','Lname':'Last','Street':'Address','Parish':'Parish of Registration'})
    df=df.rename(columns={'Phone':'Phone1','Birthdate':'DOB','Sex':'Gender','Open/Closed':'Ocstatus'})
    # Replace Girl, Boy with m f
    df.loc[:,'Gender']=df.Gender.replace('F','Girl')
    df.loc[:,'Gender']=df.Gender.replace('M','Boy')
    # Manually enter sport
    print('Enter sport for transferred players')
    sport=input()
    df.loc[:,'Sport']=sport
    df=df.dropna(subset=['First']) # remove blank rows if present
    mycols=[col for col in df if col in signups]
    df=df[mycols]
    df=formatnamesnumbers(df)
    # place date/transfer in timestamp
    mystamp=datetime.strftime(datetime.now(),'%m/%d/%y')+' transfer'
    df.loc[:,'Timestamp']=mystamp
    mycols=signups.columns
    signups=signups.append(df, ignore_index=True)
    signups=signups[mycols]
    return signups

def packagetransfers(teams, Mastersignups, famcontact, players, season, year, acronyms, messfile):
    ''' Package roster and contact info by sport- school and save as separate xls files 
    also generate customized e-mails in single log file (for cut and paste send to appropriate persons)
    args:
        teams - loaded team list
        mastersignups - signups w/ team assignment
        players -player DB
        famcontact - family contact db
        season - Fall, Winter or Spring
        year - starting sports year (i.e. 2019 for 2019-20 school year)
        acronyms -  school/parish specific abbreviations
        messfile - e-mail message template w/ blanks
    returns:
        
    '''
    teams=teams[pd.notnull(teams['Team'])]
    transferteams=np.ndarray.tolist(teams[teams['Team'].str.contains('#')].Team.unique())
    transSU=Mastersignups[Mastersignups['Team'].isin(transferteams)]
    # ensure that these are from correct season/year
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    transSU=transSU.loc[(transSU['Sport'].isin(sportlist)) & (transSU['Year']==year)] # season is not in mastersignups... only individual sports
    # get family contact info from famcontacts
    transSU=pd.merge(transSU, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))
    # get school from players.csv
    transSU=pd.merge(transSU, players, how='left', on=['Plakey'], suffixes=('','_r2'))
    # get division from Teams xls (for roster)
    transSU=pd.merge(transSU, teams, how='left', on=['Team'], suffixes=('','_r3')) # effectively adds other team info for roster toall players
    transSU.loc[:,'Role']='Player' # add column for role
    # transSU['Open/Closed']='Closed'

    # Sort by grade pre-split
    transSU.loc[:,'Grade']=transSU.Grade.replace('K',0)
    transSU.loc[:,'Grade']=transSU.Grade.apply(int)
    transSU=transSU.sort_values(['Grade'], ascending=True)
    transSU.loc[:,'Grade']=transSU.Grade.replace(0,'K') # replace K with zero to allow sorting
    # Column for sorting by transferred to school
    transSU.loc[:,'Transchool']=transSU['Team'].str.split('#').str[0]
    grouped=transSU.groupby(['Sport','Transchool'])
    for [sport, school], group in grouped:
        # prepare roster tab
        xlsname=cnf._OUTPUT_DIR+'\\Cabrini_to_'+school+'_'+sport+'_'+str(year)+'.xlsx'
        writer=pd.ExcelWriter(xlsname, engine='openpyxl')
        Transferroster=organizeroster(group)
        Transferroster=Transferroster.sort_values(['Team', 'Sex', 'Grade'], ascending=True)
        Transferroster=replaceacro(Transferroster,acronyms)
        Transferroster.to_excel(writer,sheet_name='roster',index=False)
        # prep contacts tab 
        mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone1', 'Text1','Email1', 'Phone2', 'Text2', 
        'Email2', 'Pfirst1', 'Plast1', 'Pfirst2', 'Plast2', 'Team']
        Transfercontacts=group[mycols]
        Transfercontacts.to_excel(writer, sheet_name='contacts', index=False)
        writer.save()
    # Now generate list of e-mails for all schools/directors 
    logfile='transfers_director_emails_log.txt'
    with open(logfile,'w+') as emaillog:
        # Read generic file to sport director
        with open(messfile, 'r') as file:
            blankmessage=file.read()
        for [sport, school], group in grouped:
            plagroup=group.groupby(['Grade', 'Gender'])
            platypes=[] # list of # of players by grade, gender
            gradedict={'K':'K', 1:'1st', 2:'2nd',3:'3rd',4:'4th',5:'5th',6:'6th', 7:'7th',8:'8th'}
            genderdict={'f':'girls', 'F':'girls','m':'boys','M':'boys'}
            for [grade, gender], group in plagroup:
                numplays=str(int(group['Grade'].count()))
                grname=gradedict.get(grade)
                genname=genderdict.get(gender)
                platypes.append(numplays+' '+grname+' '+genname)
            plalist=combinephrases(platypes)
            thismess=blankmessage.replace('$SCHOOL', school)
            thismess=thismess.replace('$SPORT', sport)
            thismess=thismess.replace('$PLALIST', plalist)
            emaillog.write(thismess)
            emaillog.write('\n\n')
    return

def findcards():
    '''Search ID cards folder and return player # and file link
    cards resized to 450x290 pix jpg in photoshop (scripts-image processor)
    keys are either player number as string or coach CYC ID, vals are links to files'''
    cardlist=glob.glob('%s\\IDcards\\*.jpg' %cnf._OUTPUT_DIR, recursive=True)
    # construct list of [card #, filename]
    nums=[i.split('\\')[-1] for i in cardlist]
    nums=[i.split('_')[0] if '_' in i else i.split('--')[0] for i in nums ]    
    cards={} # dict for card numbers/filenames
    for i,num in enumerate(nums):
        cards.update({num: cardlist[i]})
    return cards

def makethiscard(IDlist, team):
    ''' Passes link to ID card or player name (if missing) From team's list of player numbers (in alphabetical order), find/open card links, and create single image'''
    # make the master image and determine image array size
    margin=10 # pix on all sides
    if len(IDlist)<11: # use 2 x 5 array (horiz)
        wide=2
        high=5
    elif len(IDlist)<13: # 4w x 3 h (vert)
        wide=4
        high=3
    elif len(IDlist)<22: # 3x by 5-7 high (horiz); max 21
        wide=3
        high=math.ceil(len(IDlist)/3)
    else: # more than 21 ... yikes
        wide=3
        high=math.ceil(len(IDlist)/3)
    cardimage = Image.new('RGB', (450*wide+2*margin, 300*high+2*margin), "white") # blank image of correct size
    draw=ImageDraw.Draw(cardimage) # single draw obj for adding missing card names 
    ttfont=ImageFont.truetype('arial.ttf', size=36)
    for i,fname in enumerate(IDlist):
        row=i//high # remainder is row
        col=i%high # mod is correct column
        xpos=margin+row*450
        ypos=margin+col*300        
        try:
            thiscard=Image.open(fname)
            thiscard=thiscard.resize((450, 300), Image.ANTIALIAS)
            cardimage.paste(im=thiscard, box=(xpos, ypos)) # paste w/ xpos,ypos as upper left        
        except: # occurs when "first last" present instead of file name/path
            # blankcard=Image.new('RGB', (450, 300)) # make blank image as placeholder
            draw.text((xpos+50,ypos+100),fname,font=ttfont, fill="red")
    return cardimage
''' TESTING
i=0  team=teamlist[i]
'''
def makeCYCcards(df, players, teams, coaches, season, year, **kwargs):
    ''' From mastersignups and teams, output contact lists for all teams/all sports separately 
    team assignments must be finished 
    args:
        df -- mastersignups dataframe
        players - player info dataframe
        teams - this year's teams csv
        coaches - full coach CYC info list
        season - Fall, Winter or Spring
    kwargs:
        showmissing - True (shows missing player's name); False- skip missing player
        otherSchools - default False (also make card sheets for transferred teams/players)
    kwargs={'showmissing':False}
    missing = makeCYCcards(Mastersignups, players, teams, coaches, season, year, **{'showmissing':True} )
    missing = makeCYCcards(Mastersignups, players, teams, coaches, season, year, **{'showmissing':False} )
    '''
    # Slice by sport: Basketball (null for winter?), Soccer, Volleyball, Baseball, T-ball, Softball, Track) 
    cards=findcards() # dictionary with number: filename combo for existing CYC cards
    df=df[(df['Year']==year)] 
    df=df.reset_index(drop=True)   
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    df=df[df['Sport'].isin(sportlist)] # season is not in mastersignups... only individual sports
    # Make list of teams that need cards (all track and others >1st grade)
    def processGrade(val):
        if val=='K':
            return 0
        else:
            return int(val)
    teams.loc[:,'Grade'] = teams['Grade'].apply(lambda x:processGrade(x))
    if not kwargs.get('otherSchools', False):
        # all transfer teams with contain # (e.g. SMOS#3G) so remove these
        # dropped by default
        teams = teams[~teams['Team'].str.contains('#')]
    # need track teams or any team from grades 2+ 
    cardTeamList= teams[ (teams['Grade']>1) | (teams['Sport']=='Track') ]['Team'].unique()
    df=df[ df['Team'].isin(cardTeamList) ]
    df=df.sort_values(['Last'])
    # plakeys as string will be easiest for below matching
    df.loc[:,'Plakey']=df['Plakey'].astype(int)
    df.loc[:,'Plakey']=df['Plakey'].astype(str)

    def getName(gr, pk):
        # get name from plakey as string
        match=gr[gr['Plakey']==pk]
        name=match.iloc[0]['First'] + ' ' + match.iloc[0]['Last']
        return name
    
    teamgrouped = df.groupby(['Team'])
    missinglist=[] # list of plakeys with missing card        
        
    for team, gr in teamgrouped:
        # keys in card dict are strings
        IDlist = [str(int(i)) for i in gr.Plakey.unique()]
        missinglist.extend([i for i in gr.Plakey.unique() if i not in cards.keys() ])
        if not kwargs.get('showmissing', False):
            # Shows only valid cards, drops missing names 
            IDlist = [ cards.get(i) for i in IDlist if i in cards.keys() ]
            filename='Cards_'+ team +'.jpg'
        else: # show cards and missing name when card image not in IDcards folder
            IDlist = [cards.get(i) if i in cards.keys() else getName(gr, i) for i in IDlist ]
            filename='Cards_'+ team +'_all.jpg'
        # get team's coaches 
        IDlist.extend(getcoachIDs(team, teams, coaches, cards)) # add coach ID image file or first/last if missing
        cardimage =makethiscard(IDlist, team) # directly saved
        # save the card file
        cardimage.save(cnf._OUTPUT_DIR+'\\'+filename)
    missingcards=players[players['Plakey'].isin(missinglist)]
    missingcards=missingcards.sort_values(['Grade','Last'])
    return missingcards

def getcoachIDs(team, teams, coaches, cards):
    ''' Returns CYC IDs for all team's coaches '''
    thisteam=teams[teams['Team']==team]
    IDlist=[]
    if len(thisteam)!=1:
        print(team, 'not found in current teams list')
        return IDlist # blank list
    thisteam=thisteam.dropna(subset=['Coach ID'])
    if len(thisteam)!=1:
        print('Coach ID not found for', team)
        return IDlist # blank list
    if thisteam.iloc[0]['Coach ID']!='': # possibly blank
        thisID=thisteam.iloc[0]['Coach ID'].strip()
        if thisID in cards:
            IDlist.append(cards.get(thisID,'')) # file path to this coach's ID
        else: # get first/last
            thiscoach=coaches[coaches['Coach ID']==thisID]
            if len(thiscoach)==1:
                IDlist.append(thiscoach.iloc[0]['Fname']+' '+thiscoach.iloc[0]['Lname'])
            else:
                print("Couldn't find coach ", thisID)        
    thisteam=thisteam.dropna(subset=['AssistantIDs'])
    if len(thisteam)==1: # grab asst IDs if they exist
        asstIDs=thisteam.iloc[0]['AssistantIDs'] 
        asstIDs=[str(s).strip() for s in asstIDs.split(",")]
        for i, asstID in enumerate(asstIDs):
            if asstID in cards:
                IDlist.append(cards.get(asstID,'')) # found assistant coaches ID card image
            else: # can't find ... get assistant first last
                thisasst=coaches[coaches['Coach ID']==asstID] # matching asst coach row
                if len(thisasst)==1:
                    IDlist.append(thisasst.iloc[0]['Fname']+' '+thisasst.iloc[0]['Lname'])
                else:
                    print("Couldn't find coach ", asstID)        
    return IDlist

def autocsvbackup(df, filename, newback=True):
    ''' Pass df (i.e players for backup and basename (i.e. "family_contact" for file.. finds list of existing backups and keeps ones of 
    certain ages based on targetdates list; 
    can't remember why was newback=False was needed (always true here to make new backup)

    ''' 
    # TODO fix this!
    pass
    return

def parseDate(val):
    '''
    Conversion of date string to datetime.date (always header line 2 40:60)
    Possible date formats: 20180316 (newer style) or 03/15/2018 (older style)
    For NGA files Date format changed from 03/15/2018 to 20180316 (on jday 75 in 2018)
    
    time format: 221100 or 22:11:00 (sometimes w/ UTC)
    not terribly concerned w/ time

    possible date formats: 0) 03/01/2018, 3/1/2018, 3/1/18 or 03/01/18 
    2) 1/1/19 2) 2019-1-1 3) 2019-01-01

    '''
    if not isinstance(val, str):
        return val

    else:
        if ' ' in val: # Remove time substring (but will fail for 3 Oct 2019)
            val=val.split(' ')[0] # strip time substring if present

        patterns=['\d{1,2}/\d{1,2}/\d{2,4}', '\d{4}-\d{1,2}-\d{1,2}', '\d{1,2}-\d{1,2}-\d{4}']

        for i, patt in enumerate(patterns):
            match=re.search(r'%s' %patt, val)
            if match:
                if i==0: # Extract 03/16/2018 (or rarely 28/10/2019 style)
                    try:
                        (mo,dy,yr)=[int(i) for i in val.split('/')]
                        if yr<100 and len(str(yr))==2: # handle 2 digit year
                            yr=int('20'+str(yr))
                        if mo < 13: # normal US version (month first)
                            return datetime(yr, mo, dy).date()
                        # handle day month reversal 
                        elif dy<13: # possible month day reverse
                            print('Month and day reverse for %s' %val)
                            return datetime(yr, dy, mo).date() # Assume month/day switch
                    except:
                        print('Problem extracting date from ', val)
                        return None

                if i==1: # extract 2017-01-01 style (year first)
                    try:
                        (yr,mo,dy)=[int(i) for i in val.split('-')]
                        if mo < 13: # normal US version (month first)
                            return datetime(yr, mo, dy).date()
                        # handle day month reversal 
                        elif dy<13: # possible month day reverse
                            print('Month and day reverse for %s' %val)
                            return datetime(yr, dy, mo).date() # Assume month/day switch                        
                    except:
                        print('Problem extracting date from ', val)
                        return val
                if i==2: # extract 01-01-2019 style (year last)
                    try:
                        (mo,dy,yr)=[int(i) for i in val.split('-')]
                        if mo < 13: # normal US version (month first)
                            return datetime(yr, mo, dy).date()
                        # handle day month reversal 
                        elif dy<13: # possible month day reverse
                            print('Month and day reverse for %s' %val)
                            return datetime(yr, dy, mo).date() # Assume month/day switch                        
                    except:
                        print('Problem extracting date from ', val)
                        return val

def loadProcessPlayerInfo():
    '''Loads and processes players & family contacts (but not signup file)
    args:
        gsignups -- google signups 
        season - 'Fall', 'Winter', 'Spring
        year - 4 digit int (uses fall value all school year.. ie. 2018-19 year is always
        2018)
        ''' 
    players=pd.read_csv(cnf._INPUT_DIR + '\\players.csv', encoding='cp437') # load existing player data (need to find correct DOB format)
    players.loc[:,'Grade']=players.Grade.replace('K',0)
    players.loc[:,'Grade']=players.Grade.replace('pK',0) # just make them 0s for now
    players.loc[:,'Grade']=players.Grade.astype(int)
    # TODO use fixdates if players imports as string (various formats possible)
    # players['DOB']=players['DOB'].apply(lambda x: fixDates(x))
    if not isinstance(players.DOB[0], pd.Timestamp):# sometimes direct import to pd timestamp works, other times not
        players.loc[:,'DOB']=players.DOB.apply(lambda x: parseDate(x))

    famcontact=pd.read_csv(cnf._INPUT_DIR + '\\family_contact.csv', encoding='cp437') # load family contact info
    famcontact=formatnamesnumbers(famcontact)
    
    return players, famcontact

def loadProcessGfiles(gsignups, season, year):
    '''Loads and processes players, family contacts and signup file, gets active 
    season and year 
    args:
        gsignups -- google signups 
        season - 'Fall', 'Winter', 'Spring
        year - 4 digit int (uses fall value all school year.. ie. 2018-19 year is always
        2018)
        ''' 
    players=pd.read_csv(cnf._INPUT_DIR + '\\players.csv', encoding='cp437') # load existing player data (need to find correct DOB format)
    players.loc[:,'Grade']=players.Grade.replace('K',0)
    players.loc[:,'Grade']=players.Grade.replace('pK',0) # just make them 0s for now
    players.loc[:,'Grade']=players.Grade.astype(int)
    # TODO use fixdates if players imports as string (various formats possible)
    # players['DOB']=players['DOB'].apply(lambda x: fixDates(x))
    if not isinstance(players.DOB[0], pd.Timestamp):# sometimes direct import to pd timestamp works, other times not
        players.loc[:,'DOB']=players.DOB.apply(lambda x: parseDate(x))

    famcontact=pd.read_csv(cnf._INPUT_DIR + '\\family_contact.csv', encoding='cp437') # load family contact info
    if season=='Winter':
        gsignups['Sport']='Basketball'
    # TODO determine where multiple sports converted to separate lines
    duplicated=gsignups[gsignups.duplicated(subset=['First', 'Last','Grade','Sport'])]
    if len(duplicated)>0:
        print('Remove duplicate signups for %s' %", ".join(duplicated.Last.unique().tolist()))
    gsignups=gsignups.drop_duplicates(subset=['First', 'Last','Grade','Sport'])
    gsignups.loc[:,'Sport']=gsignups['Sport'].replace({'Volleyball':'VB'}, regex=True)
    missing=[i for i in ['Famkey','Plakey'] if i not in gsignups.columns]
    for col in missing: # add blank vals
        gsignups.loc[gsignups.index, col]=np.nan
    # convert assorted DOB strings to datetime.date
    if not isinstance(gsignups.DOB[0], pd.Timestamp):# sometimes direct import to pd timestamp works, other times not
        gsignups.loc[:,'DOB']=gsignups.DOB.apply(lambda x: parseDate(x))

    # Get year from signup file name    
    outputduplicates(gsignups) # quick check of duplicates output in console window (already removed from signups)
    gsignups=formatnamesnumbers(gsignups) # format phone numbers, names to title case, standardize schools, etc.
    famcontact=formatnamesnumbers(famcontact)
    
    def processGkey(val):
        ''' Some plakey/famkey copied to drive... must convert nan(float), whitespace or 
        number as string to either nan or int
        '''
        if isinstance(val, str):
            val=''.join(val.split(' '))
            if val=='':
                return np.nan
            else:
                try:
                    return int(val)
                except:
                    return np.nan
        else:
            return np.nan
    # ensure gsignups has only int or nan (no whitespace)
    gsignups.loc[:,'Plakey']=gsignups['Plakey'].apply(lambda x: processGkey(x))
    gsignups.loc[:,'Famkey']=gsignups['Famkey'].apply(lambda x: processGkey(x))
    return players, famcontact, gsignups

def loadprocessfiles(signupfile):
    '''Loads and processes players, family contacts and signup file, gets active 
    season and year ''' 
    players=pd.read_csv(cnf._INPUT_DIR + '\\players.csv', encoding='cp437') # load existing player data (need to find correct DOB format)
    players.loc[:,'Grade']=players.Grade.replace('K',0)
    players.loc[:,'Grade']=players.Grade.replace('pK',0) # just make them 0s for now
    players.loc[:,'Grade']=players.Grade.astype(int)
    # TODO use fixdates if players imports as string (various formats possible)
    # players['DOB']=players['DOB'].apply(lambda x: fixDates(x))
    if type(players.DOB[0])!=pd.Timestamp: # sometimes direct import to pd timestamp works, other times not
        try:
            players.loc[:'DOB']=parseDate(players.DOB) # return properly converted date columns series 
        except:
            print('Failure converting player DOB to datetime/timestamp')
    famcontact=pd.read_csv(cnf._INPUT_DIR + '\\family_contact.csv', encoding='cp437') # load family contact info
    # read this season's sports signup file and rename columns
    if signupfile.endswith('.csv'):
        SUraw=pd.read_csv(signupfile)
    elif 'xls' in signupfile:
        try:
            SUraw=pd.read_excel(signupfile, sheetname='Raw') # may or may not have plakey/famkey
        except:
            SUraw=pd.read_excel(signupfile) 
    if SUraw.shape[1]==30 and 'Plakey' in SUraw.columns: 
        SUraw.columns=['Timestamp','First','Last','DOB','Gender','School','Grade',
            'Address','Zip','Parish','Sport','AltPlacement','Ocstatus','Pfirst1',
            'Plast1','Phone1','Text1','Email','Othercontact','Coach','Pfirst2','Plast2',
            'Phone2','Text2','Email2','Coach2','Unisize','Unineed','Plakey','Famkey']
    elif SUraw.shape[1]==28 and 'Plakey' in SUraw.columns: 
        SUraw.columns=['Timestamp','First','Last','DOB','Gender','School','Grade',
            'Address','Zip','Parish','Sport','AltPlacement','Ocstatus','Pfirst1',
            'Plast1','Phone1','Text1','Email','Othercontact','Coach','Pfirst2','Plast2',
            'Phone2','Text2','Email2','Coach2','Plakey','Famkey']
    elif SUraw.shape[1]==26 and 'Plakey' not in SUraw.columns: # Raw value without plakey and famkey
        SUraw.columns=['Timestamp','First','Last','DOB','Gender','School',
            'Grade','Address','Zip','Parish','Sport','AltPlacement','Ocstatus',
            'Pfirst1','Plast1','Phone1','Text1','Email','Othercontact','Coach',
            'Pfirst2','Plast2','Phone2','Text2','Email2','Coach2']
    elif SUraw.shape[1]==28 and 'Plakey' not in SUraw.columns: # Raw value without plakey and famkey
        SUraw.columns=['Timestamp','First','Last','DOB','Gender','School',
            'Grade','Address','Zip','Parish','Sport','AltPlacement','Ocstatus',
            'Pfirst1','Plast1','Phone1','Text1','Email','Othercontact','Coach',
            'Pfirst2','Plast2','Phone2','Text2','Email2','Coach2','Unisize','Unineed']
        SUraw.loc[SUraw.index,'Plakey']=np.nan # add if absent
        SUraw.loc[SUraw.index,'Famkey']=np.nan
    signups=SUraw.drop_duplicates(subset=['First', 'Last','Grade','Sport'])
    signups['Sport'].replace({'Volleyball':'VB'},inplace=True, regex=True)
    # Get year from signup file name    
    season=re.match(r'(\D+)', signupfile).group(0) # season at string beginning followed by year (non-digit)
    if '\\' in season: # remove file path problem
        season=season.split('\\')[-1]
    year=int(re.search(r'(\d{4})', signupfile).group(0)) # full year should be only number string in signups file
    outputduplicates(SUraw) # quick check of duplicates output in console window (already removed from signups)
    signups=formatnamesnumbers(signups) # format phone numbers, names to title case, standardize schools, etc.
    famcontact=formatnamesnumbers(famcontact)
    return players, famcontact, signups, season, year
    
def findavailablekeys(df, colname, numkeys):
    '''Pass df and colname, return a defined number of available keys list
    used for players, families, signups, etc.
    '''
    # list comprehension    
    allnums=[i for i in range(1,len(df))]
    usedkeys=df[colname].unique()
    usedkeys=np.ndarray.tolist(usedkeys)
    availkeys=[i for i in allnums if i not in usedkeys]
    if len(availkeys)<numkeys: # get more keys starting at max+1
        needed=numkeys-len(availkeys)
        for i in range(0,needed):
            nextval=int(max(usedkeys)+1) # if no interior vals are available find next one
            availkeys.append(nextval+i)
    availkeys=availkeys[:numkeys] # truncate and only return the requested number of needed keys 
    return availkeys
        
def organizeroster(df):
    ''' Renaming, reorg, delete unnecessary columns for CYC roster output
    already split by sport and year''' 
    df=df.rename(columns={'First':'Fname','Last':'Lname','Address':'Street','Parish_registration':'Parish of Registration'})
    df=df.rename(columns={'Parish_residence':'Parish of Residence','Phone1':'Phone','DOB':'Birthdate','Gender':'Sex'})
    df=df.rename(columns={'Email1':'Email'})
    # replace Girl, Boy with m f
    df.loc[:,'Sex']=df.Sex.replace('Girl','F').replace('Boy','M')
    df.loc[:,'Sex']=df.Sex.str.upper() # ensure uppercase
    # Convert date format to 8/25/2010 string format
    mycols=['Fname', 'Lname', 'Street', 'City', 'State', 'Zip', 'Phone', 'Email', 'Birthdate', 'Sex', 'Role', 'Division', 'Grade', 'Team', 'School', 'Parish of Registration', 'Parish of Residence', 'Coach ID']
    df=df[mycols] # put back in desired order
    df=df.sort_values(['Team'])
    return df

'''TESTING  row=tempplay.iloc[7]
signups=signups[signups['Last']=='Elston']    
'''

def processdatachanges(signups, players, famcontact, year):
    '''Pass SC signups subset from google drive, update address for more up-to-date 
    contact information, new address, etc. 
    must start here if troubleshooting
    
    args:
        signups -- online signups file (normally google drive)
        players -  player DOB, grade, etc
        famcontact-  family contact info
        year - sports year (int); e.g. 2019 for 2019-20 school year
    '''
    # Using all entries from signups (manual and gdrive)
    # Updates from paper signups should be done directly to famcontact and players csv files (skip entirely)
    '''
    signups.Timestamp=pd.to_datetime(signups.Timestamp, errors='coerce') # converts to naT or timestamp 
    gdsignups=signups.dropna(subset=['Timestamp']) # drops manual entries (no google drive timestamp)
    '''
    # merge w/ players and update grade, recalc grade adjustment, and school
    # must use left merge to keep correct indices from players df (inner causes reindexing)
    players=players.reset_index(drop=True)
    tempplay=pd.merge(players, signups, how='inner', on=['Plakey'], suffixes=('','_n'))
    tempplay=tempplay.dropna(subset=['Gender_n']) # this drops all without a google drive entry

    for index, row in tempplay.iterrows():
        upkwargs={}
        # Skip approval for grade updates
        if row.Grade!=row.Grade_n: # grade discrepancy between players.csv and current signup
            match=players[players['Plakey']==row.Plakey]
            if len(match)==1:
                thisind=match.index[0]
                # update player grade (no approval)
                players.loc[thisind,'Grade']=row.Grade_n # set to new value from current signup file
                print (row.First," ",row.Last," grade changed to ", row.Grade_n)
        if row.School!=row.School_n and str(row.School_n)!='nan':
            upkwargs.update({'school':True})
        # Check for DOB inconsistency between google drive and players.csv
        if row.DOB!=row.DOB_n: # don't change grade adjustment if DOB discrepancy
            if row.DOB_n.year!=year: # skip birthday instead of DOB error
                upkwargs.update({'DOB':True})
        else: # recalculate grade adjustment
            # Direct adjustment to gradeadj in players (if indicated)
            players=updategradeadjust(row, players, year)
        if 'school' in upkwargs or 'DOB' in upkwargs:
            # Interactively approve school or DOB changes
            players=updateplayer_tk(row, players, **upkwargs)
    autocsvbackup(players,'players', newback=True) # run autobackup script
    outname=cnf._OUTPUT_DIR+'\\players.csv'
    players.to_csv(outname,index=False) # direct save of changes from google drive info
    # now update new info into family contacts
    # faminfo=gdsignups.drop_duplicates(subset=['Famkey']) # only process first kid from family 
    faminfo=signups.drop_duplicates(subset=['Famkey']) 
    famcontact=prepcontacts(famcontact)
    faminfo=prepcontacts(faminfo)
    tempfam=pd.merge(famcontact, faminfo, how='inner', on=['Famkey'], suffixes=('','_n')) # same indices as famcontact
    tempfam=tempfam.dropna(subset=['Zip_n']) # drops those without timestamped google drive entry 
    for index,row in tempfam.iterrows():
        # Update/reshuffle phone, email, parent list, parish of registration (direct to famcontact)
        famcontact=update_contact(row, famcontact) # update/reshuffle phone,text (list of lists)
    autocsvbackup(famcontact,'family_contact', newback=True) # run autobackup script
    outname=cnf._INPUT_DIR+'\\family_contact.csv'
    famcontact.to_csv(outname, index=False)
    return players, famcontact

def updatefamcon_tk(row, famcontact, **upkwargs):
    ''' Interactive approval of family contact changes
    changes directly made to famcontacts (but not yet autosaved)
    upkwargs: phone, email, address
    '''    
    root = tk.Tk()
    root.title('Update family contact info')
    choice=tk.StringVar() # must be define outside of event called functions
    rownum=0
    mytxt='Family: '+row.Family+' # '+str(row.Plakey)
    tk.Label(root, text=mytxt).grid(row=rownum, column=0)
    tk.Label(root, text='Deselect to remove').grid(row=rownum, column=1)
    rownum+=1
    
    # Use listbox of common schools?
    if 'parlist' in upkwargs: # indicates new parent found
        colnum=0
        parlist=upkwargs.get('parlist',[])
        # Checkboxes to add new parent
        if 'newpar1' in upkwargs:
            addpar1=tk.BooleanVar()
            addpar1.set(True)
            try:
                mytext='Add parent: '+ (' '.join(upkwargs.get('newpar1',[]))+'?')
            except:
                print('Error adding parent 1', )
                mytext=''
            tk.Checkbutton(root, variable=addpar1, text=mytext).grid(row=rownum, column=colnum)
            colnum+=1
        if 'newpar2' in upkwargs:
            addpar2=tk.BooleanVar()
            addpar2.set(True)
            try:
                mytext='Add parent: '+ (' '.join(upkwargs.get('newpar2',[]))+'?')
            except:
                mytext=''
            tk.Checkbutton(root, variable=addpar2, text=mytext).grid(row=rownum, column=colnum)
            colnum+=1
        # Checkbutton and boolvar for each parent (default true)
        pbools=[] # List of bools for parent inclusion
        for i in range(0,len(parlist)):
            pbools.append(tk.BooleanVar())
            pbools[i].set(True)
            tempstr=parlist[i]
            tk.Checkbutton(root, variable=pbools[i], text=tempstr).grid(row=rownum, column=colnum)
            rownum+=1
    rownum+=1
    if 'emails' in upkwargs: # indicates new parent found
        emaillist=upkwargs.get('emails',[])
        # Checkboxes to add new parent
        colnum=0
        if 'email1' in upkwargs:
            addemail1=tk.BooleanVar()
            addemail1.set(True)
            email1=tk.StringVar()
            email1.set(upkwargs.get('email1',''))
            tk.Checkbutton(root, variable=addemail1, text='Add new email1').grid(row=rownum, column=colnum)
            rownum+=1
            tk.Entry(root, textvariable=email1).grid(row=rownum, column=colnum)
            rownum+=1
        if 'email2' in upkwargs:
            addemail2=tk.BooleanVar()
            addemail2.set(True)
            email2=tk.StringVar()
            email2.set(upkwargs.get('email2',''))
            tk.Checkbutton(root, variable=addemail2, text='Add new email2').grid(row=rownum, column=colnum)
            rownum+=1
            tk.Entry(root, textvariable=email2).grid(row=rownum, column=colnum)
            colnum+=1
        # Checkbutton and boolvar for each email (default true)
        ebools=[] # List of bools for parent inclusion
        for i in range(0,len(emaillist)):
            ebools.append(tk.BooleanVar())
            tempstr=emaillist[i]
            ebools[i].set(True)
            tk.Checkbutton(root, variable=ebools[i], text=tempstr).grid(row=rownum, column=colnum)
            rownum+=1
    rownum+=1
    if 'phones' in upkwargs: # indicates new parent found
        phlist=upkwargs.get('phones',[])
        # Checkboxes to add new parent
        colnum=0
        if 'phone1' in upkwargs:
            addphone1=tk.BooleanVar()
            addphone1.set(True)
            try:
                mytext='Add phone/text: '+ upkwargs.get('phone1','')
            except:
                mytext=''
            tk.Checkbutton(root, variable=addphone1, text=mytext).grid(row=rownum, column=colnum)
            colnum+=1
        if 'phone2' in upkwargs:
            addphone2=tk.BooleanVar()
            addphone2.set(True)
            try:
                mytext='Add phone/text: '+ ', '.join(upkwargs.get('phone2',[]))
            except:
                mytext=''
            tk.Checkbutton(root, variable=addphone2, text=mytext).grid(row=rownum, column=colnum)
            colnum+=1
        # Checkbutton and boolvar for each email (default true)
        phbools=[] # List of bools for parent inclusion
        for i in range(0,len(phlist)):
            phbools.append(tk.BooleanVar())
            tempstr=phlist[i]
            phbools[i].set(True)
            tk.Checkbutton(root, variable=phbools[i], text=tempstr).grid(row=rownum, column=colnum)
            rownum+=1
        
    if 'address' in upkwargs:
        colnum=0
        tk.Label(root, text='Possible change of address').grid(row=rownum, column=colnum)
        rownum+=1
        newaddrbool=tk.BooleanVar()
        newaddr=tk.StringVar()
        newaddrbool.set(False)
        newaddr.set(row.Address_n)
        newzip=tk.StringVar()
        try:
            newzip.set(int(row.Zip_n))
        except:
            print('Non-standard zip value',str(row.Zip_n))
        tk.Checkbutton(root, variable=newaddrbool, text='Change address?').grid(row=rownum, column=colnum)
        colnum+=1
        tk.Label(root, text='Current address').grid(row=rownum, column=colnum)
        colnum=0
        rownum+=1
        tk.Entry(root, textvariable=newaddr).grid(row=rownum, column=colnum)
        rownum+=1
        tk.Entry(root, textvariable=newzip).grid(row=rownum, column=colnum)
        colnum+=1
        tempstr=str(row.Address)+' '+str(row.Zip)
        tk.Label(root, text=tempstr).grid(row=rownum, column=colnum)
    rownum+=1

    # Now set up select/close buttons
    def skip(event):
        choice.set('skip')        
        root.destroy()        
    def change(event):
        choice.set('change')        
        root.destroy()
    
    f=tk.Button(root, text='Skip')
    f.bind('<Button-1>', skip)
    f.grid(row=rownum, column=0)
    g=tk.Button(root, text='Change')
    g.bind('<Button-1>', change)
    g.grid(row=rownum, column=1)
    root.mainloop()
    
    mychoice=choice.get()
    if mychoice=='change':
        # Find matching row for family (needed for all changes below)
        famkey=row.Famkey
        match=famcontact[famcontact['Famkey']==famkey]
        if len(match)==1:
            thisind=match.index[0]
        else: 
            print('Problem finding unique entry for famkey', str(famkey))
            return famcontact # return unaltered
        # Reconstruct parent list
        if 'parlist' in upkwargs:
            newparlist=[] # constructing entirely new parent list from checkbox choices
            if 'newpar1' in upkwargs:
                if addpar1.get():
                    newparlist.append(upkwargs.get('newpar1',[np.nan,np.nan]))
                    #TODO fix nan error
                    print('Added parent',' '.join(upkwargs.get('newpar1')),' to ',str(row.Family))
            for i, val in enumerate(pbools):
                if pbools[i].get():
                    newparlist.append(parlist[i]) # [first, last] format
            if 'newpar2' in upkwargs:
                if addpar2.get():
                    newparlist.append(upkwargs.get('newpar2',[np.nan,np.nan]))
                    print('Added parent 2',' '.join(upkwargs.get('newpar2')),' to ',str(row.Family))
            # Now direct update of parents in this family's famcontact entry
            newparlist=newparlist[0:3] # limit to 3 entries
            while len(newparlist)<3:
                newparlist.append([np.nan,np.nan]) # pad with nan entries if necessary 
            # now reset parent name entries
            for i in range(1,4): # reset 3 existing parents entries
                fname='Pfirst'+str(i)
                lname='Plast'+str(i)
                famcontact.loc[thisind, fname] = newparlist[i-1][0]
                famcontact.loc[thisind, lname] = newparlist[i-1][1]
        # Reconstruct email list
        if 'emails' in upkwargs:
            newemaillist=[]
            if 'email1' in upkwargs:
                if addemail1.get():
                    newemaillist.append(email1.get())
                    print('Added email1', email1.get(), ' to ', str(row.Family))
            for i, val in enumerate(ebools):
                if ebools[i].get():
                    newemaillist.append(emaillist[i])
            if 'email2' in upkwargs:
                if addemail2.get():
                    # insert in 2nd position
                    newemaillist.insert(1, email2.get())
                    print('Added email2', email2.get(), ' to ', str(row.Family))
            # Now update emails in famcontact entry
            # Direct update of parent list
            newemaillist=newemaillist[0:3] # limit to 3 entries
            while len(newemaillist)<3:
                newemaillist.append(np.nan) # pad with nan entries if necessary 
            # now reset parent name entries
            for i in range(1,4): # reset 3 existing parents entries
                colname='Email'+str(i)
                famcontact.loc[thisind, colname]= newemaillist[i-1]
        # Reconstruct phone list
        if 'phones' in upkwargs:
            newphlist=[]
            if 'phone1' in upkwargs:
                if addphone1.get():
                    newphlist.append(upkwargs.get('phone1', [np.nan,np.nan]))
                    print('Added phone1', ','.join(upkwargs.get('phone1',[])), ' to ', str(row.Family))
            for i, val in enumerate(phbools):
                if phbools[i].get():
                    newphlist.append(phlist[i])
            # added at end... probably should go 
            if 'phone2' in upkwargs:
                if addphone2.get():
                    # insert in 2nd position
                    newphlist.insert(1, upkwargs.get('phone2',[np.nan,np.nan]))
                    print('Added phone2', ','.join(upkwargs.get('phone2',[])), ' to ', str(row.Family))
            # Now update phone, text in famcontact entry
            newphlist=newphlist[0:4] # limit to 4 entries
            while len(newphlist)<4:
                newphlist.append([np.nan, np.nan]) # pad with nan entries if necessary 
            # now reset parent name entries
            for i in range(1,5): # reset max 4 phone entries
                phname='Phone'+str(i)
                textname='Text'+str(i)
                famcontact.loc[thisind, phname] = newphlist[i-1][0]
                famcontact.loc[thisind, textname] = newphlist[i-1][1]
        # Handle change of address (direct change if approved)
        # Also change associated zip code and reset parish of residence 
        if 'address' in upkwargs:
            if newaddrbool:
                print('Address changed for ', str(row.Family))
                famcontact.loc[thisind, 'Address'] = newaddr.get()
                # Reset parish of residence to nan (manually find and replace)
                famcontact.loc[thisind, 'Parish_residence'] = np.nan
                try:
                    famcontact.loc[thisind,'Zip']=int(newzip.get())
                except:
                    print('Problem converting zip code ', newzip.get())
        # TODO ... handle parish of registration 
    return famcontact

def update_contact(row, famcontact):
    '''Update phone and textable list from google drive entries; 
    google drive raw entries first processed in process_data_changes (then update
    contacts is called)
    row is a merge of existing famcontact info and new signup info
    existing entries from fam_contact listed first;
    pass/modify/return series for family; reorder/replace numbers 
    has fairly long list of changes made w/o interactive approval:
     1) changing order of email or phone numbers (e.g. swap phone1 and phone2)
     2) add phone2 (or email2) if current phone2(email2) is nan
     3) change order of parents (new parent1)
    
    All other changes done w/ interactive approval using update_famcon_tk
    '''
    # [phone, text, order]
    thisfam=row.Family
    match=famcontact[famcontact['Famkey']==row.Famkey]
    if len(match)==1:
        thisind=match.index[0] # correct index for updating this family in famcontacts
    else:
        print(str(row.Family), " not found in famcontacts.. shouldn't happen")
        return famcontact
    upkwargs={} # empty dict for monitoring all changes
    # check for possible change in address (housenum as trigger)
    match1=re.search(r'\d+', row.Address)
    match2=re.search(r'\d+', row.Address_n)
    if match1 and match2:
        num1=match1.group(0)
        num2=match2.group(0)
        if num1!=num2: # change in address number strongly suggestive of actual change
            upkwargs.update({'address':True})
    else:
        print('No address # found for', str(thisfam))
    phonelist=[] # list of lists with number and textable Y/N
    for i in range(1,5): # get 4 existing phone entries (phone1, phone2, etc.)
        phname='Phone'+str(i)
        txtname='Text'+str(i)
        if str(row[phname])!='nan':
            phonelist.append([row[phname],row[txtname]]) # as phone and text y/N
    # New google drive entries will be Phone1_n.. look for phone/text pair in existing list
    if str(row.Phone1_n)!='nan' and [row.Phone1_n,row.Text1_n] in phonelist: # new ones phone is required entry
        # default move of phone1, text1 to top of list - no confirmation
        if [row.Phone1_n,row.Text1_n]!=phonelist[0]: # move if not in first position
            phonelist.insert(0,phonelist.pop(phonelist.index([row.Phone1_n,row.Text1_n])))
            print('Phone 1 changed for ', str(thisfam))
            upkwargs.update({'phchange':True})
    if str(row.Phone1_n)!='nan' and [row.Phone1_n,row.Text1_n] not in phonelist: # new ones phone is required entry
        if [row.Phone1_n, np.nan] in phonelist: # remove if # present but w/o text indication (no confirm)
            phonelist.remove([row.Phone1_n,np.nan])
            phonelist.insert(0,[row.Phone1_n,row.Text1_n]) # insert in first position
            print('Updated phone 1 to', row.Phone1_n,' for ',str(thisfam))
            upkwargs.update({'phchange':True})
        else:
            # phone1 change to be confirmed
            upkwargs.update({'phone1':[row.Phone1_n,row.Text1_n]})
            upkwargs.update({'phones': phonelist})
    if str(row.Phone2_n)!='nan': # check for phone2 entry (with _n suffix)
        if [row.Phone2_n,row.Text2_n] not in phonelist: # add second phone to 2nd position if not present
            if [row.Phone2_n,np.nan] in phonelist: # remove if # present but w/o text indication
                phonelist.remove([row.Phone2_n,np.nan]) 
                phonelist.insert(1,[row.Phone2_n,row.Text2_n])
                print ('Updated phone 2 to ', str(row.Phone2_n), 'for ', str(thisfam))
                upkwargs.update({'phchange':True})
            else: # get approval for phone 2 addition
                upkwargs.update({'phone2':[row.Phone2_n,row.Text2_n]})
                upkwargs.update({'phones': phonelist})
    # Construct existing list of known email addresses
    emaillist=[] 
    for i in range(1,4): # get 3 existing email entries
        emailname='Email'+str(i)
        if str(row[emailname])!='nan':
            emaillist.append(row[emailname].lower())
    # Find new email1 entry in google drive data
    if str(row.Email)!='nan' and '@' in row.Email: # real primary gd named email
        if row.Email.lower() in emaillist: # add in first position if not present (no confirmation)
            if row.Email.lower()!=emaillist[0]: # check if in first position already
                emaillist.insert(0,emaillist.pop(emaillist.index(row.Email)))
                upkwargs.update({'emchange':True})
                print ('Updated email 1 ', str(row.Email.lower()), 'for family', str(thisfam))
        else: # confirm email1 if not present
            upkwargs.update({'email1':row.Email})
            upkwargs.update({'emails':emaillist})
    # look for new email in email2 position and add 
    if str(row.Email2_n)!='nan' and '@' in row.Email2_n:
        if row.Email2_n.lower() not in emaillist: # add second email to 2nd position if not present
            upkwargs.update({'email2':row.Email2_n})
            upkwargs.update({'emails':emaillist})           
    # Update list of parent names (max 3 entries)
    parlist=[] # construct existing list from family contacts
    # skip if all nan for entered parents (non-gd entry)
    for i in range(1,4): # construct existing parents list
        fname='Pfirst'+str(i)
        lname='Plast'+str(i)
        if str(row[fname])!='nan':
            parlist.append([row[fname],row[lname]]) # list of lists [first, last]
    if str(row.Pfirst1_n)!='nan': # skip if parent name is nan
        if [row.Pfirst1_n,row.Plast1_n] in parlist: # reorder in list
            if [row.Pfirst1_n,row.Plast1_n]!=parlist[0]: # check if already in first
                # move to first position (everything else requires approval)
                parlist.insert(0,parlist.pop(parlist.index([row.Pfirst1_n,row.Plast1_n])))
                parlist.insert(0,[row.Pfirst1_n, row.Plast1_n]) # insert in first position
                upkwargs.update({'parchange':True})
        else: # parent not in list (confirm)
            upkwargs.update({'newpar1':[row.Pfirst1_n,row.Plast1_n]})
            upkwargs.update({'parlist':parlist})
    # inserts in first position while simultaneously removing other entry
    if str(row.Pfirst2_n)!='nan': # Check for parent 2 entry
        if [row.Pfirst2_n,row.Plast2_n] not in parlist: # add second phone to 2nd position if not present
            upkwargs.update({'newpar2':[row.Pfirst2_n,row.Plast2_n]})
            upkwargs.update({'parlist':parlist})
    # Save auto-changes in phone to family contacts
    if 'phchange' in upkwargs: # Record altered phonelist in famcontacts
        if 'phones' in upkwargs: # if present in upkwargs, update list 
            upkwargs.update({'phones': phonelist}) # ensure most current copy
        phonelist=phonelist[0:3] # construct proper list
        while len(phonelist)<4:
            phonelist.append([np.nan,np.nan]) # pad with nan entries if necessary 
        for i in range(1,5): # reset 4 existing phone entries
            phname='Phone'+str(i)
            txtname='Text'+str(i)
            famcontact.loc[thisind, phname] = phonelist[i-1][0] # first of tuple is phone
            famcontact.loc[thisind, txtname] = phonelist[i-1][1] # 2nd of tuple is text y/n
        del upkwargs['phchange']
        print('automatic phone changes for', thisfam)
    # Save auto-changes in emails to family contacts
    if 'emchange' in upkwargs: # Record altered phonelist in famcontacts
        if 'emails' in upkwargs: # if present in upkwargs, update list 
            upkwargs.update({'emails': emaillist}) # ensure most current copy
        emaillist=emaillist[0:2] # construct proper list
        while len(emaillist)<3:
            emaillist.append(np.nan) # pad with nan entries if necessary 
        for i in range(1,4): # reset 4 existing phone entries
            emname='Email'+str(i)
            famcontact.loc[thisind, emname] =emaillist[i-1]
        del upkwargs['emchange']
        print('automatic email changes for', thisfam)
    if 'parchange' in upkwargs: # Record altered parents list in famcontacts
        if 'parlist' in upkwargs: # if present in upkwargs, update list 
            upkwargs.update({'parlist': parlist}) # ensure most current copy
        parlist=parlist[0:2] # construct proper list 
        while len(parlist)<3:
            parlist.append(np.nan) # pad with nan entries if necessary (3 total)
        for i in range(1,4): # reset 4 existing phone entries
            fname='Pfirst'+str(i)
            lname='Plast'+str(i)
            try:
                famcontact.loc[thisind, fname] =parlist[i-1][0]
                famcontact.loc[thisind, lname] =parlist[i-1][1]
            except:
                print('Error updating parents for', thisfam)
        del upkwargs['parchange']
        print('automatic parent changes for', thisfam)
    # now check for any changes needing interactive approval
    if len(upkwargs)>0: # something needs interactive approval            
        famcontact=updatefamcon_tk(row, famcontact, **upkwargs)
    return famcontact

def updateplayer_tk(row, players, **upkwargs):
    ''' Interactive approval of player info updates (except date) 
    changes directly made to players (but not yet autosaved)
    called by processdatachanges
    '''
    commonschools=['Cabrini','Soulard','SLPS','Charter','Private']
    root = tk.Tk()
    root.title('Update player info')
    choice=tk.StringVar() # must be define outside of event called functions
    rownum=0
    mytxt='Player:'+row.First+' '+row.Last+' # '+str(row.Plakey)
    tk.Label(root, text=mytxt).grid(row=rownum, column=0)
    rownum+=1
    # Use listbox of common schools?
    if 'DOB' in upkwargs: # indicates discrepancy
        DOB1=date(row.DOB) 
        DOB2=date(row.DOB_n)
        # create and display DOB variables
        def ChooseDOB1(event):
            DOB.set(datetime.strftime(DOB1,'%m/%d/%y'))
        def ChooseDOB2(event):
            DOB.set(datetime.strftime(DOB2,'%m/%d/%y'))  
        DOB=tk.StringVar()
        DOB.set(datetime.strftime(DOB1,'%m/%d/%y')) # defaults to original
        tk.Label(root, text='Update date of birth?').grid(row=rownum, column=0)
        mytxt='current DOB:'+datetime.strftime(DOB1,'%m/%d/%y')
        b=tk.Button(master=root, text=mytxt)
        b.bind('<Button-1>', ChooseDOB1)
        b.grid(row=rownum, column=1)
        mytxt='New DOB:'+datetime.strftime(DOB2,'%m/%d/%y')
        b=tk.Button(master=root, text=mytxt)
        b.bind('<Button-1>', ChooseDOB2)
        b.grid(row=rownum, column=2)            
        tk.Entry(master=root, textvariable=DOB).grid(row=rownum, column=3)
        rownum+=1
    if 'school' in upkwargs:
        school=tk.StringVar()
        school.set(row.School) # default to existing value
        tk.Label(root, text='Update school?').grid(row=rownum, column=0)
        rownum+=1
        def newschool(event):
            school.set(row.School_n)
        def oldschool(event):
            school.set(row.School)
        def pickschool(event):
            # double-click to pick standard school choice
            items=lb.curselection()[0] # gets selected position in list
            school.set(commonschools[items])
        tk.Entry(root, textvariable=school).grid(row=rownum, column=2)
        mytxt='new school:'+str(row.School_n)
        b=tk.Button(master=root, text=mytxt)
        b.bind('<Button-1>', newschool)
        b.grid(row=rownum, column=1)
        mytxt='existing school:'+str(row.School)
        b=tk.Button(master=root, text=mytxt)
        b.bind('<Button-1>', oldschool)
        b.grid(row=rownum, column=0)
        # also include selectable listbox of common school choices
        lb=tk.Listbox(master=root, selectmode=tk.SINGLE)
        lb.bind("<Double-Button-1>", pickschool)
        lb.grid(row=rownum, column=3)
        for i,sch in enumerate(commonschools):
            lb.insert(tk.END, sch)
        rownum+=1
    # Now set up select/close buttons
    def skip(event):
        choice.set('skip')        
        root.destroy()        
    def change(event):
        choice.set('change')        
        root.destroy() 
    
    f=tk.Button(root, text='Skip')
    f.bind('<Button-1>', skip)
    f.grid(row=rownum, column=0)
    g=tk.Button(root, text='Change')
    g.bind('<Button-1>', change)
    g.grid(row=rownum, column=1)
    root.mainloop()
    
    mychoice=choice.get()
    if mychoice=='change':
        try:
            # make changes directly to players after finding correct index using plakey
            plakey=row.Plakey
            match=players[players['Plakey']==plakey]
            thisind=match.index[0]
            if 'school' in upkwargs:
                players.loc[thisind,'School']= school.get()
            if 'DOB' in upkwargs:
                newDOB=datetime.strptime(DOB.get(),'%m/%d/%y')
                players.loc[thisind,'DOB']= newDOB
        except:
            print('Error updating info for', row.Plakey, row.First, row.Last)
    return players

def prepcontacts(df):
    ''' Prepare for update contacts/ matching with google drive info 
    avoids possible problems/spaces in manually entered info '''
    mycols=['Pfirst1', 'Plast1','Pfirst2', 'Plast2', 'Pfirst3', 'Plast3',
    'Phone1', 'Text1','Phone2', 'Text2', 'Phone3', 'Text3', 'Phone4', 
    'Text4', 'Email1','Email2', 'Email3']
    for i, col in enumerate(mycols):
        try:
            df.loc[:,col]=df[col].str.strip()
        except: # maybe only nan or not present (i.e. in signups)
            pass
    mycols=['Text1','Text2','Text3']
    for i, col in enumerate(mycols):
        try:
            df.loc[:,col]=df[col].str.replace('No','N', case=False)
            df.loc[:,col]=df[col].str.replace('Yes','Y', case=False)
        except:
            pass
    return df

def findyearseason(df):
    ''' Pass raw signups and determine year and sports season '''
    # get year from system clock and from google drive timestamp    
    now=datetime.now()
    val=df.Timestamp[0] # grab first timestamp    
    if val!=datetime: # if not a timestamp (i.e. manual string entry find one
        while type(val)!=datetime:
            for index, row in df.iterrows():
                val=df.Timestamp[index]
    year=val.year # use year value from signup timestamps
    if now.year!=val.year:
        print ('Possible year discrepancy: Signups are from ',str(val.year))
    # now find sports season
    mask = np.column_stack([df['Sport'].str.contains("occer", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Fall'
    mask = np.column_stack([df['Sport'].str.contains("rack", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Spring'
    mask = np.column_stack([df['Sport'].str.contains("asket", na=False)])
    if len(df.loc[mask.any(axis=1)])>0:
        season='Winter'   
    return season, year

def outputduplicates(df):
    '''Prints out names of players with duplicated entries into console... can then delete from google drive signups '''
    tempdf=df[df.duplicated(['First','Last','Sport'])] # series with 2nd of duplicated entries as True
    firsts=tempdf.First.tolist()
    lasts=tempdf.Last.tolist()
    for f,l in zip(firsts, lasts):
        print('Duplicated signup for player: {} {}'.format(f,l))
    return

def formatphone(df):
    ''' Convert all entered phone numbers in dfs phone columns to 314-xxx-xxxx string and standardize text field '''
    
    def phoneFormat(val):
        # lambda function phone number reformatting
        if not isinstance(val, str):
            return val
        # replace/remove any white space
        val="".join(val.split(' '))
        if val=='': # blank phone causes problems
            return np.nan
        if not re.search(r'(\d+-\d+-\d+)', val):
            val=re.sub("[^0-9]", "", val) # substitute blank for non-number
            if len(val)==7:
                return '314'+val
            elif len(val)==11 and val.startswith('1'): # remove starting 1 if present
                return val[1:11]
            elif len(val)!=10: # sometimes has --- 
                # print('Bad number: ',val)
                return val
            else:
                return val[0:3]+'-'+val[3:6]+'-'+val[6:10]
        else:
            return val # already good
    # find phone columns (named phone, phone2, etc.)
    phlist=[str(s) for s in df.columns if 'Phone' in s]
    for col in phlist:
        df.loc[:,col]=df[col].apply(lambda x: phoneFormat(x))
    # now change yes in any text field to Y
    txtlist=[str(s) for s in df.columns if 'Text' in s]
    for col in txtlist:
        df.loc[:,col]=df[col].replace('yes','Y')
        df.loc[:,col]=df[col].replace('Yes','Y')
    return df
    
def standardizeschool(df):
    ''' can pass any frame with school column and standardize name as Cabrini and Soulard''' 
    schstr='frances' + '|' + 'cabrini' + '|' + 'sfca' # multiple school matching string    
    tempdf=df[df['School'].str.contains(schstr, na=False, case=False)]
    df.loc[tempdf.index,'School']='Cabrini'
    tempdf = df[df['School'].str.contains('soulard', na=False, case=False)]
    df.loc[tempdf.index,'School']='Soulard'
    tempdf = df[df['School'].str.contains('public', na=False, case=False)]
    df.loc[tempdf.index,'School']='Public'
    schstr='city garden' + '|' + 'citygarden'  # multiple school matching string    
    tempdf = df[df['School'].str.contains(schstr, na=False, case=False)]
    df.loc[tempdf.index,'School']='City Garden'
    return df
      
def formatnamesnumbers(df):
    '''Switch names to title case, standardize gender, call phone/text reformat and standardize school name''' 
    def titleStrip(val):
        try:
            return val.title().strip()
        except:
            return val
    processCols=['First','Last','Family','Pfirst1','Plast1','Pfirst2','Plast2','Email','Email2']
    processCols=[i for i in processCols if i in df.columns]
    for col in processCols:
        df.loc[:, col]=df[col].apply(lambda x: titleStrip(x))

    if 'Gender' in df:        
        df.loc[:,'Gender']=df.Gender.replace('Girl','f')
        df.loc[:,'Gender']=df.Gender.replace('Boy','m')  
    if 'Grade' in df:
        df.loc[:,'Grade']=df.Grade.replace('K',0)
        df.loc[:,'Grade']=df.Grade.replace('pK',0)
        try:
            df.loc[:,'Grade']=df.Grade.astype(int)
        except:
            print('Player grade likely missing from raw signup file... enter manually')
    df=formatphone(df) # call phone reformatting string

    if 'School' in df:
        df=standardizeschool(df) # use "Cabrini" and "Soulard" as school names
    return df

def graduate_players(players, year):
    ''' Recalc grade based on grade adjustment, school year (run once per year in fall) and age.  
    some player grades will already have been updated (generally google drive entries)... however recalc shouldn't
    change grade ''' 
    players.loc[:,'Grade']=players.Grade.replace('K',0)
    for index,row in players.iterrows():
        # replace K with zero         
        grade=int(players.iloc[index]['Grade']) # get currently listed grade
        gradeadj=players.iloc[index]['Gradeadj']
        dob=players.iloc[index]['DOB']
        if str(gradeadj)=='nan' or str(dob)=='NaT': # skip grade update if info is missing
            continue
        dob=date(dob)
        # calculate current age at beginning of this school on 8/1 
        age=date(year,8,1)-dob
        age = (age.days + age.seconds/86400)/365.2425 
        # assign grade based on age (and grade adjustment)
        newgrade=int(age)+int(gradeadj)-5
        if grade!=newgrade:
            first=players.iloc[index]['First']
            last=players.iloc[index]['Last']
            print('Grade changed from',grade,'to',newgrade,'for', first, last)
            players.loc[index, 'Grade'] = newgrade
    players.loc[:,'Grade']=players.Grade.replace(0,'K')
    return players

def removeEmptyFams(players, famcontact):
    '''
    Remove empty families (with no remaining players)
    '''
    # Remove families with no active players
    plaset=[int(i) for i in list(players.Famkey.unique())]
    famset=[int(i) for i in list(famcontact.Famkey.unique())]
    # Empty families 
    emptykey=[i for i in famset if i not in plaset]
    empty=famcontact[famcontact['Famkey'].isin(emptykey)]
    print('Remove empty families:')
    for ind, row in empty.iterrows():
        print(row.Family, ':',row.Pfirst1, row.Plast1)
    choice=input("Remove empty families (Y,N)?\n")
    if choice.upper()=='Y':
        famcontact=famcontact[~famcontact['Famkey'].isin(emptykey)]
    outname=cnf._INPUT_DIR+'\\family_contact.csv'
    famcontact.to_csv(outname, index=False)
    return famcontact

def removeHSkids(players):
    ''' Drop graduated players (9th graders) from list '''
    grlist=[i for i in range(0,9)]
    grlist.append('K')
    Hs=players.loc[~(players.Grade.isin(grlist))]
    for ind, row in Hs.iterrows():
        print(row.First, row.Last)
    choice=input('Remove above HS players (Y/N)?\n')
    if choice.upper()=='Y':
        players=players.loc[(players.Grade.isin(grlist))]
        print('HS Players removed but not autosaved')
    return players


def estimategrade(df, year):
    '''Estimate grade for this sports season based on DOB.. not commonly used ''' 
    for index, row in df.iterrows():        
        grade=df.loc[index]['Grade']
        if str(grade)=='nan': # skips any players who already have assigned grade
            dob=df.loc[index]['DOB']
            dob=date(dob) # convert to datetime date from timestamp
            first=df.loc[index]['First']
            last=df.loc[index]['Last']
            if str(dob)=='nan':
                print ('DOB missing for ', first,' ', last)
                continue # skip to next if dob entry is missing
            currage=date(year,8,1) - dob 
            currage = (currage.days + currage.seconds/86400)/365.2425 # age on first day of school/ sports season
            gradeest=int(currage-5)
            if gradeest==0:
               gradeest='K' 
            print(first, last, 'probably in grade', gradeest)
            df.loc[index,'Grade']=gradeest
    return df

def updateoldteams(teams, year):
    ''' Load old teams after copy to teams tab in teams_coaches, then auto-update year-grade 
    must be manually saved with saveteams... then any adjustments made manually in Excel'''
    # check to ensure teams are not already updated
    if teams.iloc[0]['Year']==year:
        print('Teams already updated for ', year,' school year')
        return teams # pass back unaltered
    # temporarily make the K to 0 replacements
    teams.Grade=teams.Grade.replace('K',0)
    teams.loc[:'Graderange']=teams['Graderange'].astype(str) # convert all to string
    teams.loc[:,'Year']=year
    teams.loc[:,'Grade']+=1    
    for index, row in teams.iterrows():
        grade=teams.loc[index]['Grade']
        div=teams.loc[index]['Division'] # division must match grade
        div=div.replace('K','0')  # replace any Ks in string
        newdiv=''.join([s if not s.isdigit() else str(grade) for s in div]) # find replace for unknown # w/ new grade
        teams.loc[index,'Division'] = newdiv
        cycname=teams.loc[index]['Team'] # update grade portion of team name
        if cycname.startswith('K'):
            newcycname='1'+ cycname[1:]
            teams.loc[index,'Team'] = newcycname
        elif cycname[0].isdigit(): # now update teams beginning w/ numbers
            newcycname=str(grade)+ cycname[1:]
            teams.loc[index,'Team']= newcycname
        # update grade ranges
        grrange=teams.loc[index]['Graderange'] # should be all numbers
        grrange=grrange.replace('K','0')
        newrange=''.join([str(int(i)+1) for i in grrange])
        teams.loc[index,'Graderange'] = newrange # grade range stored as string, right?
        # no auto-save... save with saveteams after checking for proper changes
    return teams

def splitcoaches(df):
    ''' Pass CYC teams list, split and duplicate rows with comma separated vals in colname for extra coaches'''    
    df.loc[:,'Role']='Coach' # add col for head or asst (first entry for head coach)
    # df['Open/Closed']='Closed'
    assistants=df.dropna(subset=['AssistantIDs']) # drop teams w/ no asst coaches
    for index, rows in assistants.iterrows():
        val=assistants.loc[index,'AssistantIDs']
        asstcoaches=[str(s) for s in val.split(',')] #list of assistants for single team 
        for i,asst in enumerate(asstcoaches):
            newrow=assistants.loc[index] # duplicate entry as series
            asst=asst.strip() # strip leading, trailing blanks
            newrow.loc['Coach ID'] = asst # set this asst coaches ID 
            newrow.loc['Role'] = 'Assistant Coach'
            df=df.append(newrow)
    df=df.sort_values(['Team'],ascending=True) 
    return df

def addcoachestoroster(teams, coaches):
    '''Creates roster entries for coaches for each CYC team
    pass teams and coaches (with coach roster info)
    needed roster cols are all below (except sport used in output parsing)
    args: teams -- team table w/ head and asst coach CYC ids 
        coaches - coaches table with CYC Id (key) and associated info
    returns: coachroster --separate df to be appended to main player roster
        '''
    # Add team coaches (match by CYC-IDs)
    thismask = teams['Team'].str.contains('-', case=False, na=False) # finds this season's CYC level teams
    CYCcoach=teams.loc[thismask] # also has associated sport
    CYCcoach=splitcoaches(CYCcoach) # makes new row for all assistant coaches on CYC teams
    CYCcoach=pd.merge(CYCcoach, coaches, how='left', on=['Coach ID'], suffixes=('','_r')) 
    mycols=['Sport','Fname', 'Lname', 'Street', 'City', 'State', 'Zip', 'Phone', 'Email', 'Birthdate', 'Sex', 'Role', 'Division', 'Grade', 'Team', 'School', 'Parish of Registration', 'Parish of Residence', 'Coach ID']
    for col in [col for col in mycols if col not in CYCcoach.columns]:
        CYCcoach[col]='' # birthdate generally missing
    CYCcoach=CYCcoach[mycols] # put back in desired order
    # drop duplicates on CYC ID, team (sometimes occurs during merge)
    CYCcoach=CYCcoach.drop_duplicates(['Coach ID','Team']) 
    return CYCcoach

def countteamplayers(df, teams, season, year):
    ''' For each team, summarize number of players (subset those that are younger or older) and list of names 
    passing mastersignups'''
    df=df[df['Year']==year] # removes possible naming ambiguity 
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season,[])
    df=df[df['Sport'].isin(sportlist)] # only this sports season
    df.Grade=df.Grade.replace('K',0)
    df.Grade=df.Grade.astype('int')  
    teams.loc[:,'Grade']=teams.Grade.replace('K',0)
    teams.loc[:,'Grade']=teams.Grade.astype('int')
    teams.loc[:,'Playerlist']=teams.Playerlist.astype('str')
    for index, row in teams.iterrows():
        teamname=teams.loc[index]['Team']
        match=df[df['Team']==teamname] # all players on this team from master_signups
        teams.loc[index,'Number'] = len(match) # total number of players		
        # compose player list (First L.) and add to teams
        playerlist=[]
        for ind, ro in match.iterrows():
            first=match.loc[ind]['First']
            last=match.loc[ind]['Last']
            strname=first+' ' +last[0]
            playerlist.append(strname)
        players=", ".join(playerlist)
        teams.loc[index,'Playerlist'] = players
        # count players above or below grade level
        thisgrade=int(teams.loc[index]['Grade'])
        teams.loc[index,'Upper'] = (match.Grade > thisgrade).sum()
        teams.loc[index,'Lower'] = (match.Grade < thisgrade).sum()
    writetoxls(teams, 'Teams', 'Teams_coaches.xlsx')
    return teams

def writecontacts(df, famcontact, players, season, year):
    ''' From mastersignups and teams, output contact lists for all teams/all sports separately '''
    # Slice by sport: Basketball (null for winter?), Soccer, Volleyball, Baseball, T-ball, Softball, Track) 
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    df=df.loc[(df['Sport'].isin(sportlist)) & (df['Year']==year)] # season is not in mastersignups... only individual sports
    
    '''# for transfers to same school (but different grades), combine all into single list for given school    
    for index,row in df.iterrows():
        if str(df.loc[index]['Team'])!='nan': # avoids nan team screwups
            if '#' in df.loc[index]['Team']: # this combines Ambrose#2B, Ambrose#3G to single tab
                df=df.set_value(index,'Team',df.loc[index]['Team'].split('#')[0]) 
    '''
    # get family contact info from famcontacts
    df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))
    # get school from players.csv
    df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))
    # Sort by grade pre-split
    df.loc[:,'Grade']=df.Grade.replace('K',0)
    df.loc[:,'Grade']=df.Grade.apply(int)
    df=df.sort_values(['Grade'], ascending=True)
    df.loc[:,'Grade']=df.Grade.replace(0,'K') # replace K with zero to allow sorting
    df.loc[:,'Team']=df.Team.replace(np.nan,'None') # still give contacts if team not yet assigned
    df.loc[:,'Team']=df.Team.replace('','None')
    # Standard sport contacts output for soccer, VB, basketball
    if season!='Spring':
        for i, sport in enumerate(sportlist):
            fname=cnf._OUTPUT_DIR+'\\'+sport+'_'+str(year)+'_contacts.xlsx'
            writer=pd.ExcelWriter(fname, engine='openpyxl')
            Thissport=df[df['Sport']==sport]
           
            teamlist= Thissport.Team.unique()  
            teamlist=np.ndarray.tolist(teamlist)
            # Combine transfers to same school
            transchools=[s.split('#')[0] for s in teamlist if '#' in s]
            teamlist=[s for s in teamlist if '#' not in s]
            teamlist.extend(transchools) # all to same school as single "team"
            # now can organize contacts (and drop sport)
            mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone1', 'Text1','Email1', 'Phone2', 'Text2', 
            'Email2', 'Team', 'Pfirst1', 'Plast1', 'Pfirst2', 'Plast2', 'Plakey', 'Famkey', 'Family']
            Thissport=Thissport[mycols] # drop columns and rearrange
            for i, team in enumerate(teamlist):
                thisteam=Thissport[Thissport['Team'].str.contains(team)]
                thisteam.to_excel(writer,sheet_name=team,index=False) # this overwrites existing file
            writer.save()
    else: # handle spring special case
        Balls=df[df['Sport']!='Track'] # all ball-bat sports together
        mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone1', 'Text1','Email1', 'Phone2', 'Text2', 
        'Email2', 'Team', 'Pfirst1', 'Plast1', 'Pfirst2', 'Plast2', 'Plakey', 'Famkey', 'Family']

        Balls=Balls[mycols] 
        teamlist= Balls.Team.unique()  
        teamlist=np.ndarray.tolist(teamlist)
        # Combine transfers 
        transchools=[s.split('#')[0] for s in teamlist if '#' in s]
        teamlist=[s for s in teamlist if '#' not in s]
        teamlist.extend(transchools) # all to same school as single "team"
        fname=cnf._OUTPUT_DIR+'\\'+'Batball'+'_'+str(year)+'_contacts.xlsx'
        writer=pd.ExcelWriter(fname, engine='openpyxl')
        # create a separate tab for each team and write the contacts
        for i, team in enumerate(teamlist):
            thisteam=Balls[Balls['Team'].str.contains(team)]
            thisteam.to_excel(writer,sheet_name=team,index=False) # this overwrites existing file
        writer.save() # overwrites existing
        # Entire track team as single file
        Track=df[df['Sport']=='Track']
        Track=Track[mycols] # drop columns and rearrange        
        fname=cnf._OUTPUT_DIR+'\\'+'Track'+'_'+str(year)+'_contacts.xlsx'
        writer=pd.ExcelWriter(fname, engine='openpyxl')
        Track.to_excel(writer,sheet_name='Track',index=False) 
        writer.save() 
    return

def makegoogcont(df, famcontact, players, season, year):
    '''Create and save a google contacts file for all Cabrini teams
    save to csv '''
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    df=df.loc[(df['Sport'].isin(sportlist)) & (df['Year']==year)] # season is not in mastersignups... only individual sports
    
    '''# for transfers to same school (but different grades), combine all into single list for given school    
    for index,row in df.iterrows():
        if str(df.loc[index]['Team'])!='nan': # avoids nan team screwups
            if '#' in df.loc[index]['Team']: # this combines Ambrose#2B, Ambrose#3G to single tab
                df=df.set_value(index,'Team',df.loc[index]['Team'].split('#')[0]) 
    '''
    # get family contact info from famcontacts
    df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))
    # get school from players.csv
    df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))
    # Drop any players not yet assigned
    df=df.dropna(subset=['Team'])
    # Full contacts list format for android/google

    for i, sport in enumerate(sportlist):
        Thissport=df[df['Sport']==sport]
        teamlist= Thissport.Team.unique()  
        teamlist=np.ndarray.tolist(teamlist)
        # drop if team is not yet assigned
        teamlist=[s for s in teamlist if str(s) != 'nan']
        # drop if team is 'drop'
        teamlist=[s for s in teamlist if str(s) != 'drop']
        # Drop all non-Cabrini transferred teams (which must contain #)
        teamlist=[s for s in teamlist if '#' not in s]
        # Combine track subteams to single team
        teamlist=[s[0:5] if 'Track' in s else s for s in teamlist]
        teamlist=set(teamlist)
        teamlist=list(teamlist)
        # now create google contacts list for each Cabrini team and save
        for j, team in enumerate(teamlist):
            thisteam=Thissport[Thissport['Team'].str.contains(team)]
            # Drop duplicate from same family
            thisteam=thisteam.drop_duplicates('Phone1')
            thisteam.loc[:,'Name']=thisteam['First']+' '+thisteam['Last']
            thisteam.loc[:,'Group']=sport+str(year)
            mycols=['Name','Pfirst1','Last','Phone1','Phone2','Email1','Email2','Group']
            newcols=['Name','Additional Name','Family Name','Phone 1 - Value','Phone 2 - Value',
            'E-mail 1 - Value','E-mail 2 - Value','Group Membership']
            thisteam=thisteam[mycols]
            thisteam.columns=newcols
            thisteam=thisteam.replace(np.nan,'')
            fname=cnf._OUTPUT_DIR+'\\google'+team+'.csv'
            thisteam.to_csv(fname, index=False)
    return

def createsignups(df, Mastersignups, season, year):     
    ''' pass signups and add signups to master list, also returns list of current player keys by sport
    typically use writesignupstoExcel instead
    '''   
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season) 
    # Use comma sep on multiple sport entries?? 
    now=datetime.now()
    thisdate=date.strftime(now,'%m/%d/%Y') # for signup date
    df.loc[:,'SUdate']=thisdate # can do this globally although might also add to signups
    startlen=len(Mastersignups) # starting number of signups 
    intcols=['SUkey','Year']
    for i, col in enumerate(intcols):
        if col not in df:
            df.loc[df.index, col]=np.nan
    mycols=Mastersignups.columns.tolist() # desired column order
    for i, col in enumerate(mycols):
        if col not in df:
            df.loc[df.index,col]=np.nan
    # TODO one option here would be to clone comma-separated sport entries (i.e. track and softball)
    for i, sport in enumerate(sportlist):
        # Use caution here due to Tball in Softball string problem (currently set to T-ball)
        thissport=df.loc[df['Sport'].str.contains(sport, na=False, case=False)] # also handles multi-sports
        # Prepare necessary columns
        for index, row in thissport.iterrows():
            thissport.loc[index,'Sport'] = sport # set individually to formal sport name
            thissport.loc[index,'Year'] = int(year)
            thissport.loc[index,'SUkey'] = 0 # assigned actual key below 
        # Now organize signups and add year
        Mastersignups=pd.concat([thissport,Mastersignups], ignore_index=True)
    Mastersignups=Mastersignups[mycols] # put back in original order
    # drop duplicates and save master signups file (keep older signup if present... already assigned SUkey) 
    Mastersignups=Mastersignups.sort_values(['Plakey', 'Sport','Year','SUkey'], ascending=False) # keeps oldest signup
    Mastersignups=Mastersignups.drop_duplicates(subset=['Plakey', 'Sport','Year'])  # drop duplicates (for rerun with updated signups)
    newsignups=len(Mastersignups)-startlen # number of new signups added this pass 
    print('Added ', str(newsignups),' new ', season, ' signups to master list.') 

    # add unique SUkey (if not already assigned)
    neededkeys = Mastersignups[(Mastersignups['SUkey']==0)] # filter by year 
    availSUkeys=findavailablekeys(Mastersignups, 'SUkey', len(neededkeys)) # get necessary # of unique SU keys      
    keycounter=0
    for index, row in neededkeys.iterrows():
        Mastersignups.loc[index,'SUkey'] = availSUkeys[keycounter] # reassign SU key in source master list
        keycounter+=1 # move to next available key
    Mastersignups.loc[:,'Grade']=Mastersignups.Grade.replace('K',0)
    Mastersignups=Mastersignups.sort_values(['Year', 'Sport', 'Gender','Grade'], ascending=False) 
    Mastersignups.loc[:,'Grade']=Mastersignups.Grade.replace(0,'K')
    # autocsvbackup(Mastersignups,'master_signups', newback=True)    
    Mastersignups.to_csv(cnf._INPUT_DIR + '\\master_signups.csv', index=False, date_format='mm/dd/yy') # automatically saved
    return Mastersignups
    
def replaceacro(df, acronyms):
    ''' Pass df column and return with acronyms replaced with full translations (parishes and schools 
    currently used only for CYC rosters '''
    for index, row in acronyms.iterrows():
        acro=acronyms.loc[index]['acronym']
        transl=acronyms.loc[index]['translation']
        # TODO only for parish columns 
        df.loc[:,'Parish of Registration']=df['Parish of Registration'].replace(acro, transl)
        df.loc[:,'Parish of Residence']=df['Parish of Residence'].replace(acro, transl)
        df.loc[:,'School']=df['School'].replace(acro, transl)
    return df

def createrosters(df, season, year, players, teams, coaches, famcontact, acronyms):
    ''' From Mastersignups of this season creates Cabrini CYC roster and transfers (for separate sports)
    and all junior sports (calculates ages for Judge Dowd);  pulls info merged from famcontact, players, teams, and coaches
    teams should already be assigned using teams xls and assigntoteams function
    
    returns:  None ... direct save to OUTPUT_DIR
        '''  
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    specials=['Chess','Track']
    sports=sportsdict.get(season)
    sportlist=[sport for sport in sports if sport not in specials]
    speciallist=[sport for sport in sports if sport in specials] # for track, chess, other oddballs
    
    Specials=df[(df['Year']==year) & (df['Sport'].isin(speciallist))] # deal with these at bottom
    # Proceed with all normal South Central sports 
    df = df[(df['Year']==year) & (df['Sport'].isin(sportlist))] # filter by year 

    # make duplicate entry row for double-rostered players (multiple team assignments)
    thismask = df['Team'].str.contains(',', na=False) # multiple teams are comma separated
    doubles=df.loc[thismask]
    for index, rows in doubles.iterrows():
        team=doubles.loc[index,'Team']
        team=team.split(',')[1] # grab 2nd of duplicate teams
        doubles.loc[index, 'Team'] = team
    df=pd.concat([df,doubles], ignore_index=True) # adds duplicate entry for double-rostered players with 2nd team
    thismask = df['Team'].str.contains(',', na=False) # multiple teams are comma separated
    for index, val in thismask.iteritems():
        if val:
            team=df.loc[index]['Team']
            team=team.split(',')[0] # grab 1st of duplicate teams
            df.loc[index, 'Team'] = team # removes 2nd team from first entry
    
    # now grab all extra info needed for CYC rosters 
    # Street, City, State, Zip, Phone, email, Parishreg, parishres from fam-contact        
    df=pd.merge(df, famcontact, how='left', on=['Famkey'], suffixes=('','_r'))        
    # Get division from Teams xls
    df=pd.merge(df, teams, how='left', on=['Team'], suffixes=('','_r2')) # effectively adds other team info for roster toall players
    # DOB, School  from players.csv
    df=pd.merge(df, players, how='left', on=['Plakey'], suffixes=('','_r3'))        
    df.loc[:,'Role']='Player' # add column for role
    # df['Open/Closed']=np.nan
    df.loc[:,'Coach ID']=''
    
    def formatDOB(val):
        # Pat moore date format is 4/4/19.. reformat as string for csv output
        try:
            return datetime.strftime(val, "%m/%d/%y")
        except:
            # print('Problem converting %s of type %s to date string format' %(val, type(val)) )
            return ''
            
    # Find Cabrini CYC names (containing hyphen)        
    thismask = df['Team'].str.contains('-', case=False, na=False)
    CabriniCYC=df.loc[thismask] # all  players on Cabrini CYC teams all sports this season
        
    # Finds info for CYC coaches (all sports) and generate roster entries
    coachroster=addcoachestoroster(teams, coaches) # coaches roster already in correct format + sport column
    if len(CabriniCYC)>1: # skip if all transfers or junior (i.e. in spring)
    # Split by sport 
        for i, sport in enumerate(sportlist):
            Sportroster=CabriniCYC[CabriniCYC['Sport']==sport]
            # reformat this mess as single CYC roster
            Sportroster=organizeroster(Sportroster) 
            # Add coaches from this sport to roster
            Rostercoaches=coachroster[coachroster['Sport']==sport]
            Rostercoaches=organizeroster(Rostercoaches)
            Sportroster=pd.concat([Sportroster,Rostercoaches], ignore_index=True) # adds coaches and players together      
            Sportroster=Sportroster.sort_values(['Team','Role','Grade','Lname'])
            fname=cnf._OUTPUT_DIR+'\\Cabrini_'+sport+'roster'+str(year)+'.csv'
            Sportroster=replaceacro(Sportroster, acronyms) # replace abbreviations
            Sportroster.loc[:,'Birthdate']=Sportroster['Birthdate'].apply(lambda x: formatDOB(x))
            Sportroster.to_csv(fname, index=False)
        # done with Cabrini CYC rosters

    # Break out all other types of teams (transfers, junior teams, Chess/Track)
    thismask = df['Team'].str.contains('-', case=False, na=False)
    Others=df.loc[~thismask] # no hyphen for all non Cabrini CYC level (Cabrini junior and transfers)
    
    # Cabrini transferred players to CYC teams with # (i.e. Ambrose#8B, OLS#3G)
    # Non-CYC cabrini junior teams start with number
    thismask = Others['Team'].str.contains('#', na=True) # flag  nans and set to true (usually jr teams w/o assignment)
    # Transferred teams contain # such as OLS#3G
    Transfers=Others.loc[thismask] # transferred teams have # but no hyphen
    for i, sport in enumerate(sportlist): # output roster for all transfers (all grades in case of CYC)
        Transferroster=Transfers[Transfers['Sport']==sport]
        Transferroster=organizeroster(Transferroster)
        Transferroster=Transferroster.sort_values(['Team', 'Sex', 'Grade'], ascending=True)
        fname=cnf._OUTPUT_DIR+'\\CYC'+sport+'transfers.csv'
        Transferroster=replaceacro(Transferroster,acronyms)
        Transferroster.loc[:,'Birthdate']=Transferroster['Birthdate'].apply(lambda x: formatDOB(x))
        Transferroster.to_csv(fname, index=False)        
    # Now deal with junior cabrini (should be only thing left after Cabrini CYC< 
    # transfers, special sports
    Juniorteams=Others.loc[~thismask]  # remove transfers
    Juniorteams=Juniorteams[Juniorteams['Team']!='drop'] # remove dropped players
        
    # now output all junior teams in same format (sometimes needed by Judge Dowd)
    # also calculate current age
    if len(Juniorteams)>0:
        Juniorteams=organizeroster(Juniorteams) # put in standard South Central roster format
        # Calculate current age from DOBs (renamed to Birthdate for roster only)
        Juniorteams.loc[:,'Age']=calcage(Juniorteams['Birthdate'])
        fname=cnf._OUTPUT_DIR+'\\Cabrini_junior_teams_'+str(year)+'.csv'
        Juniorteams=replaceacro(Juniorteams, acronyms)
        Juniorteams.loc[:,'Birthdate']=Juniorteams['Birthdate'].apply(lambda x: formatDOB(x))
        Juniorteams.to_csv(fname, index=False)

    # Deal with special cases -Track and Chess
    # Get DOB/school from players.. anything else needed by Butch Rosier?
    Specials=pd.merge(Specials, players, how='left', on='Plakey', suffixes=('','_r'))
    # needs address
    Specials=pd.merge(Specials, famcontact, how='left', on='Famkey', suffixes=('','_r2'))
    for i, sport in enumerate(speciallist): # output roster for all transfers (all grades in case of CYC)
        Specials=Specials[Specials['Sport']==sport]
        Specials=Specials.rename(columns={'DOB':'Birthdate'})
        mycols=['First', 'Last','Gender','Team','Grade','Birthdate','School','Address','Zip']
        Specials=Specials[mycols]
        Specials=Specials.sort_values(['Gender', 'Birthdate', 'Grade'], ascending=True)
        Specials.loc[:,'Birthdate']=Specials['Birthdate'].apply(lambda x: formatDOB(x))
        fname= cnf._OUTPUT_DIR+'\\'+ sport+'_'+str(year)+'_rosters.csv'
        Specials.to_csv(fname, index=False)   
    return    
   

def makemultiteam(df):
    '''Small utility called by assigntoteams to make temp teams df that has separate entry for each grade if team is mixed grade
    then merge to assign teams is straightforward
    twoteams- '''
    # TODO annoying problem with combining teams due to K1 (string but not int)

    mycols=df.dtypes.index
    # Deal with K1, K2 and such teams
    kteams=[str(s) for s in np.ndarray.tolist(df.Graderange.unique()) if 'K' in str(s)]
    kteams=[s for s in kteams if len(s)>1] # combo teams only
    kteams=df[df['Graderange'].isin(kteams)]
    xtrateams=pd.DataFrame(index=np.arange(0,0),columns=mycols) # empty df
    # clones rows to match lower grades in range
    for index, row in kteams.iterrows():  
        tempstr= kteams.loc[index]['Graderange']
        gr1=0 # 0 for grade K 
        gr2=int(tempstr[1])
        for gr in range(gr1,gr2):
            newrow=kteams.loc[index] # grabs row as series
            newrow.loc['Grade'] = gr # set to correct grade
            xtrateams=xtrateams.append(newrow) # add single row to temp df
    df.loc[:,'Grade']=df.Grade.replace('K','0', regex=True)

    # get rid of K string problem
    df.loc[:,'Graderange']=df.Graderange.replace('K','0', regex=True)
    df.loc[:,'Graderange']=df.Graderange.astype('int')
    # now handle numbered multiteams (e.g. 45 78 two digit ints)    
    multiteams=df.loc[df['Graderange']>9] # subset of teams comprised of multiple grades
    for index, row in multiteams.iterrows(): # check for 3 or more grades
        # TODO make sure it's not 3 grades (i.e. K-2)
        tempstr= str(multiteams.loc[index]['Graderange'])
        gr1=int(tempstr[0])
        gr2=int(tempstr[1])
        for gr in range(gr1,gr2):
            newrow=multiteams.loc[index] # grabs row as series
            newrow.loc['Grade'] = gr # set to correct grade
            xtrateams=xtrateams.append(newrow) # add single row to temp df
    # Detect gender-grade-sport w/ two teams
    # now combine with original df
    df=pd.concat([df,xtrateams], ignore_index=True) # complete team set
    df=df[mycols] # back in original order
    df=df.sort_values(['Gender','Grade'], ascending=True)
    # After cloning by grade, look for two teams per grade options 
    twoteams=df[df.duplicated(['Sport','Gender','Grade'])]
    return df, twoteams

def detectrosterchange(PMroster, myroster):
    '''Compare submitted and returned rosters to look for unique rows (altered by Pat Moore)
    first row is Pat Moore version (presumably correct to match CYC database) and second row is my 
    submitted version... make any corrections to appropriate source data files 
    datetime format conversions can be problematic '''
    # all columns by default, false drops both duplicates leaving unique rows
    bothrosters=pd.concat([PMroster,myroster])
    mycols=bothrosters.columns
    nanrows=bothrosters[pd.isnull(bothrosters['Birthdate'])]
    nanrows=nanrows.drop_duplicates(keep=False)
    # ensure player rows are both in correct format
    myroster=myroster[pd.notnull(myroster['Birthdate'])]
    PMroster=PMroster[pd.notnull(PMroster['Birthdate'])]
    
    def removeLeadZero(val):
        if val.startswith('0'):
            return val[1:]
        else:
            return val
    myroster.loc[:,'Birthdate']=myroster['Birthdate'].apply(lambda x:pd.to_datetime(x).strftime('%m/%d/%Y'))
    PMroster.loc[:,'Birthdate']=PMroster['Birthdate'].apply(lambda x:pd.to_datetime(x).strftime('%m/%d/%Y'))

    myroster.loc[:,'Birthdate']=myroster['Birthdate'].apply(lambda x:removeLeadZero(x))
    PMroster.loc[:,'Birthdate']=PMroster['Birthdate'].apply(lambda x:removeLeadZero(x))

    bothrosters=pd.concat([PMroster,myroster])
    bothrosters=bothrosters.sort_values(['Fname','Lname'])
    # Fix date string differences
    alteredrows=bothrosters.drop_duplicates(keep=False)
    alteredrows=alteredrows.append(nanrows)
    alteredrows=alteredrows[mycols]
    alteredrows=alteredrows.sort_values(['Lname','Fname'])
    return alteredrows

def saveteams(teams):
    '''Save teams tab into teams_coaches.xlsx after changes have been made '''
    book=load_workbook('Teams_coaches.xlsx')
    writer=pd.ExcelWriter('Teams_coaches.xlsx', engine='openpyxl')
    writer.book=book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    teams.to_excel(writer,sheet_name='Teams',index=False) # this overwrites existing file
    writer.save() # saves xls file with all modified data
    
def assigntoteams(df, season, year, teams, overwrite=False):
    '''From mastersignups finds CYC team name based on year, grade, gender and sport from teams tab 
    (which only contains names from this season/year to avoid screwing up old custom team assignments''' 
    # teamsmult has multi grade range teams with duplicates for merge matching
    # twoteams is multiple teams for same grade
    Teamsmult, Twoteams =makemultiteam(teams) # makes duplicates team entries to match both grades
    # compare grades as ints with K=0 
    df.loc[:,'Grade']=df.Grade.replace('K','0', regex=True) # convert Ks to zeros
    df.loc[:,'Grade']=df['Grade'].astype('int')
    Teamsmult.loc[:,'Grade']=Teamsmult['Grade'].astype('int') # ensure these are ints
    # left merge keeps all master_signups oentries
    df=pd.merge(df, Teamsmult, how='left', on=['Year','Grade','Gender','Sport'], suffixes=('','_r'))
    # need to drop SUkey duplicates (keeping first)... occurs if >1 team per grade
    df=df.drop_duplicates(subset=['SUkey']) # drops any duplicates by unique SUkey
    # Consider all sports except Track (team assignment done separately by DOB)
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    # this is post-merge so no chance of getting indices screwed up 
    # select current sports & year and subset with new team assignment
    CurrentSU=df.loc[(df['Sport'].isin(sportlist)) & (df['Year']==year) & (pd.notnull(df['Team_r']))]
    if overwrite==False: # if no overwrite, keep only those with nan for team
        CurrentSU=CurrentSU.loc[pd.isnull(CurrentSU['Team'])]
    # Never overwrite team assignment for known drops
    CurrentSU=CurrentSU[CurrentSU['Team']!='drop']
    counter=0
    for index, row in CurrentSU.iterrows(): 
        # all remaining can be overwritted (those w/ existing team dropped above)
        match=df[df['SUkey']==CurrentSU.loc[index]['SUkey']]
        if len(match)==1:
            thisind=match.index[0]
            # add new team assignment to correct index in original master signups
            df.loc[thisind, 'Team'] = CurrentSU.loc[index]['Team_r']
            counter+=1
    print(str(counter),' player(s) newly assigned to teams')
    # now drop extra columns and sort 
    mycols=['SUkey','First', 'Last', 'Grade', 'Gender', 'Sport', 'Year', 'Team', 'Plakey','Famkey', 'Family', 
    'SUdate', 'Issue date', 'Uniform#','UniReturnDate'] 
    df.loc[:,'Grade']=df.Grade.replace('K',0)
    df=df.sort_values(['Year','Sport', 'Gender', 'Grade'], ascending=True)
    df.loc[:,'Grade']=df.Grade.replace('0','K', regex=True) # make sure any 0 grades are again replaced with K
    df=df[mycols]
    autocsvbackup(df,'master_signups', newback=True) # autobackup of master signups
    df.to_csv(cnf._INPUT_DIR + '\\master_signups.csv', index=False) # save/overwrite existing csv
    return df

def assigntrackgroup(df, year, players):
    '''Assign to different track team based on age on May 31 of this year (school year+1)
    '''
    Track=df[(df['Sport']=='Track') & (df['Year']==year)]
    Track=pd.merge(Track,players, how='left', on=['Plakey'], suffixes=('','2'))
    numunassigned=len(Track[pd.isnull(Track['Team'])])
    for index, row in Track.iterrows():
        DOB=Track.loc[index]['DOB'] # merged from players.csv
        if isinstance(DOB,str):
            DOB=datetime.strptime(DOB,"%m/%d/%Y").date() # convert string to datetime
        elif isinstance(DOB, pd.tslib.Timestamp):
            DOB=DOB.date() # convert timestamp to datetime
        trackage=date(year+1,5,31)-DOB # age on prior year's May 31st (same as school year in current convention)
        trackage=(trackage.days + trackage.seconds/86400)/365.2425 # as decimal
        trackage=math.floor(trackage)
        if trackage <=7:
            team='Track7'
        elif 8 <= trackage <=9: 
            team='Track89'
        elif 10 <= trackage <=11: 
            team='Track1011'
        elif 12 <= trackage <=13: 
            team='Track1213'
        elif 14 <= trackage <=15:
            team='Track1415'
        else: # probably some entry error
            mystr=Track.loc[index]['First']+' '+Track.loc[index]['Last']+' Grade:'+Track.loc[index]['Grade']
            print('Suspected DOB error for',mystr, 'DOB:', datetime.strftime(DOB, "%m/%d/%y") )
            team=''
        # Now write back altered subset to mastersignups (index is lost so use SUkey)
        SUkey=int(Track.loc[index]['SUkey'])
        match=df[df['SUkey']==SUkey] # This gives correct index
        df.loc[match.index[0], 'Team'] = team # alter/assign team for this signup
    newlyassigned=numunassigned-len(Track[pd.isnull(Track['Team'])])
    print(newlyassigned,' players assigned to track age group.')
    return df


def readbackevents(trackevents):
    '''
    Reads back choices of track events from summary sheet and prep for 
    copy to Pat Moore spreadsheet
    in 4x100, 4x200, 4x400 col enter start order 1,2,3,4,1A,2A
    '''
    regcols=['Last', 'First', 'Middle', 'Gender',
       'DOB', 'Team Code','Event#1', 'Event#2', 'Event#3', 'Event#4']
    # Manually enter order of runners and alternates for relays
    events=['50M', '100M', '200M', '400M', '800M', '1600M', 'SoftThrow', 
            'ShotPut','StdLongJump', 'RunLongJump']
    regfile=pd.DataFrame(columns=regcols)
    regfile.loc[:,'Team Code']=='SFC'
    for index, row in trackevents.iterrows():
        # get events for which player is signed up
        playerevents=[]
        for i, event in enumerate(events):
            if str(row[event])!='nan':
                playerevents.append(event)
                print(event,' for ',row.First, row.Last)
        # Check for relay type separately
        if row['Relay'] in ['4x100', '4x200','4x400']:
            playerevents.append(row['Relay'])
            print(row['Relay'],' for ',row.First, row.Last)
        if len(playerevents)>4:
            print('Too many events for ', row.First, row.Last)
        # Now construct player's entry in regfile
        thisentry=row
        thisentry['Middle']=''
        thisentry['Team Code']='SFC'
        # Gender is upper case M or F
        thisentry['Gender']=thisentry['Gender'].upper()
        for i, event in enumerate(playerevents):
            colname='Event#'+str(i+1)
            thisentry[colname]=event
        regfile=regfile.append(thisentry, ignore_index=True)
    regfile=regfile[regcols]
    return regfile

def maketracksummary(df, year, players):
    '''Assign to different track team based on age on May 31 of this year (school year+1)
    '''
    Track=df[(df['Sport']=='Track') & (df['Year']==year)]
    Track=pd.merge(Track,players, how='left', on=['Plakey'], suffixes=('','2'))
    Track.loc[Track.index,'Trackage']=np.nan
    for index, row in Track.iterrows():
        DOB=Track.loc[index]['DOB'] # merged from players.csv
        if isinstance(DOB,str):
            DOB=datetime.strptime(DOB,"%m/%d/%Y").date() # convert string to datetime
        elif isinstance(DOB, pd.tslib.Timestamp):
            DOB=DOB.date() # convert timestamp to datetime
        trackage=date(year+1,5,31)-DOB # age on prior year's May 31st (same as school year in current convention)
        trackage=(trackage.days + trackage.seconds/86400)/365.2425 # as decimal
        Track.loc[index,'Trackage'] = trackage
        trackage=math.floor(trackage)
        if trackage <=7:
            team='Track7'
        elif 8 <= trackage <=9: 
            team='Track89'
        elif 10 <= trackage <=11: 
            team='Track1011'
        elif 12 <= trackage <=13: 
            team='Track1213'
        elif 14 <= trackage <=15:
            team='Track1415'
        else: # probably some entry error
            mystr=Track.loc[index]['First']+' '+Track.loc[index]['Last']+' Grade:'+Track.loc[index]['Grade']
            print('Suspected DOB error for',mystr, 'DOB:', datetime.date.strftime(DOB, "%m/%d/%y") )
            team=''
        Track.loc[index,'Team'] = team
    Track=Track.sort_values(['Trackage'])
    mycols=['First', 'Last', 'Grade', 'Gender','DOB','Team','Trackage']
    Track=Track[mycols]
    return Track

def findrecruits(df, players, famcontact, season, year, signupfile):
    '''Read list of signed-up player keys from xls file; compare with last year's set of 
    players from master Signups log
    7/2018 mod... grab DOB to allow easier manual additions to signups '''
    mycols=df.columns.tolist() # Same columns as mastersignups 
    sportsdict={'Fall':['VB','Soccer'], 'Winter':['Basketball'],'Spring':['Track','Softball','Baseball','T-ball']}
    sportlist=sportsdict.get(season)
    Recruits=pd.DataFrame(columns=mycols) # empty frame for recruits
    for i, sport in enumerate(sportlist):
        thissport=df[df['Sport']==sport]
        thissport=thissport.sort_values(['Year'], ascending=False) # most current signups at top
        plakeylist=thissport.Plakey.unique() # ndarray with list of unique soccer players      
        keylist=plakeylist.tolist()
        for i, key in enumerate(keylist):
            match=thissport[thissport['Plakey']==key]
            # recruits ... played in year -1 but not in year
            if year-1 in match.Year.unique() and year not in match.Year.unique():
                match=match[0:1] # take only last season's signup
                Recruits=pd.concat([Recruits,match], ignore_index=True)
    # plakey, famkey, first, last, grade, gender, 
    Recruits.loc[:,'Grade']=Recruits.Grade.replace('K',0) # replace K with zero to allow sorting
    Recruits.loc[:,'Grade']=Recruits.Grade.astype(int)
    Recruits.loc[:,'Grade']=Recruits.Grade+1 # adjust to correct grade for this year 
    # Drop if graduated
    Recruits=Recruits[Recruits['Grade']<=8]
    # adjust grade such that players current grade is in list
    # join with famcontact on famkey to get contact info (emails, phones, etc.)
    # Inner join on famkey adds the necessary info
    Recruits=pd.merge(Recruits, famcontact,how='inner', on='Famkey', suffixes=('','_r'))
    # Now need to look up school from players.csv    
    Recruits=pd.merge(Recruits, players, how='inner', on='Plakey', suffixes=('','_r'))    
    
    mycols=['First', 'Last', 'DOB', 'Gender', 'School', 'Grade',  'Address', 'Zip', 
            'Parish_registration', 'Sport', 'Phone1', 'Text1','Email1', 'Phone2', 
            'Text2', 'Email2', 'Plakey', 'Famkey', 'Family'] 
    Recruits=Recruits[mycols]
    Recruits.loc[:,'Grade']=Recruits.Grade.replace('K',0)
    Recruits=Recruits.sort_values(['Grade'], ascending=True)
    Recruits.loc[:,'Grade']=Recruits.Grade.replace(0,'K') # replace K with zero to allow sorting
    
    Recruits=Recruits.sort_values(['Sport', 'Gender', 'Grade'], ascending=True)
    # now write recruits to tab in master signups file
    if signupfile.endswith('.csv'):
        fname=cnf._OUTPUT_DIR + '\\%s%s_recruits.csv' %(season, year)
        Recruits.to_csv(fname, index=False)
        print("Info on possible recruits saved to", fname)
    else: # should be excel file
        book=load_workbook(signupfile)
        writer=pd.ExcelWriter(signupfile, engine='openpyxl')
        writer.book=book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        Recruits.to_excel(writer,sheet_name='Recruits',index=False) # this overwrites existing file
        writer.save() # saves xls file with all modified data   
        print("Info on possible recruits saved in", signupfile)
    return

def summarizesignups(df, season, year, **kwargs):
    '''Write out summary to date of players by sport, gender, grade with 
    abbreviated playerlist
    can work on either Mastersignups or single season signup
    
    kwargs: 'XLSpath': path to xls signup file (save as separate sheet here)
        'toDf': don't save ... return as dataframe
    '''
    mycols=['Sport','Gender','Grade','Number','Playerlist','Plakeys']
    sportsum=pd.DataFrame(columns=mycols)
    # Determine if this is mastersignups (or single season raw signup file)
    if 'Basketball' in df.Sport.unique() and 'Track' in df.Sport.unique():
        df=df[(df['Year']==year)] # this year only
        df=df.reset_index(drop=True)
        if season=='Fall':
            sports=['VB', 'Soccer']        
            CurrentSU=df[df.Sport.isin(sports)]
        if season=='Winter':
            sports=['Basketball']        
            CurrentSU=df[df.Sport.isin(sports)] # winter (bball) signups
        if season=='Spring':
            sports=['Track','Softball','Baseball', 'T-ball']
            CurrentSU=df[df.Sport.isin(sports)]
    else:
        CurrentSU=df # single season signup
        sports=CurrentSU.Sport.unique().tolist()
    # Replace K with 0
    CurrentSU.loc[:,'Grade']=CurrentSU.Grade.replace('K',0)
    CurrentSU.loc[:,'Grade']=CurrentSU.Grade.astype(int) # convert all to int for sorting
    if 'Team' in CurrentSU: # remove drops if using mastersignups
        CurrentSU=CurrentSU[CurrentSU['Team']!='drop']
    # loop through sport, gender, grade
    for i, sport in enumerate(sports):
        thismask = CurrentSU['Sport'].str.contains(sport, na=False, case=False)
        thissport=CurrentSU.loc[thismask]
        genders=thissport.Gender.unique()
        for j, gend in enumerate(genders):
            thisgensport=thissport[thissport['Gender']==gend]
            grades=thisgensport.Grade.unique()
            for k, grade in enumerate(grades):
                # create single rowed df for this sport-gender-grade
                theseplayers=thisgensport[thisgensport['Grade']==grade] # for getting player list
                playerlist=[] # first l. for all in gender grade 
                plakeylist=[]
                for ind, ro in theseplayers.iterrows():
                    first=theseplayers.loc[ind]['First']
                    last=theseplayers.loc[ind]['Last']
                    strname=str(first)+' ' +str(last)
                    playerlist.append(strname)
                    plakeylist.append(theseplayers.loc[ind]['Plakey'])
                thisgendergrade=pd.DataFrame(index=np.arange(0,1),columns=mycols)
                thisgendergrade.loc[0,'Sport'] = sport
                thisgendergrade.loc[0,'Gender'] = gend
                thisgendergrade.loc[0,'Grade'] = grade
                thisgendergrade.loc[0,'Number'] = len(playerlist)
                thisgendergrade.loc[0,'Playerlist'] = playerlist
                thisgendergrade.loc[0,'Plakeys'] = plakeylist
                sportsum=sportsum.append(thisgendergrade) # adds row to 
    sportsum=sportsum.sort_values(['Sport','Gender','Grade'])
    sportsum.Grade=sportsum.Grade.replace(0,'K')
    if kwargs.get('toDf'): # don't save ... return as dataframe
        return sportsum
    # now write recruits to tab in master signups file
    if 'XLSpath' in kwargs: # optional excel write with provided path
        signupfile=kwargs.get('XLSpath')
        book=load_workbook(signupfile)
        # xls signups file should be in SC_files (input dir)
        writer=pd.ExcelWriter(cnf._INPUT_DIR+'\\'+signupfile, engine='openpyxl')
        writer.book=book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        sportsum.to_excel(writer,sheet_name='Summary',index=False) # this overwrites existing file
        writer.save() # saves xls file with all modified data
    else: # default is write to csv file in std output dir
        fname="%s\\%s_%i_signup_summary.csv" %(cnf._OUTPUT_DIR, season, year)
        sportsum.to_csv(fname,index=False) 
    return

def findmissinginfo(df, players, famcontact):
    ''' Using player and family keys, update nan values in SC signups (mainly 
    for paper/word-of-mouth entries needed for writesignupstoExcel '''
    for index, row in df.iterrows():
        # manual entries won't have google drive timestamp
        if type(row.Timestamp)!=pd._libs.tslib.Timestamp:
            thisplakey=row.Plakey
            thisfamkey=row.Famkey
            # get first, dob, school from master players list
            match = players[(players['Plakey']==thisplakey)]
            if len(match)==1: # update school, gender
                df.loc[index,'School'] = match.iloc[0]['School']
                df.loc[index,'Gender'] = match.iloc[0]['Gender']
            # get address, zip, parish, phone/text, email, phone2, text2, email2 from famcontact 
            match = famcontact[(famcontact['Famkey']==thisfamkey)]
            if len(match)==1:
                df.loc[index,'Address'] = match.iloc[0]['Address']
                try:
                    df.loc[index,'Zip'] = int(match.iloc[0]['Zip'])
                except:
                    print('Problem w/ zip code for', row.Last)
                df.loc[index,'Parish'] = match.iloc[0]['Parish_registration']
                df.loc[index,'Phone'] = match.iloc[0]['Phone1']
                df.loc[index,'Text1'] = match.iloc[0]['Text1']
                df.loc[index,'Email1'] = match.iloc[0]['Email1']
                df.loc[index,'Phone2'] = match.iloc[0]['Phone2']
                df.loc[index,'Text2'] = match.iloc[0]['Text2']
                df.loc[index,'Email2'] = match.iloc[0]['Email2']
    return df

def makephonedict(famcontact):
    ''' Construct dictionary w/ all active numbers and associated famkeys for matching operations
    called by findplayers'''
    phonedict={}
    mylist=['Phone1','Phone2','Phone3','Phone4']
    for i, col in enumerate(mylist):        
        phones=famcontact.loc[pd.notnull(famcontact[col])]
        for index, row in phones.iterrows():
            phnum=phones.loc[index][col]
            famkey=phones.loc[index]['Famkey']
            phonedict.update({phnum:famkey})
    return phonedict

def makelastlist(df, df2):
    ''' Construct list of tuples (not dict due to possible duplicate keys) w/ all active last names and 
    associated famkeys for matching operations '''
    lnamelist=[] # list of tuples
    # handle last names in players.csv
    df=df.drop_duplicates(subset=['Famkey']) # one entry per family
    for index, row in df.iterrows():
        lnamelist.append((df.loc[index]['Last'],df.loc[index]['Famkey']))
    mylist=['Plast1','Plast2','Plast3']
    for i, col in enumerate(mylist):        
        temp=df2.loc[pd.notnull(df2[col])]
        for index, row in temp.iterrows():
            last=temp.loc[index][col]
            famkey=temp.loc[index]['Famkey']
            if (last, famkey) not in lnamelist:
                lnamelist.append((last,famkey))
    return lnamelist

''' Testing 
row=unmatched.iloc[0]
index=37
'''
def findplayers(signups, players, famcontact, year):
    '''Find player key from players df using multiple levels of matching (Plakey already initialized)
    if not a perfect match on all characters, create some data output structure to resolve possible problems
    plakey and famkey cols added in loadprocess '''
    savepla=False # flags to save modified files
    savefam=False

    phonedict=makephonedict(famcontact) # dict for known phone #s to famkey
    # left merge keeping index to do first/last/dob match (works since no duplicates in players)
    matches=signups.reset_index().merge(players, how='left', on=['First','Last','DOB'], suffixes=('','_2')).set_index('index')
    if len(matches)!=len(signups):
        print('Multiple match problem in players csv!')
        return signups, players, famcontact
    # after first/last/DOB match, copies over assigned plakey as found in players
    matches=matches[(pd.isnull(matches['Plakey'])) & (pd.notnull(matches['Plakey_2']))]
    signups.loc[matches.index,'Plakey'] = matches['Plakey_2']
    signups.loc[matches.index,'Famkey'] = matches['Famkey_2']
    # same matching process for (Alias, Last, DOB)
    alias=signups.copy().rename(columns={'First':'Alias'})
    matches=alias.reset_index().merge(players, how='left', on=['Alias','Last','DOB'], suffixes=('','_2')).set_index('index')
    if len(matches)!=len(alias):
        print('Multiple match problem in players csv!')
        return signups, players, famcontact
    # Find newly matched plakeys (using player alias)
    matches=matches[(pd.isnull(matches['Plakey'])) & (pd.notnull(matches['Plakey_2']))]
    signups.loc[matches.index,'Plakey'] = matches['Plakey_2']
    signups.loc[matches.index,'Famkey'] = matches['Famkey_2']
    # Continue w/ attempted id via phone or add new 
    unmatched=signups.loc[pd.isnull(signups['Plakey'])]
    nophone = unmatched[pd.isnull(unmatched['Phone1'])]
    if len(nophone)>1:
        print('Add missing phone for %s' %",".join(nophone.Last.unique().tolist()))
    unmatched=unmatched.loc[pd.notnull(unmatched['Phone1'])]
    # blank DOBs will be nonetype (after datetime conversion)
    nobd=unmatched[pd.isnull(unmatched['DOB'])]
    if len(nobd)>0:
        print('Enter DOB for %s' %",".join(nobd.Last.unique().tolist()))
    unmatched= unmatched[pd.notnull(unmatched.DOB)] # Remove no DOB

    # Birthday instead of DOB problem
    bd=unmatched[ unmatched['DOB'] > date(year-1,1,1)]
    if len(bd)>0:
        print('Fix birthday instead of DOB for %s' %",".join(bd.Last.unique().tolist()))
    unmatched= unmatched[~unmatched.index.isin(bd.index)] # Remove birthdays
    
    # lastnames=makelastlist(players, famcontact) # set of tuples with last name and assoc famkey
    for index, row in unmatched.iterrows(): # row=unmatched.iloc[7]
        # no exact match and no phone1... skip and correct manual entry
        if str(unmatched.loc[index]['Phone1'])=='nan': # skip if no phone1 entry
            print('No exact match for', row.First, row.Last, row.DOB, 'add phone #')
            continue
        # no exact match ... find matching phone/ family match
        phonelist=[]
        for i, col in enumerate(['Phone1','Phone2']):
            phonelist.append(unmatched.loc[index][col])
        phonelist=[s.strip() for s in phonelist if str(s)!='nan']
        famkey=[phonedict.get(s,'') for s in phonelist if s in phonedict]
        if len(famkey)>0: # found one or mre matching phone numbers
            famkey=int(famkey[0])
            signups.loc[index, 'Famkey'] = famkey
            kids=players[players['Famkey']==famkey]
            # now find player (alias/ inexact) or add new kid
            choicedict=newplayertk(unmatched.loc[index], phonelist, kids)
            if choicedict.get('choice','')=='alias' or choicedict.get('choice','')=='ID':
                signups.loc[index, 'Plakey'] = choicedict.get('ID',0)
                if choicedict.get('choice','')=='alias':
                    players=addalias(players, signups.loc[index]) # add alias to this entry and directly save
                    savepla=True
            elif choicedict.get('choice','')=='addkid':
                # add kid to existing family
                players, plakey=addplayer(signups.loc[index], players)
                signups.loc[index, 'Plakey'] = plakey
                savepla=True
            elif choicedict.get('choice','')=='skip':
                # process add alias function
                pass
        else: # possible new family (could be manually added or new player/family or unique/changed phone)
            Ser=signups.loc[index] # convert current SU info to series
            choice=newplafamtk(Ser, phonelist) # confirm new player/family
            if choice=='adddb': # new family (and new player)
                players, famcontact, plakey, famkey=addnewplafam(Ser, players, famcontact) # pass as df row
                savepla=True
                savefam=True
                # Add new #(s) and famkey to phonedict
                for i, num in enumerate(phonelist):
                    phonedict.update({num:famkey})
                # Update signups 
                signups.loc[index, 'Famkey'] = famkey
                signups.loc[index, 'Plakey'] = plakey
            else:
                print('Skipped unidentified player', row.First, row.Last)
    if savepla: # save players data csv if modified
        players.to_csv(cnf._INPUT_DIR + '\\players.csv',index=False)
    if savefam:
        famcontact.to_csv(cnf._INPUT_DIR + '\\family_contact.csv',index=False)
    return signups, players, famcontact

def addnewplafam(Ser, players, famcontact):
    ''' Add single new player & family after confirmation with tk '''
    # Create family name for new families (added to famcontacts but not needed in signups)
    last=str(Ser.Last).title()
    plast=str(Ser.Plast1).title() # parent 1 last name
    if last==plast: # same name for parent and player 
        Ser.loc['Family'] = last # assign player last name as family name
    elif plast=='nan': # parent last name missing
        Ser.loc['Family']= last # assign player last name as family name
    elif plast in last: # if kid name is hyphenated multiname, use the hyphenated name
        Ser.loc['Family']= last
    else: # make a new family name for different first/last
        newname=last+'_'+plast
        Ser.loc['Family'] = newname
    # From new player series entries, create entries for master players table
    players, plakey =addplayer(Ser,players) # update master players list, save and return
    Ser.loc['Plakey'] = plakey
    # update master families lists, save and return
    famcontact, famkey=addfamily(Ser,famcontact)
    # Still needs to add returned famkey to this player's entry 
    match=players[players['Plakey']==plakey]
    if len(match)==1:
        thisind=match.index[0]
        players.loc[thisind,'Famkey'] = famkey
    return players, famcontact, plakey, famkey # go ahead and pass back modified versions to main

def addfamily(Ser, famcontact):
    ''' df contains new families to add to master family contact and family billing tables '''
    # TODO run autobackup for famcontact?
    # find and assign new family key
    famkey=findavailablekeys(famcontact, 'Famkey', 1)[0] # get new unique famkey
    Ser.loc['Famkey'] = famkey

    # Add all the default missing columns
    Ser.loc['City'] ='St. Louis'
    Ser.loc['State'] ='MO'
    for col in ['Parish_residence','Pfirst3','Plast3','Phone3','Text3','Phone4','Text4','Email3']:
        Ser[col]=''
    Ser=Ser.rename(columns={'Plakey': 'Players', 'Parish': 'Parish_registration', 
        'Phone': 'Phone1', 'Text': 'Text1', 'Email': 'Email1',})
    df=pd.DataFrame()
    df=df.append(Ser, ignore_index=True)
    # update family contact
    df=df.rename(columns={'Plakey': 'Players', 'Parish': 'Parish_registration', 
            'Phone': 'Phone1', 'Text': 'Text1', 'Email': 'Email1'})
    mycols=famcontact.columns.tolist()
    dfcols=df.columns.tolist()
    missing=[i for i in mycols if i not in dfcols]
    for i, col in enumerate(missing): # shouldn't happen but doublecheck
        df[col]='' 
    df=df[mycols] # put back in original order
    # concat the two frames (if same names, column order doesn't matter)
    famcontact=pd.concat([famcontact,df], ignore_index=True) 
    famcontact=famcontact.reset_index(drop=True)
    famcontact=famcontact.sort_values(['Famkey'], ascending=True)
    # autocsvbackup(famcontact, 'family_contact', newback=True)
    # famcontact=famcontact.to_csv('family_contact.csv',index =False)
    return famcontact, famkey

def addplayer(Ser, players):
    ''' gets info from confirmed newplayers,reformat and adds to main players list '''
    # first assign new player and family keys (index already reset by findfamily)
    plakey=findavailablekeys(players, 'Plakey', 1)[0] # gets new unique plakey
    # assign new player keys and create df with info to add to master players 
    Ser.loc['Plakey'] = plakey
    
    mycols=players.columns.tolist() # desired column order
    Ser.loc['Gradeadj'] = 0 # add grade adjust col and init to zero
    Ser.loc['Alias']=''# add alias column
    Ser.loc['Uni#'] = np.nan # add default uniform number (junior teams)
    df=pd.DataFrame()
    df=df.append(Ser, ignore_index=True)
    dfcols=df.columns.tolist()
    missing=[i for i in mycols if i not in dfcols]
    for i, val in enumerate(missing): # shouldn't happen  but check just in case
        df[val]='' # set missing col to string 
    df=df[mycols] # same column set
    players=pd.concat([players, df], ignore_index=True) # concat the two frames (if same names, column order doesn't matter)
    players=players.reset_index(drop=True)
    players=players.sort_values(['Plakey'], ascending=True) # sort by player number
    # players['DOB']=pd.to_datetime(players['DOB'], format='m/%d/%Y') # convert DOB to usual format
    # autocsvbackup(players,'players', newback=True) # backup of players file    
    # players=players.to_csv('players.csv',encoding='cp437', index =False)
    return players, plakey # full master list with new entries
    
#%%
def newplayertk(Ser, phonelist, kids):
    ''' Found family... possible new player or existing player (W/ data entry) or new alias
    Choices : (ID existing player from family, ID and add alias, add new players)
    pass famkey, phonelist (for convenience); no need for distinction between add player and 
    add player & add family... both handled by addplafam 
    args:
        Ser - this unmatched row
        phonelist - phone numbers entered in signup
        kids -dataframe w/ info on potential matching kids (using phone match)
        '''
    # first print out existing info in various lines
    root = tk.Tk()
    choice=tk.StringVar() # must be define outside of event called functions
    thisplanum=tk.StringVar() # tk var for entered player num (if existing).. will be converted to int
    rownum=0
    try:
        family=str(kids.iloc[0]['Family']) # get fam name from passed matches
    except:
        print('No kids found for family of', Ser.First, Ser.Last)
        family=''
    mytext='No match for player '+ Ser.First+' '+Ser.Last+', Family '+ family + ' Parent: '+str(Ser.Pfirst1) + ' '+ str(Ser.Plast1)
    a=tk.Label(root, text=mytext)
    a.grid(row=rownum)
    rownum+=1
    # Print out possible kid matches
    try:
        for index, row in kids.iterrows():
            plakey=int(kids.loc[index]['Plakey'])
            first=kids.loc[index]['First']
            last=kids.loc[index]['Last']
            mytext='Possible match: '+str(plakey)+' '+ first+' '+ last
            a=tk.Label(root, text=mytext)
            a.grid(row=rownum)
            rownum+=1
    except:
        pass
    # now add new player button, entry box for ID, entry box for ID and add alias
    def addkid(event):
        choice.set('addkid')
        root.destroy()
    def skip(event):
        choice.set('skip')        
        root.destroy()        
    def IDplayer(event):
        choice.set('ID')        
        root.destroy()  
    def alias(event):
        choice.set('alias')        
        root.destroy() 
    def abort(event):
        choice.set('abort')        
        root.destroy() 

    tk.Label(root, text='Enter existing player ID number').grid(row=rownum, column=0)
    # Entry box for player num
    tk.Entry(root, textvariable=thisplanum).grid(row=rownum, column=1)
    rownum+=1
    a=tk.Button(root, text='Add as new player')
    a.bind('<Button-1>', addkid)
    a.grid(row=rownum, column=0)
    a=tk.Button(root, text='Skip player')
    a.bind('<Button-1>', skip)
    a.grid(row=rownum, column=1)
    a=tk.Button(root, text='ID existing player')
    a.bind('<Button-1>', IDplayer)
    a.grid(row=rownum, column=2)
    a=tk.Button(root, text='ID and add alias')
    a.bind('<Button-1>', alias)
    a.grid(row=rownum, column=3)
    a=tk.Button(root, text='abort')
    a.bind('<Button-1>', abort)
    a.grid(row=rownum, column=4)

    root.mainloop()

    mychoice=choice.get()
    choices={} # dict for choice return (and possibly existing player ID number)
    if mychoice=='abort':
        print('Execution aborted')
    elif mychoice=='ID' or mychoice=='alias':
        # ensure entered # is in list
        if int(thisplanum.get()) not in kids.Plakey.unique():
            print('Invalid player number entered.')
            mychoice='skip'
        if mychoice=='ID':
            choices.update({'ID':int(thisplanum.get())})# Need to return correct chosen player key
        elif mychoice=='alias':
            choices.update({'ID':thisplanum.get()})# return correct chosen player key 
    # returned choices are addkid, skip, 
    choices.update({'choice':mychoice})    
    return choices
#%%
def newplafamtk(Ser, phonelist):
    ''' Confirm that player (and family) are new via tkinter and add to players/famcontact lists 
    pass famkey, phonelist (for convenience); no need for distinction between add player and 
    add player & add family... both handled by addplafam '''
    
    root = tk.Tk()
    choice=tk.StringVar() # must be define outside of event called functions
    def addtodb(event):
        choice.set('adddb')
        root.destroy()
    def skip(event):
        choice.set('skip')        
        root.destroy()
    famkey=Ser.Famkey # nan or zero if no family match
    rownum=0
    if famkey>0: # family matched but not player (already checked aliases, other possible errors)
        mytext='No match for player '+Ser.First+' '+Ser.Last+' in Famkey '+str(int(famkey))
        a=tk.Label(root, text=mytext)
        a.grid(row=rownum)
        rownum+=1
    else:
        mytext='New family... no match for '+','.join(phonelist)
        a=tk.Label(root, text=mytext)
        a.grid(row=rownum)
        rownum+=1
        mytext='New player '+ Ser.First + ' '+Ser.Last + ' Parent: '+str(Ser.Pfirst1) + ' '+ str(Ser.Plast1)
        b=tk.Label(root, text=mytext)
        b.grid(row=rownum)
        rownum+=1
    c=tk.Button(root, text='Add new player (and family) to database')
    c.bind('<Button-1>', addtodb)
    c.grid(row=rownum)
    rownum+=1
    d=tk.Button(root, text='Skip player')
    d.bind('<Button-1>', skip)
    d.grid(row=rownum)
    root.mainloop()
    mychoice=choice.get()    
    return mychoice


def addalias(players, Ser):
    ''' Add alias to existing player (if chosen from tk player GUI) '''
    plakey=Ser.Plakey
    first=Ser.First # new alias name to add
    match=players[players['Plakey']==plakey]
    if len(match)!=1:
        last=Ser.Last
        print('Problem adding alias for', first, last)
        return players
    alias=players.loc[match.index[0]]['Alias']
    if str(alias)=='nan':
        players.loc[match.index[0],'Alias'] = first
    else:
        newalias=alias+', '+first # just make comma separated string
        players.loc[match.index[0],'Alias'] = newalias
    # direct save of modified
    players.to_csv(cnf._INPUT_DIR + '\\players.csv', index=False)
    return players

def comparefamkeys(players,famcontact):
    '''Utility script to compare family contacts and players list '''
    fams_pla=players.Famkey.unique()
    fams_pla=np.ndarray.tolist(fams_pla)
    fams_con=famcontact.Famkey.unique()
    fams_con=np.ndarray.tolist(fams_con)
    # compare contacts and billing ... should be identical
    
    noplayers=[i for i in fams_con if i not in fams_pla]
    for i,val in enumerate(noplayers):
        print("Famkey ", val, " in family contacts but not found among players.")

    # Check for family name discrepancies between players and famcontact
    for i in range(0,len(famcontact)):
        famkey=famcontact.iloc[i]['Famkey'] # grab this key
        family=famcontact.iloc[i]['Family']
        family=family.title()   # switch to title case
        family=family.strip() # remove whitespace
        match=players[players['Famkey']==famkey]
        if len(match)==1: # should be a match already assuming above section finds no discrepancies
            family2=match.iloc[0]['Family']
            family2=family2.title()   # switch to title case
            family2=family2.strip() # remove whitespace
            if family!=family2: # different family names with same key
                print("Family key ", str(famkey), ": ", family, " or ", family2)
    return

def calcage(Ser):
    '''pass Series with DOB as timestamp and return Age column in years as floats
    return column containing age in years (e.g. 6.1 yrs)'''  
    mytime=datetime.now() 
    mytime=datetime.date(mytime) # convert time to datetime.date
    Ser=pd.to_datetime(Ser)
    # Get age in years
    Age=mytime-Ser  # age in days (timedelta)
    Age= Age.dt.total_seconds() / (24 * 60 * 60)/365.25
    return Age

def parseOls(fname):
    ''' Custom parser for goofy Ols spreadsheet containing junior basketball schedule'''
    Ols=pd.read_excel(fname, encoding='cp437')
    Ols=Ols.iloc[:, 0:7]
    mycols=['Date','Time','Junk','Team1','Team2', 'Team3','Team4']
    Ols.columns=mycols
    mycols=['Date','Time','Team1','Team2', 'Team3','Team4']
    Ols=Ols[mycols]
    
    Ols=Ols.loc[pd.notnull(Ols['Time'])]
    Ols=Ols[Ols['Time'].str.contains(':')]
    Ols=Ols.reset_index(drop=True)
    
    gooddate=Ols['Date'].apply(parseDate).dropna()
    datelist=gooddate.tolist() # correct list of dates
    starts=Ols.loc[Ols['Date'].str.contains('WEEK', na=False)] # correct parsing of dates
    starts=np.ndarray.tolist(starts.index.unique())
    starts.append(len(Ols))
    for i in range(0,len(Ols)): # correct dates column 
        # find positions of first val larger than i in ordered starts list
        pos=[v for v in starts if v > i] # first value larger than i
        position=starts.index(min(pos))-1 # corresponds to index of date value from list to assign
        Ols.loc[i,'Date'] = datelist[position]
    # now duplicate rows with 
    Ols2=Ols.copy()
    mycols=['Date','Time','Team3','Team4']
    Ols2=Ols2[['Date','Time','Team3','Team4']]
    Ols2.columns=['Date','Time','Team1','Team2']
    Ols['Location']='Court 1'
    Ols2['Location']='Court 2'
    Ols=Ols[['Date','Time','Team1','Team2','Location']]
    Ols=pd.concat([Ols,Ols2], ignore_index=True)
    Ols.columns=['Date','Time','Home','Visitor','Location']
    return Ols

def updategradeadjust(row, players, year):
    ''' From row in signup file (with correct current grade) run gradeadj to
    see if changes need to be made 
    called by processdatachanges after merge of signups and players  
    args:
        row - single entry from gsignups file (w/ parent entered grade)
        players- DB-like file w/ player DOB and other info
        year - Cabrini sports year (i.e. 2019 for 2019-20 school year)
    '''
    now=datetime.now()
    gradeadj=row.Gradeadj
    if str(row.DOB)=='NaT' or row.Grade_n=='nan': # skip players with no DOB on file
        return players
    # Already checked for DOB discrepancy betwee SC signup and players.csv
    try:
        dob=datetime.date(row.DOB)
    except: # could be already converted to datetime.date
        pass 
    if row.Grade_n=='K': # use newly entered grade from signups (not existing from players)
        grade=0
    else:
        grade=row.Grade_n
    tempyear=now.year-int(grade)
    entryage=datetime.date(tempyear,8,1)-dob 
    entryage = (entryage.days + entryage.seconds/86400)/365.2425
    # Consider all the separate cases for entry age
    if 5 < entryage <6: # normal K age
        newadj=0
    elif 4 < entryage <5: # ahead of schedule
        newadj=1
    elif 6 < entryage <7: # 1 year back
        newadj=-1       
    elif 7 < entryage <8: # working on grade school mustache
        newadj=-2
    else: # probably some entry error
        print('Suspected DOB or grade error for ', row.First, ' ', row.Last,' Grade ', row.Grade_n, 'DOB', datetime.date.strftime(dob, "%m/%d/%y") )
        return players
    if gradeadj!=newadj:
        match=players[players['Plakey']==row.Plakey]
        if len(match)==1:
            thisind=match.index[0]
            # update player grade (no approval)
            players.loc[thisind,'Gradeadj'] = newadj # set to new value from current signup file
            print('Grade adjustment changed to', str(newadj),' for ',row.First, ' ', row.Last)
    return players

#%% Legacy or one-time use functions


# LEGACY FUNCTIONS
# assignteams probably not needed (using assigntoteams and different flow of information)

def assignteams(df, Teams):
    '''Pass contacts summary and assign team name (similar to assigntoteams called by mastersignups
    Teams tab must have those for current year; merge based on grade, gender and sport from teams tab 
    (which only contains names from this season/year to avoid screwing up old custom team assignments''' 
    Teamsmult, Twoteams =makemultiteam(Teams) # makes duplicates team entries to match both grades
    Teamsmult['Grade']=Teamsmult['Grade'].astype('str') # convert grade back to string
    Teamsmult.Grade=Teamsmult.Grade.replace('K','0', regex=True) # convert Ks to grade 0
    df.loc[:,'Grade']=df['Grade'].astype('str') # these summaries already have K as grade 0
    # left merge keeps all master_signups entries
    df=pd.merge(df, Teamsmult, how='left', on=['Grade','Gender','Sport'], suffixes=('','_r'))
    # now copy over CYC team name from Teams_coaches to this df ... skip copying if null
    for i in range(0, len(df)):
        if df.iloc[i]['Team']!='nan':
            df.loc[i, 'Team'] = df.iloc[i]['Team']
    # now drop extra columns
    mycols=['First', 'Last', 'Grade', 'Gender', 'School', 'Phone', 'Text', 'Email', 'Phone2', 'Text2', 'Email2', 'Team', 'Plakey','Famkey', 'Family'] 
    dropcollist=[s for s in df.dtypes.index if s not in mycols]
    df=df.drop(dropcollist, axis=1) # drops extraneous columns    
    return df
